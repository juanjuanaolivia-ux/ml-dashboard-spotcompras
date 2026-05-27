#!/usr/bin/env python3
import json, os, math, unicodedata, re
from datetime import datetime, date

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
OUT_DIR  = os.path.join(BASE_DIR, 'dashboards')
os.makedirs(OUT_DIR, exist_ok=True)

def load_json(name):
    p = os.path.join(DATA_DIR, name)
    if not os.path.exists(p): return None
    try:
        with open(p, encoding='utf-8') as fh: return json.load(fh)
    except json.JSONDecodeError:
        print(f"  ⚠️ JSON malformado/truncado: {name} — usando vacío")
        return None

def jd(v): return json.dumps(v, ensure_ascii=False)

def fmt_money(v):
    v = float(v or 0)
    if v >= 1_000_000: return f"${v/1_000_000:.2f}M"
    if v >= 1_000: return f"${v/1_000:.0f}K"
    return f"${v:,.0f}"

def delta_pct(cur, pri):
    c, p = float(cur or 0), float(pri or 0)
    if not p: return None
    return ((c - p) / p) * 100

def arrow(pct, inv=False):
    if pct is None: return "—", "neutral"
    if pct > 0: return f"▲ {pct:.1f}%", ("bad" if inv else "good")
    if pct < 0: return f"▼ {abs(pct):.1f}%", ("good" if inv else "bad")
    return "= 0%", "neutral"

# Title-based keyword → STATUS internal category (order matters: more specific first)
TITLE_RULES = [
    # COOK — real COOK items: espumadores de leche, pavas eléctricas
    (['ESPUMADOR DE LECHE', 'PAVA ELECTRICA', 'PAVA TIVOLI', 'PAVA DIGITAL'], 'PEQUEÑOS ELECTRO'),
    # PEQUEÑO ELECTRO
    (['BATIDORA', 'AMASADORA', 'STAND MIXER'], 'PEQUEÑOS ELECTRO'),
    (['FREIDORA', 'AIR FRYER', 'HORNO FREIDOR'], 'PEQUEÑOS ELECTRO'),
    (['CAFETERA'], 'PEQUEÑOS ELECTRO'),
    (['MICROONDAS'], 'PEQUEÑOS ELECTRO'),
    (['ASPIRADORA', 'ASPIRADOR', 'LAVATAPIZADO'], 'PEQUEÑOS ELECTRO'),
    (['LICUADORA'], 'PEQUEÑOS ELECTRO'),
    (['TOSTADORA'], 'PEQUEÑOS ELECTRO'),
    (['WAFLERA', 'GOFRERA'], 'PEQUEÑOS ELECTRO'),
    (['YOGURTERA'], 'PEQUEÑOS ELECTRO'),
    (['EXPRIMIDOR'], 'PEQUEÑOS ELECTRO'),
    (['MULTIRALLADOR', 'RALLADOR'], 'PEQUEÑOS ELECTRO'),
    (['PROCESADORA', 'PROCESADOR ALIM'], 'PEQUEÑOS ELECTRO'),
    (['JUGUERA'], 'PEQUEÑOS ELECTRO'),
    (['SANDWICHERA', 'SANWICHERA'], 'PEQUEÑOS ELECTRO'),
    (['PARRILLA', 'PLANCHA GRILL'], 'PEQUEÑOS ELECTRO'),
    (['FABRICA DE MINI', 'FABRICA DE DONA', 'DONUTS', 'DONA MAKER'], 'PEQUEÑOS ELECTRO'),
    # AUDIO
    (['PARTY BOX', 'PARLANTE', 'ALTAVOZ'], 'AUDIO'),
    (['AURICULAR', 'VINCHA', 'HEADPHONE', 'EARPHONE'], 'AUDIO'),
    (['TRIPODE PARA PARLANTE', 'TRIPODE PARLANTE'], 'AUDIO'),
    (['RADIO AM', 'RADIO FM'], 'AUDIO'),
    # LUCES
    (['REFLECTOR'], 'ILUMINACION'),
    (['PLAFON', 'PANEL LED', 'PANEL 595', 'DOWNLIGHT'], 'ILUMINACION'),
    (['LUZ DE EMERGENCIA', 'LUZ EMERGENCIA'], 'ILUMINACION'),
    (['DRIVER', 'FUENTE 12W', 'FUENTE 6W', 'FUENTE 18W', 'FUENTE 24W'], 'ILUMINACION'),
    (['LISTON BAJO MESADA'], 'ILUMINACION'),
    # HERRAMIENTAS
    (['TALADRO'], 'HERRAMIENTAS'),
    (['AMOLADORA'], 'HERRAMIENTAS'),
    (['HIDROLAVADORA'], 'HERRAMIENTAS'),
    (['DESTORNILLADOR', 'ATORNILLADOR'], 'HERRAMIENTAS'),
    (['LIJADORA'], 'HERRAMIENTAS'),
    # ACCESORIOS BICICLETA
    (['CICLOCOMPUTADORA', 'CICLOCOMPUTADOR'], 'CICLISMO'),
    (['COMPRESOR PORTATIL', 'COMPRESOR BICICLETA', 'COMPRESOR CUBO'], 'CICLISMO'),
    (['CYCPLUS', 'ANT+ CICLISMO', 'RECEPTOR USB ANT'], 'CICLISMO'),
    # BIENESTAR
    (['MASAJEADOR', 'PISTOLA MASAJEAD'], 'BIENESTAR'),
    (['EJERCITADOR PARA ABDOMIN', 'EJERCITADOR ABDOMIN', 'ABS-250', 'RUEDA ABDOMINAL'], 'BIENESTAR'),
    (['EJERCITADOR DE PISO', 'EJERCITADOR PISO', 'EJERCITADOR PARA PISO'], 'BIENESTAR'),
    # SALUD → BIENESTAR
    (['BALANZA', 'BASCULA'], 'BIENESTAR'),
    # CCTV
    (['CAMARA DE SEGURIDAD', 'CAMARA INDOOR', 'CAMARA OUTDOOR', 'DOBLE CAMARA'], 'CCTV'),
    # GAMING
    (['SILLA GAMER', 'SILLA GAMING'], 'GAMING'),
    # JUGUETES
    (['PISTOLA DE AGUA', 'PISTOLA AGUA', 'BURBUJERO', 'PISTOLA BURBUJAS'], 'JUGUETES'),
    (['MALETIN DE ARTE', 'SET DE ARTE', 'KIT DE ARTE', 'BLOQUES MAGNETICO', 'JUEGO MAGNETICO'], 'JUGUETES'),
    (['CONTROL REMOTO', 'BUGGY', 'AUTO A CONTROL', 'CAMIONETA A CONTROL'], 'JUGUETES'),
    # TEMPORADA
    (['VENTILADOR'], 'CLIMATIZACION'),
    # ARTICULOS PARA OFICINA
    (['CONTADOR DE BILLETE', 'CONTADORA DE BILLETE', 'CONTADORA BILLETE'], 'UTILES'),
    # ARTICULOS PARA EL HOGAR
    (['MAQUINA DE COSER', 'MÁQUINA DE COSER'], 'HOGAR'),
    # PET
    (['CORTAPELO', 'SECADORA PARA MASCOTA', 'CORTAPELO ASPIRADORA'], 'PET'),
    # PC - INFORMATICA
    (['BASE PARA NOTEBOOK', 'BASE NOTEBOOK'], 'VARIOS'),
    (['SOPORTE PARA MONITOR', 'SOPORTE MONITOR', 'BASE MONITOR', 'MONITOR SOPORTE'], 'VARIOS'),
]

def _norm(s):
    """Uppercase + strip accents for accent-insensitive matching."""
    s = (s or '').upper()
    return unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode('ascii')

def classify_title(title):
    """Classify ML listing title into STATUS internal category."""
    t = _norm(title)
    for keywords, cat in TITLE_RULES:
        if any(k in t for k in keywords):
            return cat
    return 'VARIOS'

# ML category ID → STATUS internal category (used for PRIOR period only)
ML_CAT_TO_INTERNAL = {
    'MLA10085': 'PEQUEÑOS ELECTRO',   # Batidoras planetarias
    'MLA456045': 'PEQUEÑOS ELECTRO',  # Freidoras de aire
    'MLA416847': 'PEQUEÑOS ELECTRO',  # Fábricas donuts
    'MLA1577':   'PEQUEÑOS ELECTRO',  # Microondas
    'MLA4340':   'PEQUEÑOS ELECTRO',  # Cafeteras de filtro
    'MLA4622':   'HOGAR',             # Máquinas de coser
    'MLA4339':   'PEQUEÑOS ELECTRO',  # Procesadoras / ralladores
    'MLA8618':   'AUDIO',             # Parlantes portátiles
    'MLA370796': 'CICLISMO',          # Compresores Cycplus
    'MLA4337':   'PEQUEÑOS ELECTRO',  # Aspiradoras de mano
    'MLA126142': 'PEQUEÑOS ELECTRO',  # Yogurteras
    'MLA1588':   'ILUMINACION',       # Plafones LED
    'MLA447782': 'GAMING',            # Sillas gamer
    'MLA409403': 'PEQUEÑOS ELECTRO',  # Wafleras
    'MLA74279':  'PEQUEÑOS ELECTRO',  # Exprimidores eléctricos
    'MLA5232':   'HERRAMIENTAS',      # Taladros inalámbricos
    'MLA10104':  'HERRAMIENTAS',      # Lijadoras
    'MLA401457': 'PEQUEÑOS ELECTRO',  # Aspiradoras robot
    'MLA104680': 'PEQUEÑOS ELECTRO',  # Licuadoras personales
    'MLA370486': 'PET',               # Aspiradoras mascotas
    'MLA373611': 'PEQUEÑOS ELECTRO',  # Parrillas / sandwicheras
    'MLA417723': 'BIENESTAR',         # Masajeadores
    'MLA371883': 'CICLISMO',          # GPS Cycplus
    'MLA30840':  'HERRAMIENTAS',      # Hidrolavadoras
    'MLA388313': 'JUGUETES',          # Burbujeros
    'MLA414164': 'JUGUETES',          # Kits de arte
    'MLA9554':   'BIENESTAR',         # Balanzas digitales
    'MLA126135': 'PEQUEÑOS ELECTRO',  # Sandwicheras
    'MLA86840':  'PEQUEÑOS ELECTRO',  # Aspiradoras tapizados
    'MLA3697':   'AUDIO',             # Auriculares
    'MLA10068':  'PEQUEÑOS ELECTRO',  # Tostadoras
    'MLA30803':  'HERRAMIENTAS',      # Destornilladores
    'MLA407557': 'UTILES',            # Contadoras de billetes
    'MLA10569':  'AUDIO',             # Trípodes parlante
    'MLA74278':  'PEQUEÑOS ELECTRO',  # Jugueras
    'MLA5229':   'HERRAMIENTAS',      # Amoladoras
    'MLA456969': 'HERRAMIENTAS',      # Taladros percutores
    'MLA418043': 'VARIOS',            # Soportes monitor doble
    'MLA373504': 'ILUMINACION',       # Reflectores LED
    'MLA5848':   'ILUMINACION',       # Drivers LED
    'MLA417835': 'CCTV',              # Cámaras seguridad
    'MLA455865': 'CLIMATIZACION',     # Ventiladores portátiles
    'MLA438177': 'BIENESTAR',         # Ejercitadores abdominales
    'MLA9760':   'CICLISMO',          # Receptor ANT+ Cycplus
    'MLA418042': 'VARIOS',            # Bases monitor
    'MLA90321':  'JUGUETES',          # Pistolas de agua
    'MLA417724': 'BIENESTAR',         # Masajeadores de mano
    'MLA417130': 'VARIOS',            # Limpialentes
    'MLA125102': 'ILUMINACION',       # Luces emergencia
    'MLA435473': 'PEQUEÑOS ELECTRO',  # Exprimidores manuales
    'MLA3098':   'BIENESTAR',         # Ejercitadores piso
    'MLA413405': 'PEQUEÑOS ELECTRO',  # Marcadores / sellos de carne
    'MLA417716': 'VARIOS',            # Aro de luz
}


def build_dashboard():
    mc    = load_json('metrics_current.json') or {}
    mp    = load_json('metrics_prior.json') or {}
    ec    = load_json('enrich_current.json') or {}
    ep    = load_json('enrich_prior.json') or {}
    sm    = load_json('summary.json') or {}
    rep   = load_json('reputation.json') or {}
    ship_perf  = load_json('shipment_performance.json') or {}
    # Historial diario de envios reales (join orders x shipments_log)
    import re as _re2, collections as _col2
    from datetime import date as _dt2, timedelta as _td2
    _today2 = _dt2.today()
    _mon2   = _today2 - _td2(days=_today2.weekday())
    _w0_s   = (_mon2 - _td2(weeks=3)).isoformat()
    _fp2 = ship_perf.get('flex', {})
    _cp2 = ship_perf.get('colecta', {})
    def _parse_abs2(s):
        m = _re2.search(r'(\d+)\s+de\s+(\d+)', str(s or ''))
        return (int(m.group(1)), int(m.group(2))) if m else (0, 0)
    _f_ok2, _f_tot2 = _parse_abs2(_fp2.get('envios_correctos_abs',''))
    _c_ok2, _c_tot2 = _parse_abs2(_cp2.get('envios_correctos_abs',''))
    _fx_fail = 1 - (_f_ok2/_f_tot2 if _f_tot2 else 1)  # tasa fallo semanal Flex
    _co_fail = 1 - (_c_ok2/_c_tot2 if _c_tot2 else 1)  # tasa fallo semanal Colecta
    # Join orders_current x shipments_log para counts reales por dia y tipo
    _ship_log2 = load_json('shipments_map.json') or {}  # {ship_id: logistic_type} — actualizado en run_daily 1.4b
    _daily_flex = _col2.defaultdict(int)
    _daily_col  = _col2.defaultdict(int)
    for _o2 in (load_json('orders_current.json') or []):
        _d2s = (_o2.get('date_created','') or '')[:10]
        if _d2s < _w0_s: continue
        _sid2 = str(_o2.get('shipping_id','') or '')
        _lt2  = _ship_log2.get(_sid2, '')
        if _lt2 == 'self_service':   _daily_flex[_d2s] += 1
        elif _lt2 == 'cross_docking': _daily_col[_d2s] += 1
    # Construir ship_daily con fail rate de la semana actual
    ship_daily = {}
    for _wi2 in range(4):
        _ws2 = _mon2 - _td2(weeks=3-_wi2)
        for _di2 in range(7):
            _day2 = (_ws2 + _td2(days=_di2)).isoformat()
            _f2   = _daily_flex.get(_day2, 0)
            _c2   = _daily_col.get(_day2, 0)
            _fb2  = round(_f2 * _fx_fail)
            _cb2  = round(_c2 * _co_fail)
            if _f2 + _c2 == 0: continue  # omitir dias sin datos
            ship_daily[_day2] = {
                'flex_ok': max(0, _f2 - _fb2), 'flex_bad': _fb2,
                'col_ok':  max(0, _c2 - _cb2), 'col_bad':  _cb2,
            }
    neg_revs   = load_json('negative_reviews.json') or {}
    ads        = load_json('ads_manual.json') or {}
    ads_daily_f= load_json('ads_daily.json') or {}
    ads_d      = ads_daily_f.get('daily', {})
    costs = load_json('costs_manual.json') or {}
    mon   = load_json('monthly_2026.json') or []
    stk   = load_json('stock_status.json') or []
    items_list = load_json('items.json') or []

    today   = date.today()
    max_day = today.day

    # ── FUENTE MAESTRA: MLA → SKU (Excel maestro) ───────────────────────────
    mla_to_sku   = {}   # item_id → sku_code (authoritative)
    mla_es_combo = {}   # item_id → units_per_combo
    cat_excel    = {}   # sku_code → {categoria, sub_cat, descripcion}

    _excel_path = os.path.join(OUT_DIR, 'Categoría y subcategoría por código.xlsx')
    if os.path.exists(_excel_path):
        try:
            import openpyxl
            _wb = openpyxl.load_workbook(_excel_path, data_only=True)
            # Hoja MLA-Código → mapping MLA → SKU
            if 'MLA-Código' in _wb.sheetnames:
                _ws = _wb['MLA-Código']
                for _r in range(2, _ws.max_row + 1):
                    _mla  = str(_ws.cell(_r, 1).value or '').strip()
                    _cod  = str(_ws.cell(_r, 2).value or '').strip()
                    _cmb  = str(_ws.cell(_r, 3).value or '').strip().upper()
                    _unit = _ws.cell(_r, 5).value
                    if _mla and _cod:
                        mla_to_sku[_mla]   = _cod
                        if _cmb == 'SI':
                            mla_es_combo[_mla] = int(_unit) if _unit else 1
            # Hoja Sheet1 → SKU → categoria/sub_cat
            if 'Sheet1' in _wb.sheetnames:
                _ws2 = _wb['Sheet1']
                for _r in range(2, _ws2.max_row + 1):
                    _cod  = str(_ws2.cell(_r, 1).value or '').strip()
                    _desc = str(_ws2.cell(_r, 2).value or '').strip()
                    _cat  = str(_ws2.cell(_r, 3).value or '').strip()
                    _sub  = str(_ws2.cell(_r, 4).value or '').strip()
                    if _cod:
                        cat_excel[_cod] = {'descripcion': _desc, 'categoria': _cat, 'sub_cat': _sub}
        except Exception as _e:
            print(f'  [warning] Excel maestro: {_e}')

    # También cargar STATUS.xlsx para stock físico
    stk_by_code = {}
    for s in stk:
        cod = (s.get('codigo') or '').upper().strip()
        if cod:
            stk_by_code[cod] = s

    # ── FUENTE AUTORITATIVA: cruce_mla_result.json (MLA ID → categoría top-level) ──
    _cruce_data = load_json('cruce_mla_result.json') or {}
    _cruce_rows = _cruce_data.get('rows', []) if isinstance(_cruce_data, dict) else []
    _CRUCE_NORM = {
        'De Mano': 'HERRAMIENTAS', 'Eléctricos': 'HERRAMIENTAS',
        'Amoladoras Angulares': 'HERRAMIENTAS', 'Destornilladores': 'HERRAMIENTAS',
        'Taladros': 'HERRAMIENTAS', 'Hidrolavadoras': 'HERRAMIENTAS',
        'Manuales': 'HERRAMIENTAS', 'De Aire': 'HERRAMIENTAS',
        'Cafeteras': 'PEQUEÑOS ELECTRO', 'Tostadoras': 'PEQUEÑOS ELECTRO',
        'Jugueras': 'PEQUEÑOS ELECTRO', 'Licuadoras': 'PEQUEÑOS ELECTRO',
        'Donuts Maker': 'PEQUEÑOS ELECTRO', 'Aspiradoras Robot': 'PEQUEÑOS ELECTRO',
        'Aspiradoras': 'PEQUEÑOS ELECTRO', 'Exprimidores Eléctricos': 'PEQUEÑOS ELECTRO',
        'Exprimidores Manuales': 'PEQUEÑOS ELECTRO', 'Sandwicheras': 'PEQUEÑOS ELECTRO',
        'Microondas': 'PEQUEÑOS ELECTRO', 'Kits Parrilleros': 'PEQUEÑOS ELECTRO',
        'Parlantes Portátiles': 'AUDIO', 'Auriculares': 'AUDIO',
        'Soportes para Parlantes': 'AUDIO',
        'Iluminadores': 'ILUMINACION',
        'Kits de Seguridad': 'CCTV',
        'Ventiladores Portátiles': 'CLIMATIZACION',
        'Contadoras de Billetes': 'UTILES',
        'Máquinas de Coser': 'HOGAR',
        'Balanzas de Baño': 'BIENESTAR', 'Aparatos para Abdominales': 'BIENESTAR',
        'Secadores': 'PET',
        'Pistolas de Agua': 'JUGUETES', 'Vehículos a Control Remoto': 'JUGUETES',
        'Soportes': 'VARIOS', 'Flash para Celulares': 'VARIOS',
        'ARTICULOS PARA EL HOGAR': 'HOGAR', 'ARTICULOS PARA OFICINA': 'UTILES',
        'SALUD': 'BIENESTAR', 'TEMPORADA': 'CLIMATIZACION',
        'ACCESORIOS BICICLETA': 'CICLISMO', 'PC - INFORMATICA': 'VARIOS',
    }
    cruce_mla_cat = {}  # mla_id → categoría normalizada top-level
    for _cr in _cruce_rows:
        _cid = _cr.get('mla_id', '')
        _ccat = _cr.get('categoria', '')
        if _cid and _ccat:
            cruce_mla_cat[_cid] = _CRUCE_NORM.get(_ccat, _ccat)


    # ── Fallback matching (solo para items NO en Excel maestro) ──────────────
    _STOPW = {'DE','LA','EL','EN','CON','Y','A','E','O','POR','UN','UNA','LOS',
              'LAS','DEL','AL','SU','SE','LO','X','V','I','II','III','SUS'}
    _BRND  = {'TIVOLI','CYCPLUS','AIWA','ICSEE','ENERGY','SAFE','SUZUKI'}
    _COLR  = {'NEGRO','NEGRA','GRIS','BLANCO','BLANCA','AZUL','ROJO','ROJA',
              'ROSA','NARANJA','VIOLETA','AMARILLO','AMARILLA','BEIGE','INOX','VERDE','COLOR'}
    _GENR  = {'ELECTRICA','ELECTRICO','AUTOMATICO','AUTOMATICA','MANUAL','DIGITAL',
              'PORTATIL','MECANICA','MECANICO','INDUSTRIAL','PERSONAL','PREMIUM','PROFESIONAL'}

    def _norm_s(s):
        s = unicodedata.normalize('NFD', str(s).upper())
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return re.sub(r'[^A-Z0-9 ]', ' ', s)

    def _mw(s):
        exc = _STOPW | _BRND | _COLR
        return {w for w in _norm_s(s).split()
                if w and w not in exc and not w.isdigit() and len(w) > 1}

    _sku_codes = [s.get('codigo','') for s in stk if s.get('codigo')]
    _sku_idx   = [(s.get('codigo',''), _mw(s.get('descripcion',''))) for s in stk if s.get('codigo')]

    def _match_fallback(title):
        tu = _norm_s(title)
        for cod in sorted(_sku_codes, key=lambda c: -len(c)):
            if re.search(r'(?<![A-Z0-9])' + re.escape(_norm_s(cod)) + r'(?![A-Z0-9])', tu):
                return cod
        tw = _mw(title)
        if not tw: return None
        best_s, best_c = 0.0, None
        for cod, sw in _sku_idx:
            if not sw: continue
            ov = tw & sw
            if not ov or not (ov - _GENR): continue
            sc = len(ov) / len(sw)
            if sc > best_s: best_s, best_c = sc, cod
        return best_c if best_s >= 0.5 else None

    def _get_sku(item_id, title):
        """Excel maestro primero; fallback semántico si no está mapeado."""
        if item_id in mla_to_sku:
            return mla_to_sku[item_id], 'master'
        fb = _match_fallback(title)
        return (fb, 'fallback') if fb else (item_id, 'unmapped')

    def _get_cat(sku_code, title_fallback):
        """Categoría: cruce_mla_result > Excel maestro > STATUS > TITLE_RULES."""
        CAT_REMAP = {'COOK': 'PEQUEÑOS ELECTRO', 'PEQUEÑO ELECTRO': 'PEQUEÑOS ELECTRO', 'LUCES': 'ILUMINACION',
                     'ACCESORIOS BICICLETA': 'CICLISMO', 'TEMPORADA': 'CLIMATIZACION',
                     'ARTICULOS PARA EL HOGAR': 'HOGAR', 'ARTICULOS PARA OFICINA': 'UTILES',
                     'SALUD': 'BIENESTAR', 'PC - INFORMATICA': 'VARIOS'}
        if sku_code in cruce_mla_cat:
            return cruce_mla_cat[sku_code], ''
        if sku_code in cat_excel:
            cat = cat_excel[sku_code].get('categoria','') or ''
            return CAT_REMAP.get(cat, cat), cat_excel[sku_code].get('sub_cat','')
        if sku_code in stk_by_code:
            cat = stk_by_code[sku_code].get('categoria','') or ''
            return CAT_REMAP.get(cat, cat), stk_by_code[sku_code].get('sub_cat','')
        return classify_title(title_fallback), ''

    # ── SINGLE SOURCE OF TRUTH: computar TODOS los totales desde orders ──────
    orders_cur      = load_json('orders_current.json') or []
    orders_pri_data = load_json('orders_prior.json') or []

    # Computar daily_stats directamente desde orders (para reconciliación exacta)
    daily_cur_computed = {}
    daily_pri_computed = {}

    # Current period
    units_by_item = {}   # item_id → {units, gmv, title}
    dow_units_raw = [0.0] * 7

    for o in orders_cur:
        ds  = o.get('date_closed') or o.get('date_created') or ''
        st  = o.get('status', '')
        try:
            dobj = datetime.fromisoformat(ds[:19])
            day  = dobj.day
            dow  = dobj.weekday()
        except Exception:
            day, dow = 1, 0

        if day not in daily_cur_computed:
            daily_cur_computed[day] = {'gmv': 0.0, 'units': 0, 'paid': 0, 'canc': 0}

        if st == 'cancelled':
            daily_cur_computed[day]['canc'] += 1   # FIX: 1 por orden, no por unidades
        else:
            daily_cur_computed[day]['paid'] += 1    # FIX: 1 por orden, fuera del loop items
            for it in o.get('items', []):
                iid   = it.get('item_id', '?')
                qty   = it.get('quantity', 0) or 0
                price = float(it.get('unit_price', 0) or it.get('full_unit_price', 0) or 0)
                title = it.get('title', '')
                daily_cur_computed[day]['gmv']   += qty * price
                daily_cur_computed[day]['units']  += qty
                dow_units_raw[dow] += qty
                if iid not in units_by_item:
                    units_by_item[iid] = {'units': 0, 'gmv': 0.0, 'title': title}
                units_by_item[iid]['units'] += qty
                units_by_item[iid]['gmv']   += qty * price

    # Prior period daily
    for o in orders_pri_data:
        ds  = o.get('date_closed') or o.get('date_created') or ''
        st  = o.get('status', '')
        try:
            day = int(ds[8:10])
        except Exception:
            day = 1
        if day not in daily_pri_computed:
            daily_pri_computed[day] = {'gmv': 0.0, 'units': 0, 'paid': 0, 'canc': 0}
        if st == 'cancelled':
            daily_pri_computed[day]['canc'] += 1   # FIX: 1 por orden
        else:
            daily_pri_computed[day]['paid'] += 1    # FIX: 1 por orden, fuera del loop items
            for it in o.get('items', []):
                qty   = it.get('quantity', 0) or 0
                price = float(it.get('unit_price', 0) or 0)
                daily_pri_computed[day]['gmv']   += qty * price
                daily_pri_computed[day]['units']  += qty

    # Stringify keys for JS
    daily_cur = {str(k): {'gmv': round(v['gmv']), 'units': v['units'],
                           'paid': v['paid'], 'canc': v['canc']}
                 for k, v in daily_cur_computed.items()}
    daily_pri = {str(k): {'gmv': round(v['gmv']), 'units': v['units'],
                           'paid': v['paid'], 'canc': v['canc']}
                 for k, v in daily_pri_computed.items()}

    # Totales derivados de orders (fuente única)
    gmv_c  = round(sum(v['gmv'] for v in daily_cur_computed.values()))
    gmv_p  = round(sum(v['gmv'] for v in daily_pri_computed.values()))
    un_c   = sum(v['units'] for v in daily_cur_computed.values())
    un_p   = sum(v['units'] for v in daily_pri_computed.values())
    pd_c   = sum(v['paid']  for v in daily_cur_computed.values())
    pd_p   = sum(v['paid']  for v in daily_pri_computed.values())
    canc_c = sum(v['canc']  for v in daily_cur_computed.values())
    canc_p = sum(v['canc']  for v in daily_pri_computed.values())
    tk_c   = round(gmv_c / un_c) if un_c else 0
    # ── SYNC: parchear mes actual en MONTHLY con totales de órdenes días 1..today.day (misma fuente que tab0) ──
    _td = today.day
    _gmv_sync   = round(sum(v['gmv']   for k,v in daily_cur_computed.items() if 1 <= k <= _td))
    _un_sync    = sum(v['units'] for k,v in daily_cur_computed.items() if 1 <= k <= _td)
    _pd_sync    = sum(v['paid']  for k,v in daily_cur_computed.items() if 1 <= k <= _td)
    _ca_sync    = sum(v['canc']  for k,v in daily_cur_computed.items() if 1 <= k <= _td)
    _tk_sync    = round(_gmv_sync / _pd_sync) if _pd_sync else 0
    _cr_sync    = round(_ca_sync / (_pd_sync + _ca_sync) * 100, 2) if (_pd_sync + _ca_sync) else 0
    for _m in mon:
        if _m.get('month') == today.month and _m.get('year') == today.year:
            _m['gmv']         = _gmv_sync
            _m['units']       = _un_sync
            _m['paid']        = _pd_sync
            _m['cancelled']   = _ca_sync
            _m['cancel_rate'] = _cr_sync
            _m['fees']        = round(mc.get('fees', _m.get('fees', 0)) or 0)
            _m['avg_ticket']  = _tk_sync
            break
    tk_p   = round(gmv_p / un_p) if un_p else 0
    cr_c   = round(canc_c / (pd_c + canc_c) * 100, 2) if (pd_c + canc_c) else 0
    cr_p   = round(canc_p / (pd_p + canc_p) * 100, 2) if (pd_p + canc_p) else 0
    fees_c     = mc.get('fees', 0) or 0
    fee_rate_c = mc.get('fee_rate', 0) or 0
    fee_rate_p = mp.get('fee_rate', 0) or 0
    # Normalize: stored as percentage (19.25) → convert to decimal (0.1925)
    if fee_rate_c > 1: fee_rate_c = round(fee_rate_c / 100, 4)
    if fee_rate_p > 1: fee_rate_p = round(fee_rate_p / 100, 4)
    unique_days = mc.get('unique_days', 1) or 1

    # ── SKU-CONSOLIDATED aggregation (MASTER MAPPING FIRST) ──────────────────
    sku_agg_cur = {}   # sku_key → {id, title, categoria, sub_cat, units, gmv}
    item_to_sku = {}   # item_id → sku_key

    for iid, d in units_by_item.items():
        s_key, method = _get_sku(iid, d['title'])
        item_to_sku[iid] = s_key
        cat, sub_cat = _get_cat(s_key, d['title'])
        if s_key not in sku_agg_cur:
            sku_agg_cur[s_key] = {
                'id':        s_key,
                'title':     d['title'],
                'categoria': cat,
                'sub_cat':   sub_cat,
                'units':     0,
                'gmv':       0.0,
                'method':    method,
            }
        sku_agg_cur[s_key]['units'] += d['units']
        sku_agg_cur[s_key]['gmv']   += d['gmv']

    sku_top_items = sorted(sku_agg_cur.values(), key=lambda x: -x['gmv'])
    for x in sku_top_items:
        x['gmv'] = round(x['gmv'])

    # ── Prior period SKU aggregation ──────────────────────────────────────────
    _sku_pri = {}
    for _o in orders_pri_data:
        if _o.get('status') == 'cancelled': continue
        for _it in _o.get('items', []):
            _iid  = _it.get('item_id', '?')
            _sku, _ = _get_sku(_iid, _it.get('title',''))
            _qty  = _it.get('quantity', 0) or 0
            _g    = (_it.get('unit_price', 0) or 0) * _qty
            if _sku not in _sku_pri:
                _sku_pri[_sku] = {'gmv': 0.0, 'units': 0}
            _sku_pri[_sku]['gmv']   += _g
            _sku_pri[_sku]['units'] += _qty
    sku_top_pri = {k: {'gmv': round(v['gmv']), 'units': v['units']} for k, v in _sku_pri.items()}

    # ── Daily category breakdown (for Ventas range filter) ───────────────────
    # Build MLA→category lookup from sku_agg_cur + item_to_sku
    _iid_to_cat = {}
    for _iid_k, _sku_k in item_to_sku.items():
        _iid_to_cat[_iid_k] = sku_agg_cur.get(_sku_k, {}).get('categoria', 'OTROS')

    _daily_by_cat = {}   # day → cat → {gmv, units}
    for _o in orders_cur:
        _ds  = _o.get('date_closed') or _o.get('date_created') or ''
        _st  = _o.get('status', '')
        if _st == 'cancelled': continue
        try:
            _day = int(_ds[8:10])
        except Exception:
            _day = 0
        if not _day: continue
        for _it in _o.get('items', []):
            _iid  = _it.get('item_id','')
            _qty  = _it.get('quantity', 0) or 0
            _price = float(_it.get('unit_price', 0) or 0)
            _cat  = _iid_to_cat.get(_iid, 'OTROS')
            if _day not in _daily_by_cat:
                _daily_by_cat[_day] = {}
            if _cat not in _daily_by_cat[_day]:
                _daily_by_cat[_day][_cat] = {'gmv': 0.0, 'units': 0}
            _daily_by_cat[_day][_cat]['gmv']   += _qty * _price
            _daily_by_cat[_day][_cat]['units'] += _qty

    # Prior period daily by cat
    _daily_by_cat_pri = {}
    for _o in orders_pri_data:
        _ds  = _o.get('date_closed') or _o.get('date_created') or ''
        _st  = _o.get('status', '')
        if _st == 'cancelled': continue
        try:
            _day = int(_ds[8:10])
        except Exception:
            _day = 0
        if not _day: continue
        for _it in _o.get('items', []):
            _iid   = _it.get('item_id','')
            _qty   = _it.get('quantity', 0) or 0
            _price = float(_it.get('unit_price', 0) or 0)
            _sku_k, _ = _get_sku(_iid, _it.get('title',''))
            _cat2, _  = _get_cat(_sku_k, _it.get('title',''))
            _cat2 = _cat2 or 'OTROS'
            if _day not in _daily_by_cat_pri:
                _daily_by_cat_pri[_day] = {}
            if _cat2 not in _daily_by_cat_pri[_day]:
                _daily_by_cat_pri[_day][_cat2] = {'gmv': 0.0, 'units': 0}
            _daily_by_cat_pri[_day][_cat2]['gmv']   += _qty * _price
            _daily_by_cat_pri[_day][_cat2]['units'] += _qty

    daily_by_cat = {str(d): {c: {'gmv': round(v['gmv']), 'units': v['units']}
                              for c, v in cmap.items()}
                    for d, cmap in sorted(_daily_by_cat.items())}
    daily_by_cat_pri = {str(d): {c: {'gmv': round(v['gmv']), 'units': v['units']}
                                  for c, v in cmap.items()}
                        for d, cmap in sorted(_daily_by_cat_pri.items())}

    # ── Daily hour breakdown (for Horarios range filter) ─────────────────────
    _daily_hour = {}   # day → hour → {gmv, units}
    for _o in orders_cur:
        _ds  = _o.get('date_closed') or _o.get('date_created') or ''
        _st  = _o.get('status', '')
        if _st == 'cancelled': continue
        try:
            _day  = int(_ds[8:10])
            _hraw = int(_ds[11:13])
            _hour = _hraw  # keep server hour
        except Exception:
            continue
        if not _day: continue
        if _day not in _daily_hour: _daily_hour[_day] = {}
        if _hour not in _daily_hour[_day]: _daily_hour[_day][_hour] = {'gmv': 0.0, 'units': 0}
        for _it in _o.get('items', []):
            _qty   = _it.get('quantity', 0) or 0
            _price = float(_it.get('unit_price', 0) or 0)
            _daily_hour[_day][_hour]['gmv']   += _qty * _price
            _daily_hour[_day][_hour]['units'] += _qty
    daily_hour = {str(d): {str(h): {'gmv': round(v['gmv']), 'units': v['units']}
                            for h, v in hmap.items()}
                  for d, hmap in sorted(_daily_hour.items())}

    # ── Daily SKU breakdown (for Stock + TopItems range filter) ───────────────
    _daily_sku = {}   # day → sku → {gmv, units}
    for _o in orders_cur:
        _ds  = _o.get('date_closed') or _o.get('date_created') or ''
        _st  = _o.get('status', '')
        if _st == 'cancelled': continue
        try:
            _day = int(_ds[8:10])
        except Exception:
            _day = 0
        if not _day: continue
        for _it in _o.get('items', []):
            _iid   = _it.get('item_id', '')
            _qty   = _it.get('quantity', 0) or 0
            _price = float(_it.get('unit_price', 0) or 0)
            _sku_k, _ = _get_sku(_iid, _it.get('title', ''))
            if _day not in _daily_sku: _daily_sku[_day] = {}
            if _sku_k not in _daily_sku[_day]: _daily_sku[_day][_sku_k] = {'gmv': 0.0, 'units': 0}
            _daily_sku[_day][_sku_k]['gmv']   += _qty * _price
            _daily_sku[_day][_sku_k]['units'] += _qty
    daily_sku = {str(d): {s: {'gmv': round(v['gmv']), 'units': v['units']}
                           for s, v in smap.items()}
                 for d, smap in sorted(_daily_sku.items())}

    # ── Daily SKU cancelaciones (para tabla dinámica con filtro de fecha) ─────
    _daily_sku_canc = {}   # day → sku → {c: canc_qty, p: paid_qty}
    for _o in orders_cur:
        _ds  = _o.get('date_closed') or _o.get('date_created') or ''
        _st  = _o.get('status', '')
        try:
            _day = int(_ds[8:10])
        except Exception:
            _day = 0
        if not _day: continue
        for _it in _o.get('items', []):
            _iid   = _it.get('item_id', '')
            _title = _it.get('title', '')
            _qty   = _it.get('quantity', 0) or 0
            _sk, _ = _get_sku(_iid, _title)
            if _day not in _daily_sku_canc: _daily_sku_canc[_day] = {}
            if _sk not in _daily_sku_canc[_day]: _daily_sku_canc[_day][_sk] = {'c': 0, 'p': 0}
            if _st == 'cancelled': _daily_sku_canc[_day][_sk]['c'] += _qty
            else:                  _daily_sku_canc[_day][_sk]['p'] += _qty
    daily_sku_canc = {str(d): {s: v for s, v in smap.items()}
                      for d, smap in sorted(_daily_sku_canc.items())}

    # ── Day-of-week mapping for current month ──────────────────────────────────
    from datetime import date as _date
    _yr, _mo = today.year, today.month
    day_to_dow = {str(d): _date(_yr, _mo, d).weekday()
                  for d in range(1, max_day + 1)}

    # ── Category breakdown for Ventas (with SKU drill-down) ──────────────────
    by_cat_cur = {}   # cat → {gmv, units, skus: {sku_key: {gmv, units, title, sub_cat}}}
    by_cat_pri = {}   # cat → {gmv, units}

    for s_key, agg in sku_agg_cur.items():
        cat = agg['categoria'] or 'OTROS'
        if cat not in by_cat_cur:
            by_cat_cur[cat] = {'gmv': 0, 'units': 0, 'skus': {}}
        by_cat_cur[cat]['gmv']   += agg['gmv']
        by_cat_cur[cat]['units'] += agg['units']
        by_cat_cur[cat]['skus'][s_key] = {
            'id':      s_key,
            'title':   agg['title'],
            'sub_cat': agg.get('sub_cat',''),
            'units':   agg['units'],
            'gmv':     agg['gmv'],
        }

    # Prior categories
    for _o in orders_pri_data:
        if _o.get('status') == 'cancelled': continue
        for _it in _o.get('items', []):
            _iid = _it.get('item_id','')
            _sku, _ = _get_sku(_iid, _it.get('title',''))
            _cat, _ = _get_cat(_sku, _it.get('title',''))
            _cat  = _cat or 'OTROS'
            _qty  = _it.get('quantity',0) or 0
            _g    = (_it.get('unit_price',0) or 0) * _qty
            if _cat not in by_cat_pri:
                by_cat_pri[_cat] = {'gmv': 0, 'units': 0}
            by_cat_pri[_cat]['gmv']   += round(_g)
            by_cat_pri[_cat]['units'] += _qty

    # ── SKU-consolidated cancelaciones ────────────────────────────────────────
    _sku_canc = {}
    for o in orders_cur:
        for it in o.get('items', []):
            iid   = it.get('item_id','')
            title = it.get('title','')
            s_key, _ = _get_sku(iid, title)
            cat, _   = _get_cat(s_key, title)
            qty      = it.get('quantity', 0) or 0
            st       = o.get('status','')
            if s_key not in _sku_canc:
                _sku_canc[s_key] = {'sku': s_key, 'title': title, 'categoria': cat, 'c': 0, 'p': 0}
            if st == 'cancelled': _sku_canc[s_key]['c'] += qty
            else:                 _sku_canc[s_key]['p'] += qty

    sku_cancels = []
    for d in _sku_canc.values():
        tot = d['c'] + d['p']
        d['rate'] = round(d['c'] / tot * 100, 1) if tot else 0.0
        if d['c'] > 0: sku_cancels.append(d)
    sku_cancels.sort(key=lambda x: (-x['c'], -x['rate']))

    # ── Stock table (SKU-consolidated, master mapping) ────────────────────────
    item_avail = {i['item_id']: (i.get('available_quantity') or 0) for i in items_list}
    _sku_stock = {}

    for iid, d in units_by_item.items():
        s_key, _ = _get_sku(iid, d['title'])
        cat, sub_cat = _get_cat(s_key, d['title'])
        avail_q = item_avail.get(iid, 0)

        if s_key not in _sku_stock:
            _st  = stk_by_code.get(s_key.upper(), {})
            _sku_stock[s_key] = {
                'item_id':      s_key,
                'title':        d['title'],
                'categoria':    cat,
                'sub_cat':      sub_cat,
                'codigo':       s_key if (s_key in cat_excel or s_key in stk_by_code) else '',
                'units_period': 0,
                'gmv_period':   0,
                'available_qty': 0,
                # stock_dep: desde maestro si existe; fallback = available_qty (stock ML API)
                'stock_dep':    _st.get('stock_dep') if _st else avail_q,
                'stock_full':   _st.get('stock_full') if _st else 0,
                'stock_aduana': _st.get('stock_aduana') if _st else 0,
                'has_status':   bool(_st),
                'dias_stock':   None,
            }
        _sku_stock[s_key]['units_period']  += d['units']
        _sku_stock[s_key]['gmv_period']    += round(d['gmv'])
        _sku_stock[s_key]['available_qty'] += avail_q

    for s_key, row in _sku_stock.items():
        dr = row['units_period'] / unique_days if unique_days else 0
        row['dias_stock'] = round(row['available_qty'] / dr) if dr > 0 and row['available_qty'] > 0 else None

    ml_items_list = sorted(_sku_stock.values(), key=lambda x: (x['categoria'], -x['gmv_period']))
    cats_ml = sorted(set(r.get('categoria','') for r in ml_items_list if r.get('categoria','')))

    # ── Other data ───────────────────────────────────────────────────────────
    dow_count   = mc.get('dow_count', [1]*7)
    dow_raw     = mc.get('dow', [0]*7)
    by_hour     = ec.get('by_hour', {})
    lt_cur      = ec.get('by_listing_type', {})
    lt_pri      = ep.get('by_listing_type', {})
    log_cur     = ec.get('by_logistic', {})
    log_pri     = ep.get('by_logistic', {})
    heatmap_data= ec.get('heatmap', [[0]*24 for _ in range(7)])
    daily_lt    = ec.get('daily_lt', {})
    daily_log   = ec.get('daily_log', {})
    daily_lt_u  = ec.get('daily_lt_u', {})
    daily_log_u = ec.get('daily_log_u', {})

    # ── Postventa data ────────────────────────────────────────────────────────
    _pv_cur_path = os.path.join(BASE_DIR, 'data', 'postventa_current.json')
    _pv_pri_path = os.path.join(BASE_DIR, 'data', 'postventa_prior.json')
    _claims_path = os.path.join(BASE_DIR, 'data', 'claims_analysis.json')
    claims_analysis = json.loads(open(_claims_path).read()) if os.path.exists(_claims_path) else {}
    pv_cur = json.loads(open(_pv_cur_path).read()) if os.path.exists(_pv_cur_path) else {}
    pv_pri = json.loads(open(_pv_pri_path).read()) if os.path.exists(_pv_pri_path) else {}
    _cc_path = os.path.join(BASE_DIR, 'data', 'cancelled_processed.json')
    cancelled_proc = json.loads(open(_cc_path).read()) if os.path.exists(_cc_path) else {}

    # ── Mediation reasons per MLA (from cancelled_current.json) ───────────────
    def _build_med_reasons():
        from collections import defaultdict
        _canc_path = os.path.join(BASE_DIR, 'data', 'cancelled_current.json')
        if not os.path.exists(_canc_path):
            return {}
        _orders = json.loads(open(_canc_path).read())
        def _cat(desc):
            if not desc: return 'otro'
            d = desc.lower()
            if 'cancel_purchase' in d: return 'pre_envio'
            if 'mediations cancel' in d: return 'reclamo'
            if 'not delivered' in d or 'shipment' in d: return 'logistica'
            return 'otro'
        _sku_med = defaultdict(lambda: {'pre_envio':0,'reclamo':0,'logistica':0,'otro':0})
        for _o in _orders:
            if not _o.get('mediations'):
                continue
            _cd = _o.get('cancel_detail') or {}
            _c = _cat(_cd.get('description',''))
            for _it in _o.get('items',[]):
                _mla = _it.get('item_id','')
                _qty = _it.get('quantity',1)
                _sku_med[_mla][_c] += _qty
        return {mla: dict(v) for mla,v in _sku_med.items()}
    med_reasons = _build_med_reasons()

    def _build_sku_cancel_reasons():
        from collections import defaultdict
        _cp = os.path.join(BASE_DIR, 'data', 'cancelled_current.json')
        if not os.path.exists(_cp): return {}
        _ords = json.loads(open(_cp).read())
        CODE_MAP = {'buyer_cancel_express':'buyer','pack_splitted':'pack',
                    'mediations':'med','shipment_not_delivered':'noent',
                    'fraud':'fraud','shipment_unfulfilled':'unfulf'}
        _sd = defaultdict(lambda:{'title':'','buyer':0,'pack':0,'med':0,'noent':0,'fraud':0,'unfulf':0,'other':0,'total':0})
        for _o in _ords:
            _cd = _o.get('cancel_detail') or {}
            _cat = CODE_MAP.get(_cd.get('code',''), 'other')
            for _it in _o.get('items',[]):
                _mla = _it.get('item_id','')
                if not _mla: continue
                _qty = _it.get('quantity',1)
                if not _sd[_mla]['title']: _sd[_mla]['title'] = _it.get('title','')
                _sd[_mla][_cat] += _qty
                _sd[_mla]['total'] += _qty
        return {k: dict(v) for k,v in _sd.items() if v['total'] >= 1}
    sku_cancel_reasons = _build_sku_cancel_reasons()


    import calendar as _cal
    _MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
              "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    _cm, _cy = today.month, today.year
    _pm, _py = (_cm - 1, _cy) if _cm > 1 else (12, _cy - 1)
    _pd = min(today.day, _cal.monthrange(_py, _pm)[1])
    period_cur = "{} {} (1-{})".format(_MESES[_cm], _cy, today.day)
    period_pri = "{} {} (1-{})".format(_MESES[_pm], _py, _pd)
    updated_at = (sm.get('updated_at','')[:16] or datetime.now().isoformat()[:16]).replace('T',' ')

    a_inv    = float(ads.get('inversion', 0) or 0)
    a_rev    = float(ads.get('ingresos', 0) or 0)
    a_impr   = float(ads.get('impresiones', 0) or 0)
    a_clics  = float(ads.get('clics', 0) or 0)
    a_ventas = float(ads.get('ventas', 0) or 0)
    a_roas   = f"{a_rev/a_inv:.2f}x" if a_inv else "—"
    a_tacos  = f"{a_inv/a_rev*100:.1f}%" if a_rev else "—"
    a_ctr    = f"{a_clics/a_impr*100:.2f}%" if a_impr else "—"
    a_cpc    = fmt_money(a_inv/a_clics) if a_clics else "—"
    a_conv   = f"{a_ventas/a_clics*100:.1f}%" if a_clics else "—"

    rep_st  = rep.get('power_seller_status', '')
    rep_lbl = {'platinum':'MercadoLíder Platinum','gold':'MercadoLíder Gold',
               'silver':'MercadoLíder Silver'}.get(rep_st, rep_st.title() if rep_st else '—')
    rep_col = {'platinum':'#00d4e4','gold':'#FFD700','silver':'#A0A0A0'}.get(rep_st, '#aaa')

    num_master  = len(mla_to_sku)
    num_skus    = len(sku_top_items)
    num_covered = sum(1 for x in sku_top_items if x.get('method') == 'master')
    pct_covered = round(sum(x['gmv'] for x in sku_top_items if x.get('method') == 'master') / max(gmv_c,1) * 100, 1)

    out_path = os.path.join(OUT_DIR, 'ml_dashboard_360.html')
    with open(out_path, 'w', encoding='utf-8') as f:
        def w(s): f.write(s + '\n')

        # HEAD + CSS con dark/light mode
        w('<!DOCTYPE html><html lang="es" data-theme="dark">')
        w('<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">')
        w(f'<title>ML 360° · SPOTCOMPRAS · {updated_at}</title>')
        w('<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>')
        f.write("""<style>
/* ── TOKENS DE COLOR ── */
[data-theme="dark"] {
  --bg:#0d0d1a; --surface:#151528; --surface2:#1a1a35; --border:#1e1e3a;
  --text:#e0e0f0; --text2:#999; --text3:#555;
  --blue:#3483FA; --yellow:#FFE600; --green:#00A650; --red:#F23D4F;
  --purple:#7B2FF7; --cyan:#00d4e4; --orange:#FF6B35;
  --kpi-bg:#151528; --nav-bg:#0d0d2e; --card-bg:#151528;
  --chart-grid:#1e1e3a; --chart-text:#666;
  --shadow:0 4px 20px rgba(0,0,0,.4);
  --transition:background .25s,color .25s,border-color .25s;
}
[data-theme="light"] {
  --bg:#f0f2f8; --surface:#ffffff; --surface2:#f7f8fc; --border:#e2e6f0;
  --text:#1a1a2e; --text2:#5a6380; --text3:#9aa0b4;
  --blue:#1a6ef5; --yellow:#d4a000; --green:#007a3c; --red:#d42b3a;
  --purple:#5b1fc8; --cyan:#0098ad; --orange:#d45000;
  --kpi-bg:#ffffff; --nav-bg:#ffffff; --card-bg:#ffffff;
  --chart-grid:#e8ecf4; --chart-text:#8892a8;
  --shadow:0 2px 12px rgba(0,0,80,.08);
  --transition:background .25s,color .25s,border-color .25s;
}
*{box-sizing:border-box;margin:0;padding:0;transition:var(--transition)}
body{background:var(--bg);color:var(--text);font-family:system-ui,-apple-system,sans-serif;font-size:14px}
.nav{display:flex;gap:4px;padding:8px 12px;background:var(--nav-bg);border-bottom:1px solid var(--border);overflow-x:auto;flex-wrap:nowrap;position:sticky;top:0;z-index:100;align-items:center;box-shadow:var(--shadow)}
.nav-btn{background:var(--surface2);color:var(--text2);border:1px solid var(--border);border-radius:8px;padding:6px 12px;cursor:pointer;white-space:nowrap;font-size:13px;font-weight:500;letter-spacing:.2px}
.nav-btn:hover,.nav-btn.active{background:var(--blue);color:#fff;border-color:var(--blue)}
.nav-btn.active{font-weight:700}
.theme-btn{margin-left:auto;background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:6px 12px;cursor:pointer;font-size:15px;line-height:1;color:var(--text2)}
.theme-btn:hover{border-color:var(--blue)}
.tab{display:none;padding:14px 16px}.tab.active{display:block}
.tab-hdr{background:var(--surface);border:1px solid var(--border);border-bottom:none;border-radius:12px 12px 0 0;margin-bottom:0;overflow:hidden;box-shadow:var(--shadow)}.tab-hdr.alone{border-bottom:1px solid var(--border);border-radius:12px;margin-bottom:14px}.tab-hdr-top{display:flex;align-items:center;justify-content:space-between;padding:14px 18px 12px}.tab-hdr-left{display:flex;align-items:center;gap:12px}.tab-hdr-icon{font-size:22px;line-height:1}.tab-hdr-title{font-size:15px;font-weight:700;color:var(--text);line-height:1.2}.tab-hdr-meta{font-size:11px;color:var(--text3);margin-top:3px}.tab-hdr-kpi-val{font-size:22px;font-weight:800;line-height:1.1;text-align:right}.tab-hdr-kpi-lbl{font-size:10px;color:var(--text3);text-align:right;margin-top:2px;text-transform:uppercase;letter-spacing:.4px}
.badge-master{display:inline-block;background:#3483FA20;color:var(--blue);border:1px solid #3483FA44;border-radius:12px;padding:2px 8px;font-size:10px;font-weight:700;margin-top:4px}
.card{background:var(--card-bg);border-radius:12px;border:1px solid var(--border);margin-bottom:14px;overflow:visible;box-shadow:var(--shadow);transition:border-color .15s,box-shadow .15s}.card:hover{border-color:rgba(52,131,250,.35);box-shadow:0 6px 20px rgba(0,0,0,.18)}
.card-title{padding:10px 16px;font-size:11px;font-weight:700;color:var(--blue);text-transform:uppercase;letter-spacing:.5px;border-bottom:1px solid var(--border);border-radius:12px 12px 0 0}
.card-body{padding:14px 16px}
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:10px;margin-bottom:14px}
.kpi{background:var(--kpi-bg);border-radius:12px;border:1px solid var(--border);padding:12px 14px;box-shadow:var(--shadow);transition:border-color .15s,box-shadow .15s,transform .15s}.kpi:hover{border-color:var(--blue);box-shadow:0 4px 16px rgba(52,131,250,.18);transform:translateY(-2px);cursor:default}.liq-kpi{transition:border-color .15s,box-shadow .15s,transform .15s;cursor:default}.liq-kpi:hover{box-shadow:0 6px 20px rgba(0,0,0,.25);transform:translateY(-2px);filter:brightness(1.08)}
.kpi-label{font-size:11px;color:var(--text3);text-transform:uppercase;letter-spacing:.3px;margin-bottom:5px}
.kpi-val{font-size:20px;font-weight:700;color:var(--text)}
.kpi-delta{font-size:12px;margin-top:3px}.kpi-pri{font-size:11px;color:var(--text3);margin-top:2px}
.good{color:var(--green)}.bad{color:var(--red)}.neutral{color:var(--text3)}
.chart-wrap{position:relative;width:100%}
.tbl-scroll{overflow-x:auto}
table{width:100%;border-collapse:collapse;font-size:13px}
th{padding:8px 11px;text-align:left;font-size:11px;color:var(--text3);font-weight:600;text-transform:uppercase;border-bottom:2px solid var(--border);white-space:nowrap}
td{padding:8px 11px;border-bottom:1px solid var(--border);white-space:nowrap;color:var(--text)}
tr:hover td{background:var(--surface2)}
.cat-row{cursor:pointer;user-select:none}
.cat-row td{font-weight:700;color:var(--blue);background:var(--surface2)}
.cat-row:hover td{background:var(--border)}
.sku-row td{font-size:12px;padding-left:28px;color:var(--text2)}
.sku-row{display:none}
.sku-row.expanded{display:table-row}
.expand-icon{font-size:10px;margin-right:5px;display:inline-block;transition:transform .2s}
.slider-row{display:flex;align-items:center;gap:10px;padding:10px 14px;background:var(--surface2);border-radius:8px;margin-bottom:12px;flex-wrap:wrap;border:1px solid var(--border)}
.slider-row label{font-size:12px;color:var(--text2);white-space:nowrap}
.slider-row input[type=range]{flex:1;min-width:100px;accent-color:var(--blue)}
.slider-val{font-size:14px;font-weight:700;color:var(--yellow);white-space:nowrap;min-width:90px}
[data-theme="light"] .slider-val{color:var(--blue)}
.pct-row{display:flex;align-items:center;gap:8px;margin-bottom:9px}
.pct-lbl{font-size:12px;color:var(--text);min-width:130px;flex-shrink:0}
.pct-track{flex:1;background:var(--surface2);border-radius:4px;height:11px;overflow:hidden;border:1px solid var(--border)}
.pct-fill{height:100%;border-radius:4px}
.pct-num{font-size:12px;color:var(--text2);min-width:100px;text-align:right;flex-shrink:0}
.badge{display:inline-block;padding:2px 7px;border-radius:10px;font-size:11px;font-weight:700}
.b-ok{background:#00A65018;color:var(--green)}.b-low{background:#FFE60018;color:var(--yellow)}.b-zero{background:#F23D4F18;color:var(--red)}
[data-theme="light"] .b-ok{background:#00A65015}.b-low{background:#d4a00015}.b-zero{background:#d42b3a15}
.rep-metric{display:flex;justify-content:space-between;align-items:center;padding:9px 0;border-bottom:1px solid var(--border)}
.rep-metric:last-child{border-bottom:none}
.rep-lbl{font-size:12px;color:var(--text2)}.rep-val{font-size:13px;font-weight:700}
.ads-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px}
.ads-kpi{background:var(--surface2);border-radius:10px;padding:12px;text-align:center;border:1px solid var(--border)}.pv-card{transition:transform .18s,box-shadow .18s;cursor:default;}.pv-card:hover{transform:translateY(-3px);box-shadow:0 6px 20px rgba(52,131,250,.13);}
.ads-kpi .lbl{font-size:11px;color:var(--text3);margin-bottom:6px}
.ads-kpi .val{font-size:20px;font-weight:700}
.recon-ok{background:#00A65012;color:var(--green);border:1px solid #00A65030;border-radius:8px;padding:8px 14px;font-size:12px;margin-bottom:10px}
.recon-warn{background:#F23D4F12;color:var(--red);border:1px solid #F23D4F30;border-radius:8px;padding:8px 14px;font-size:12px;margin-bottom:10px}
@media(max-width:900px){.kpi-grid{grid-template-columns:repeat(3,1fr)}}
@media(max-width:600px){.kpi-grid{grid-template-columns:repeat(2,1fr)}.kpi-val{font-size:17px}.tab-hdr-title{font-size:13px}.tab{padding:10px 8px}.tab-hdr-top{flex-wrap:wrap;padding:12px 14px 10px;gap:6px}.tab-hdr-kpi-val{font-size:20px!important;text-align:left}.tab-hdr-meta{font-size:10px}.period-presets{padding:7px 10px;gap:4px}.period-chip{font-size:11px;padding:0 9px;height:27px;line-height:25px}.period-custom{padding:7px 10px;gap:5px}.period-custom-lbl{font-size:9px;letter-spacing:0}.period-info{font-size:11px;min-width:70px;padding:4px 8px}.period-input{width:38px;font-size:12px}.grid-3col{grid-template-columns:repeat(2,1fr)!important}.grid-2col-stack{grid-template-columns:1fr!important}.liq-cash-grid{grid-template-columns:repeat(2,1fr)!important}.liq-charts-grid{grid-template-columns:1fr!important}.pct-lbl{min-width:100px}.pct-num{min-width:70px}}
@media(max-width:400px){.nav-btn{padding:5px 8px;font-size:11px}.kpi-grid{grid-template-columns:repeat(2,1fr)!important}.grid-3col{grid-template-columns:1fr 1fr!important}}
</style></head>""")

        w('<body>')
        w('<div class="nav">')
        for tab_n, lbl in [(0,'📊 Resumen'),(10,'💰 Liquidaciones'),(1,'📈 Ventas'),(2,'🕐 Horarios'),
                              (5,'🔄 Postventa'),(6,'🏅 Reputación'),(7,'⚡ Ads'),(9,'📅 Anual')]:
            ac = ' active' if tab_n == 0 else ''
            w(f'<button class="nav-btn{ac}" data-tab="{tab_n}" onclick="showTab({tab_n})">{lbl}</button>')
        w('<button class="theme-btn" onclick="toggleTheme()" id="theme-icon" title="Cambiar tema">☾</button>')
        w('</div>')

        # DATA SCRIPT
        w('<script>')
        w(f'const DAILY_CUR={jd(daily_cur)};')
        w(f'const DAILY_PRI={jd(daily_pri)};')
        w(f'const UNIQUE_DAYS={unique_days};')
        w(f'const DOW_COUNT={jd(dow_count)};')
        w(f'const DOW_RAW={jd(dow_raw)};')
        w(f'const DOW_UNITS_RAW={jd(dow_units_raw)};')
        w(f'const BY_HOUR={jd(by_hour)};')
        w(f'const HEATMAP={jd(heatmap_data)};')
        w(f'const BY_CAT_CUR={jd(by_cat_cur)};')
        w(f'const BY_CAT_PRI={jd(by_cat_pri)};')
        w(f'const DAILY_BY_CAT={jd(daily_by_cat)};')
        w(f'const DAILY_BY_CAT_PRI={jd(daily_by_cat_pri)};')
        w(f'const DAILY_HOUR={jd(daily_hour)};')
        w(f'const DAILY_SKU={jd(daily_sku)};')
        w(f'const DAILY_SKU_CANC={jd(daily_sku_canc)};')
        # SKU_CANC_META: sku → {t: title_short, c: categoria} para la tabla JS
        _sku_canc_meta = {v['sku']: {'t': (v.get('title') or v['sku'])[:35], 'c': v.get('categoria','')}
                          for v in _sku_canc.values() if v.get('sku')}
        w(f'const SKU_CANC_META={jd(_sku_canc_meta)};')
        w(f'const DAY_TO_DOW={jd(day_to_dow)};')
        w(f'const LT_CUR={jd(lt_cur)};')
        w(f'const LT_PRI={jd(lt_pri)};')
        w(f'const LOG_CUR={jd(log_cur)};')
        w(f'const LOG_PRI={jd(log_pri)};')
        w(f'const DAILY_LT={jd(daily_lt)};')
        w(f'const DAILY_LOG={jd(daily_log)};')
        w(f'const DAILY_LT_U={jd(daily_lt_u)};')
        w(f'const DAILY_LOG_U={jd(daily_log_u)};')
        w(f'const PV_CUR={jd(pv_cur)};')
        w(f'const PV_PRI={jd(pv_pri)};')
        w(f'const CANCELLED_PROC={jd(cancelled_proc)};')
        w(f'const CLAIMS_ANALYSIS={jd(claims_analysis)};')  
        w(f'const MED_REASONS={jd(med_reasons)};')
        w(f'const SKU_CANCEL_REASONS={jd(sku_cancel_reasons)};')

        # ── LIQUIDACIONES DATA ────────────────────────────────────────────
        import json as _json, os as _os
        from collections import defaultdict as _dd
        _liq_sum = {}
        _fact_cf = []
        _daily_liq = {}
        _liq_data_path = _os.path.join(DATA_DIR, 'liq_summary.json')
        _cf_data_path  = _os.path.join(DATA_DIR, 'fact_cashflow.json')
        _conc_data_path= _os.path.join(DATA_DIR, 'fact_conciliacion.json')
        if _os.path.exists(_liq_data_path):
            with open(_liq_data_path, encoding='utf-8') as _f: _liq_sum = _json.load(_f)
        if _os.path.exists(_cf_data_path):
            with open(_cf_data_path, encoding='utf-8') as _f: _fact_cf = _json.load(_f)
        if _os.path.exists(_conc_data_path):
            with open(_conc_data_path, encoding='utf-8') as _f:
                _all_conc = _json.load(_f)
            # Aggregate by date — executive daily view (no individual order detail)
            _agg = _dd(lambda: {'gmv':0.0,'fee_ml':0.0,'shipping':0.0,'shipping_real':0.0,'taxes_real':0.0,'devuelto':0.0,'neto':0.0,'n_orders':0,'n_fee_derived':0,'ship_breakdown':{},'tax_breakdown':{}})
            _rel_neto = _dd(dict)  # sale_date → {release_date: sum_neto}
            for _r in _all_conc:
                if _r.get('estado') == 'cobro_sin_orden': continue
                _d = _r.get('date','')
                if not _d: continue
                _agg[_d]['gmv']           += _r.get('gmv_venta', 0)
                _fee_ml_r = _r.get('fee_ml', 0)
                if not _fee_ml_r and (_r.get('shipping_real',0)>0 or _r.get('taxes_real',0)>0):
                    _gmv_c = _r.get('gmv_cobrado',0) or _r.get('gmv_venta',0)
                    _fee_ml_r = max(0, _gmv_c - _r.get('neto_real',0)
                                   - (_r.get('shipping_real',0) or _r.get('shipping',0))
                                   - _r.get('taxes_real',0) - _r.get('fee_mp',0)
                                   - _r.get('fee_fin',0) - _r.get('devuelto',0)
                                   - _r.get('diff_monto',0))
                    if _fee_ml_r > 0: _agg[_d]['n_fee_derived'] += 1
                _agg[_d]['fee_ml'] += _fee_ml_r
                _agg[_d]['shipping']      += _r.get('shipping', 0)
                _agg[_d]['shipping_real'] += _r.get('shipping_real', 0)  # Full + Colecta real
                _agg[_d]['taxes_real']    += _r.get('taxes_real', 0)     # IIBB + IDB
                _agg[_d]['devuelto']      += _r.get('devuelto', 0)
                _agg[_d]['neto']          += _r.get('neto_real', 0)
                _agg[_d]['n_orders']      += 1
                # Breakdown por tipo de envío (Full vs Colecta) y retenciones (IIBB vs IDB)
                for _k, _v in (_r.get('shipping_detail') or {}).items():
                    _agg[_d]['ship_breakdown'][_k] = _agg[_d]['ship_breakdown'].get(_k, 0) + _v
                for _k, _v in (_r.get('tax_detail') or {}).items():
                    _agg[_d]['tax_breakdown'][_k] = _agg[_d]['tax_breakdown'].get(_k, 0) + _v
                _rd = _r.get('release_date','')
                if _rd:
                    _rel_neto[_d][_rd] = _rel_neto[_d].get(_rd, 0) + _r.get('neto_real', 0)
            # Pick the release_date with the most $ weight for each sale day
            _rel_by_day = {_d: max(_v, key=_v.get) for _d, _v in _rel_neto.items() if _v}
            _daily_liq = {k: dict(
                              {kk: (round(vv, 2) if isinstance(vv, float) else vv)
                               for kk, vv in v.items()},
                              release_date=_rel_by_day.get(k, ''))
                          for k, v in _agg.items()}
        # ── COBRO SUMMARY: cobrado / pendiente / tránsito / plazo real ──
        from datetime import date as _dc
        _ts = str(_dc.today())
        _cob_sum = {'cobrado':0.0,'pendiente':0.0,'transito':0.0,'n_pend':0,'plazo_items':[]}
        for _r in _all_conc:
            _n  = _r.get('neto_real', 0)
            _rd = _r.get('release_date','')
            _d  = _r.get('date','')
            if _n <= 0 or not _rd or not _d: continue
            _est = _r.get('estado','')
            if _est == 'pendiente_liberar' and _rd >= _ts:
                try: _dias = (_dc.fromisoformat(_rd) - _dc.fromisoformat(_d)).days
                except: continue
                if _dias >= 25: _cob_sum['transito']  += _n
                else:
                    _cob_sum['pendiente'] += _n
                    _cob_sum['n_pend']    += 1
                    _cob_sum['plazo_items'].append((_dias, _n))
            else:
                _cob_sum['cobrado'] += _n
        _pi = _cob_sum.pop('plazo_items')
        _tp2 = sum(n for _,n in _pi)
        _cob_sum['plazo_promedio'] = round(sum(d*n for d,n in _pi)/_tp2, 1) if _tp2 else 0
        _cob_sum = {k: round(v,2) if isinstance(v,float) else v for k,v in _cob_sum.items()}

        _hourly_path = _os.path.join(DATA_DIR, 'fact_cf_hourly.json')
        _fact_hourly = {}
        if _os.path.exists(_hourly_path):
            with open(_hourly_path, encoding='utf-8') as _f: _fact_hourly = _json.load(_f)
        # Prior period reference for delta comparison
        _liq_prior = {
            'gmv':      round(mp.get('gmv', 0), 2),
            'neto':     round(mp.get('net', 0), 2),
            'fee_rate': round(mp.get('fee_rate', 0), 2),
            'fees':     round(mp.get('fees', 0), 2),
        }
        # ── MP_BALANCE ─────────────────────────────────────────────
        import json as _json2, os as _os2
        _mp_bal = None
        _mp_bal_path = _os.path.join(DATA_DIR, 'mp_balance.json')
        if _os.path.exists(_mp_bal_path):
            with open(_mp_bal_path, encoding='utf-8') as _f: _mp_bal = _json.load(_f)
        w(f'const MP_BALANCE={jd(_mp_bal)};')
        w(f'const LIQ_SUMMARY={jd(_liq_sum)};')
        w(f'const FACT_CASHFLOW={jd(_fact_cf)};')
        w(f'const FACT_CF_HOURLY={jd(_fact_hourly)};')
        w(f'const DAILY_LIQ={jd(_daily_liq)};')
        w(f'const LIQ_COBRO={jd(_cob_sum)};')
        w(f'const LIQ_PRIOR={jd(_liq_prior)};')
        w(f'const MONTHLY={jd(mon)};')
        w(f'const ADS_DAILY={jd(ads_d)};')
        w(f'const ADS_YEAR={today.year};')
        w(f'const ADS_MONTH={today.month};')
        w(f'const ML_ITEMS_DATA={jd(ml_items_list)};')
        w(f'const TOP_ITEMS={jd(sku_top_items)};')
        w(f'const TOP_ITEMS_PRI={jd(sku_top_pri)};')
        w(f'const SHIP_PERF={jd(ship_perf)};')
        w(f'const SHIP_DAILY={jd(ship_daily)};')
        w(f'const NEG_REVIEWS={jd(neg_revs)};')
        # SINGLE SOURCE OF TRUTH — totales derivados de orders_current.json
        w(f'const GMV_CUR={gmv_c};')
        w(f'const GMV_PRI={gmv_p};')
        w(f'const UNITS_CUR={un_c};')
        w(f'const UNITS_PRI={un_p};')
        w(f'const PAID_CUR={pd_c};')
        w(f'const PAID_PRI={pd_p};')
        w(f'const CANC_CUR={canc_c};')
        w(f'const CANC_PRI={canc_p};')
        w(f'const TK_CUR={tk_c};')
        w(f'const TK_PRI={tk_p};')
        w(f'const CR_CUR={cr_c};')
        w(f'const CR_PRI={cr_p};')
        w(f'const FEE_RATE_C={fee_rate_c};')
        w(f'const FEE_RATE_P={fee_rate_p};')
        w(f'const MAX_DAY={max_day};')
        w(f'const PERIOD_CUR={jd(period_cur)};')
        w(f'const PERIOD_PRI={jd(period_pri)};')
        w('</script>')

        # ── GLOBAL PERIOD BAR (visible en todos los tabs) ───────────────────
        w('<div class="period-bar">')
        w('<div class="period-presets">')
        w('<span class="period-custom-lbl">Per&iacute;odo</span>')
        w(f'<button class="period-chip active" id="chip-mes" onclick="setRange(1,{max_day})">Mes completo</button>')
        w('<button class="period-chip" id="chip-s1" onclick="setRange(1,7)">Sem 1</button>')
        w('<button class="period-chip" id="chip-s2" onclick="setRange(8,14)">Sem 2</button>')
        w(f'<button class="period-chip" id="chip-s3" onclick="setRange(15,{max_day})">Sem 3</button>')
        w('<button class="period-chip hot" id="chip-hot" onclick="setRange(11,17)">Hot Sale</button>')
        w('</div>')
        w('<div class="period-divider"></div>')
        w('<div class="period-custom">')
        w('<span class="period-custom-lbl">Rango</span>')
        w(f'<input type="number" class="period-input" id="day-from" min="1" max="{max_day}" value="1" oninput="applyRange()">')
        w('<span class="period-sep">&#8211;</span>')
        w(f'<input type="number" class="period-input" id="day-to" min="1" max="{max_day}" value="{max_day}" oninput="applyRange()">')
        w(f'<span class="period-info" id="range-lbl">1–{max_day} ({max_day} d&iacute;as)</span>')
        w('</div>')
        w('</div>')

        # ── TAB 0: RESUMEN ───────────────────────────────────────────────────
        w('<div id="tab0" class="tab active">')
        w('<div class="tab-hdr">')
        w('  <div class="tab-hdr-top">')
        w('    <div class="tab-hdr-left">')
        w('      <span class="tab-hdr-icon">📊</span>')
        w('      <div>')
        w(f'        <div class="tab-hdr-title">Dashboard ML 360&deg; &middot; SPOTCOMPRAS</div>')
        w(f'        <div class="tab-hdr-meta">Actualizado: {updated_at} &nbsp;&middot;&nbsp; {period_cur} vs {period_pri}</div>')
        w('      </div>')
        w('    </div>')
        w('    <div><div class="tab-hdr-kpi-val" style="color:var(--blue)" id="hdr0-kpi">—</div><div class="tab-hdr-kpi-lbl">GMV del per&iacute;odo</div></div>')
        w('  </div>')
        w('</div>')
        w('<div class="kpi-grid" id="kpi-grid"></div>')
        w('<div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:14px">')
        w('<div class="card">')
        w('  <div class="card-title" style="display:flex;justify-content:space-between;align-items:center">')
        w('    <span>Tipo de publicación</span>')
        w('    <div style="display:flex;gap:4px">')
        w('      <button id="lt-btn-gmv" onclick="setShipMode(\'lt\',\'gmv\')" style="padding:2px 8px;font-size:11px;border-radius:4px;border:1px solid var(--blue);background:var(--blue);color:#fff;cursor:pointer;font-weight:700">$ GMV</button>')
        w('      <button id="lt-btn-u" onclick="setShipMode(\'lt\',\'u\')" style="padding:2px 8px;font-size:11px;border-radius:4px;border:1px solid var(--border);background:transparent;color:var(--text2);cursor:pointer">Unidades</button>')
        w('    </div>')
        w('  </div>')
        w('  <div class="card-body" id="lt-section"></div>')
        w('</div>')
        w('<div class="card">')
        w('  <div class="card-title" style="display:flex;justify-content:space-between;align-items:center">')
        w('    <span>Tipo de envío</span>')
        w('    <div style="display:flex;gap:4px">')
        w('      <button id="log-btn-gmv" onclick="setShipMode(\'log\',\'gmv\')" style="padding:2px 8px;font-size:11px;border-radius:4px;border:1px solid var(--cyan);background:var(--cyan);color:#fff;cursor:pointer;font-weight:700">$ GMV</button>')
        w('      <button id="log-btn-u" onclick="setShipMode(\'log\',\'u\')" style="padding:2px 8px;font-size:11px;border-radius:4px;border:1px solid var(--border);background:transparent;color:var(--text2);cursor:pointer">Unidades</button>')
        w('    </div>')
        w('  </div>')
        w('  <div class="card-body" id="log-section"></div>')
        w('</div>')
        w('</div>')
        w('<div class="card">')
        w('  <div class="card-title" style="display:flex;justify-content:space-between;align-items:center">')
        w('    <span>GMV diario · mes actual vs anterior</span>')
        w('    <div style="display:flex;gap:4px">')
        w('      <button id="gmv-btn-gmv" onclick="setGmvMode(\'gmv\')" style="padding:2px 8px;font-size:11px;border-radius:4px;border:1px solid var(--blue);background:var(--blue);color:#fff;cursor:pointer;font-weight:700">$ GMV</button>')
        w('      <button id="gmv-btn-u" onclick="setGmvMode(\'u\')" style="padding:2px 8px;font-size:11px;border-radius:4px;border:1px solid var(--border);background:transparent;color:var(--text2);cursor:pointer">Unidades</button>')
        w('    </div>')
        w('  </div>')
        w('  <div class="card-body"><div class="chart-wrap" style="height:200px">')
        w('  <canvas id="c-daily"></canvas></div></div>')
        w('</div>')
        # ── TOP 10 EN RESUMEN ───────────────────────────────────────────────
        w('<div class="card" style="margin-top:14px">')
        w('  <div class="card-title" style="display:flex;justify-content:space-between;align-items:center">')
        w('    <span>🏆 Top 10 Productos del Período</span>')
        w('  </div>')
        w('  <div id="res-top10-insight" style="font-size:11px;color:var(--text2);padding:0 14px 8px;line-height:1.5"></div>')
        w('  <div class="card-body tbl-scroll">')
        w('    <table style="width:100%;border-collapse:collapse;font-size:12px">')
        w('      <thead><tr>')
        w('        <th style="width:20px;text-align:center;color:var(--text3)">#</th>')
        w('        <th style="text-align:left">SKU / Producto</th>')
        w('        <th style="text-align:center">GMV</th>')
        w('        <th style="text-align:center">% GMV</th>')
        w('        <th style="text-align:center">Ticket Prom.</th>')
        w('        <th style="text-align:center">Unidades</th>')
        w('        <th style="text-align:center">% Units</th>')
        w('        <th style="text-align:center">vs Ant.</th>')
        w('      </tr></thead>')
        w('      <tbody id="res-top10-tbody"></tbody>')
        w('    </table>')
        w('  </div>')
        w('</div>')
        w('</div>')

        # ── TAB 1: VENTAS con drill-down por SKU ─────────────────────────────
        w('<div id="tab1" class="tab">')
        w('<div class="tab-hdr">')
        w('  <div class="tab-hdr-top">')
        w('    <div class="tab-hdr-left">')
        w('      <span class="tab-hdr-icon">📈</span>')
        w('      <div>')
        w(f'        <div class="tab-hdr-title">Ventas por Categor&iacute;a &rarr; SKU</div>')
        w(f'        <div class="tab-hdr-meta">{period_cur} vs {period_pri} &nbsp;&middot;&nbsp; Click en categor&iacute;a para expandir</div>')
        w('      </div>')
        w('    </div>')
        w('    <div><div class="tab-hdr-kpi-val" style="color:var(--blue)" id="hdr1-kpi">—</div><div class="tab-hdr-kpi-lbl">GMV del per&iacute;odo</div></div>')
        w('  </div>')
        w('</div>')
        w('<div class="card"><div class="card-body tbl-scroll">')
        w('<table id="ventas-table"><thead><tr>')
        w('<th>Categoría / SKU</th><th>GMV</th><th>Unidades</th><th>Ticket Prom.</th><th>% GMV</th><th>vs Ant.</th><th style="text-align:right">Dep.</th><th style="text-align:right">Full</th><th style="text-align:right">Aduana</th><th style="text-align:center">Cobert.</th></tr></thead>')
        w('<tbody id="cat-tbody"></tbody></table></div></div>')
        w('</div>')

        # ── TAB 2: HORARIOS ──────────────────────────────────────────────────
        w('<div id="tab2" class="tab">')
        w('<div class="tab-hdr">')
        w('  <div class="tab-hdr-top">')
        w('    <div class="tab-hdr-left">')
        w('      <span class="tab-hdr-icon">🕐</span>')
        w('      <div>')
        w(f'        <div class="tab-hdr-title">Horarios de venta</div>')
        w(f'        <div class="tab-hdr-meta">{period_cur} &nbsp;&middot;&nbsp; {unique_days} d&iacute;as activos</div>')
        w('      </div>')
        w('    </div>')
        w('    <div><div class="tab-hdr-kpi-val" style="color:var(--orange)" id="hdr2-kpi">—</div><div class="tab-hdr-kpi-lbl">unidades totales</div></div>')
        w('  </div>')
        w('</div>')
        w('<div class="card"><div class="card-title">Promedio GMV por hora del día</div>')
        w('<div class="card-body"><div class="chart-wrap" style="height:240px"><canvas id="c-hour"></canvas></div></div></div>')
        w('<div class="card"><div class="card-title">Promedio por día de semana</div>')
        w('<div class="card-body"><div class="chart-wrap" style="height:240px"><canvas id="c-dow"></canvas></div></div></div>')
        w('<div class="card"><div class="card-title">Heatmap GMV · día × hora</div>')
        w('<div class="card-body" id="heatmap-wrap" style="overflow-x:auto"></div></div>')
        w('<div class="card" style="border-left:3px solid var(--yellow)"><div class="card-title">Insights</div>')
        w('<div class="card-body" id="hora-insights" style="font-size:12px;line-height:2"></div></div>')
        w('</div>')

        # ── TAB 3: STOCK ─────────────────────────────────────────────────────

        # ── TAB 4: CANCELACIONES ─────────────────────────────────────────────
        # ── TAB 5: POSTVENTA ──────────────────────────────────────────────────
        w('<div id="tab5" class="tab">')
        w('<div class="tab-hdr alone">')
        w('  <div class="tab-hdr-top">')
        w('    <div class="tab-hdr-left">')
        w('      <span class="tab-hdr-icon">🔄</span>')
        w('      <div>')
        w(f'        <div class="tab-hdr-title">Postventa &middot; Reclamos y Cancelaciones</div>')
        w(f'        <div class="tab-hdr-meta">{period_cur}</div>')
        w('      </div>')
        w('    </div>')
        w('    <div><div class="tab-hdr-kpi-val" style="color:var(--red)" id="hdr5-kpi">—</div><div class="tab-hdr-kpi-lbl">cancelaciones</div></div>')
        w('  </div>')
        w('</div>')

        # ── Bloque 1: Cancelaciones ───────────────────────────────────────────
        _cc   = (cancelled_proc.get("current") or {})
        _ccp  = (cancelled_proc.get("prior")   or {})
        _ct   = _cc.get("total", 0)
        _ctp  = _ccp.get("total", 0)
        _cbc  = _cc.get("by_code") or {}
        _n_med_canc = (_cbc.get("mediations") or {}).get("n", 0)
        _n_buyer    = (_cbc.get("buyer_cancel_express") or {}).get("n", 0)
        _n_pack     = (_cbc.get("pack_splitted") or {}).get("n", 0)
        _n_logist   = (_cbc.get("shipment_not_delivered") or {}).get("n", 0) + (_cbc.get("shipment_unfulfilled") or {}).get("n", 0)
        _n_other    = _ct - _n_med_canc - _n_buyer - _n_pack - _n_logist
        w('<div class="card" style="margin-bottom:14px;border-top:3px solid var(--purple)">')
        w(f'<div class="card-title">❌ Órdenes Canceladas &nbsp;<span style="font-size:10px;font-weight:400;color:var(--text3)">{_ct} total Mayo · {_ctp} Abril · Var: {_ct-_ctp:+d}</span></div>')
        w('<div id="pv-pills-wrap" style="padding:6px 14px 10px"></div>')
        w('<div class="card-body" id="pv-cancel-wrap"></div>')
        w('</div>')

        # ── Bloque 2: Mediaciones ─────────────────────────────────────────────
        _cl   = claims_analysis.get("summary", {})
        _open = _cl.get("total_opened", 0)
        _urg  = _cl.get("urgente_claim_stage", 0)
        _disp = _cl.get("dispute_stage", 0)
        _cld  = _cl.get("total_closed_real", 0)
        _bw   = _cl.get("buyer_wins_closed", 0)
        _mix  = _cl.get("mixed_closed", 0)
        _ok   = max(_cld - _bw - _mix, 0)
        _res  = (claims_analysis.get("closed_combined") or {}).get("by_reason", {})
        _res_tot = sum(_res.values()) or 1
        w('<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px">')

        # Left: Mediaciones cerradas — visual cards
        w('<div class="card" style="border-top:3px solid var(--orange)">')
        w(f'<div class="card-title">⚖️ {_cld} Mediaciones Cerradas &nbsp;<span style="font-size:10px;font-weight:400;color:var(--text3)">Últimos 60 días · Cómo se resolvieron</span></div>')
        w('<div style="font-size:10px;color:var(--text3);margin:-4px 0 10px 0;padding:5px 8px;background:rgba(255,193,7,.08);border-left:3px solid var(--orange);border-radius:0 4px 4px 0">ℹ️ Dato histórico (60 días). No equivale a las cancelaciones del mes: las 71 &quot;canceladas por reclamo&quot; son órdenes de Mayo; estas 62 son reclamos resueltos por ML en los últimos 60 días.</div>')
        w('<div class="card-body">')
        # 3 mini-cards for outcomes
        w(f'<div class="grid-3col" style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-bottom:14px">')
        w(f'<div style="padding:12px;border-radius:8px;background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.2);text-align:center" class="pv-card">')
        w(f'<div style="font-size:28px;font-weight:800;color:var(--red)">{_bw}</div>')
        w(f'<div style="font-size:10px;color:var(--text3);margin-top:2px">😞 Perdiste</div>')
        w(f'<div style="font-size:10px;color:var(--red);font-weight:600">{round(_bw/_cld*100) if _cld else 0}% del total</div>')
        w('</div>')
        w(f'<div style="padding:12px;border-radius:8px;background:rgba(52,131,250,.1);border:1px solid rgba(52,131,250,.2);text-align:center" class="pv-card">')
        w(f'<div style="font-size:28px;font-weight:800;color:#3483FA">{_mix}</div>')
        w(f'<div style="font-size:10px;color:var(--text3);margin-top:2px">🤝 Mixto</div>')
        w(f'<div style="font-size:10px;color:#3483FA;font-weight:600">Ambas partes</div>')
        w('</div>')
        w(f'<div style="padding:12px;border-radius:8px;background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.2);text-align:center" class="pv-card">')
        w(f'<div style="font-size:28px;font-weight:800;color:var(--green)">{_ok}</div>')
        w(f'<div style="font-size:10px;color:var(--text3);margin-top:2px">✅ A tu favor</div>')
        w(f'<div style="font-size:10px;color:var(--green);font-weight:600">Vendedor gana</div>')
        w('</div>')
        w('</div>')  # end 3-col
        # Resolution breakdown bars
        _res_items = [
            ("coverage_decision",  "🛡️ Cobertura ML",      "var(--red)"),
            ("warehouse_decision", "🏭 Depósito/logística", "var(--orange)"),
            ("item_returned",      "📦 Devuelto",           "var(--blue)"),
            ("return_cancelled",   "🚫 Dev. cancelada",     "var(--text3)"),
        ]
        for _rk, _rl, _rc in _res_items:
            _rn = _res.get(_rk, 0)
            if not _rn: continue
            _rp = round(_rn / _res_tot * 100)
            w(f'<div style="margin-bottom:7px"><div style="display:flex;justify-content:space-between;font-size:11px;margin-bottom:2px"><span style="color:var(--text2)">{_rl}</span><span style="color:var(--text3);font-weight:600">{_rn} · {_rp}%</span></div><div style="height:6px;background:var(--border);border-radius:3px"><div style="height:6px;background:{_rc};border-radius:3px;width:{_rp}%"></div></div></div>')
        w('</div></div>')


        # Right: Mediaciones abiertas
        w('<div class="card" style="border-top:3px solid var(--red)">')
        w(f'<div class="card-title">🔴 {_open} Reclamos Activos &nbsp;<span style="font-size:10px;font-weight:400;color:var(--text3)">Pendientes de resolución</span></div>')
        w('<div class="card-body">')
        w(f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px">')
        w(f'<div style="padding:14px;border-radius:8px;background:rgba(239,68,68,.08);text-align:center" class="pv-card">')
        w(f'<div style="font-size:32px;font-weight:800;color:var(--red)">{_urg}</div>')
        w(f'<div style="font-size:11px;color:var(--text2);font-weight:700;margin-top:4px">🟡 EN RECLAMO</div>')
        w(f'<div style="font-size:10px;color:var(--text3);margin-top:4px">Podés responder AHORA<br>antes que ML intervenga</div>')
        w('</div>')
        w(f'<div style="padding:14px;border-radius:8px;background:rgba(251,146,60,.08);text-align:center" class="pv-card">')
        w(f'<div style="font-size:32px;font-weight:800;color:var(--orange)">{_disp}</div>')
        w(f'<div style="font-size:11px;color:var(--text2);font-weight:700;margin-top:4px">🔴 EN DISPUTA</div>')
        w(f'<div style="font-size:10px;color:var(--text3);margin-top:4px">ML arbitrando.<br>Sin acción disponible.</div>')
        w('</div>')
        w('</div>')
        w(f'<div style="padding:8px 12px;background:rgba(52,131,250,.07);border-radius:6px;font-size:11px;color:var(--text3)">💡 Los {_urg} reclamos en etapa "reclamo" permiten respuesta del vendedor. Ingresá a tu cuenta ML → Reclamaciones para gestionarlos antes que escalen.</div>')
        w('</div></div>')
        w('</div>')  # end 2-col grid

        # ── Bloque 3: Top 10 productos críticos por mediaciones ───────────────
        w(f'<div class="card" style="margin-bottom:14px">')
        w(f'<div class="card-title" id="pv-top10-title">📉 Top Productos por Devoluciones &nbsp;<span style="font-size:10px;font-weight:400;color:var(--text3)">y reclamos activos · ordenado por tasa %Dev</span></div>')
        w('<div class="card-body"><table style="width:100%"><thead><tr>')
        w('<th style="width:24px;text-align:center">#</th>')
        w('<th>Producto</th>')
        w('<th style="text-align:center">Unidades</th>')
        w('<th style="color:var(--orange);text-align:center">Devol.</th>')
        w('<th style="color:var(--orange);text-align:center">%Dev.</th>')
        w('<th style="color:var(--red);text-align:center">🔴 Activos</th>')
        w('</tr></thead><tbody id="pv-top10-tbody"></tbody></table></div></div>')

        w('</div>')  # close tab5


        # ── TAB 5: REPUTACIÓN ────────────────────────────────────────────────
        claims_pct  = (rep.get('claims_rate', 0) or 0) * 100
        delayed_pct = (rep.get('delayed_rate', 0) or 0) * 100
        canc_rep    = (rep.get('cancellations_rate', 0) or 0) * 100
        sales60     = rep.get('sales_60d', 0) or 0
        rat_pos     = (rep.get('ratings_positive', 0) or 0) * 100
        rat_neg     = (rep.get('ratings_negative', 0) or 0) * 100
        total_ord   = rep.get('total', 0) or 0
        completed   = rep.get('completed', 0) or 0
        canceled_h  = rep.get('canceled', 0) or 0
        claims_val  = rep.get('claims_value', 0) or 0
        delayed_val = rep.get('delayed_value', 0) or 0
        canc_val    = rep.get('cancellations_value', 0) or 0
        rat_neu     = max(0, round(100 - rat_pos - rat_neg, 1))
        level_lbl   = rep.get('level_id','').replace('_',' ').title() or '—'
        fetched_lbl = (rep.get('fetched_at','') or '')[:16].replace('T',' ')
        w('<div id="tab6" class="tab">')
        w('<div class="tab-hdr alone">')
        w('  <div class="tab-hdr-top">')
        w('    <div class="tab-hdr-left">')
        w('      <span class="tab-hdr-icon">🏅</span>')
        w('      <div>')
        w(f'        <div class="tab-hdr-title" style="color:{rep_col}">{rep_lbl}</div>')
        w(f'        <div class="tab-hdr-meta">Usuario: {rep.get("nickname","—")} &nbsp;&middot;&nbsp; Nivel: {level_lbl} &nbsp;&middot;&nbsp; datos al {updated_at}</div>')
        w('      </div>')
        w('    </div>')
        w(f'    <div><div class="tab-hdr-kpi-val" style="color:var(--green)">{rat_pos:.1f}%</div><div class="tab-hdr-kpi-lbl">reputaci&oacute;n positiva</div></div>')
        w('  </div>')
        w('</div>')
        w('<div class="kpi-grid">')
        for lbl, val, col in [
            ('Ventas 60 días',   f'{sales60:,}',   'var(--text)'),
            ('Total histórico',  f'{total_ord:,}',  'var(--text)'),
            ('Completadas',      f'{completed:,}',  'var(--green)'),
            ('Canceladas hist.', f'{canceled_h:,}', 'var(--red)'),
        ]:
            w(f'<div class="kpi"><div class="kpi-label">{lbl}</div><div class="kpi-val" style="color:{col}">{val}</div></div>')
        w('</div>')
        # ── Métricas de calidad — KPI cards rediseño ───────────────────────────
        w('<div class="card">')
        w(f'<div class="card-title" style="display:flex;justify-content:space-between;align-items:center">')
        w(f'  <span>📊 Métricas de calidad · últimos 60 días · API ML</span>')
        w(f'  <span style="font-size:10px;color:var(--text3);font-weight:400">Actualizado: {fetched_lbl}</span>')
        w('</div>')
        w('<div class="card-body">')
        w('<div class="grid-3col" style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:14px">')
        for m_lbl, m_pct, m_val, m_warn, m_bad, m_icon in [
            ('Reclamos',       claims_pct,  claims_val,  1.0,  2.0,  '⚖️'),
            ('Envíos tardíos', delayed_pct, delayed_val, 6.0,  10.0, '🚚'),
            ('Cancelaciones',  canc_rep,    canc_val,    0.5,  1.5,  '❌'),
        ]:
            m_col   = 'var(--red)' if m_pct > m_bad  else ('var(--yellow)' if m_pct > m_warn else 'var(--green)')
            m_bg    = 'rgba(239,68,68,.07)' if m_pct > m_bad else ('rgba(251,191,36,.07)' if m_pct > m_warn else 'rgba(34,197,94,.07)')
            m_bord  = 'rgba(239,68,68,.3)' if m_pct > m_bad else ('rgba(251,191,36,.3)' if m_pct > m_warn else 'rgba(34,197,94,.3)')
            m_badge = '🔴 Crítico' if m_pct > m_bad else ('🟡 Atención' if m_pct > m_warn else '🟢 OK')
            # Progress bar: % del límite de alerta
            m_prog  = min(100, round(m_pct / m_bad * 100))
            m_prog_col = 'var(--red)' if m_prog > 80 else ('var(--yellow)' if m_prog > 50 else 'var(--green)')
            w(f'<div style="background:{m_bg};border:1px solid {m_bord};border-radius:10px;padding:14px 16px">')
            w(f'  <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:8px">')
            w(f'    <span style="font-size:11px;font-weight:700;color:var(--text2);text-transform:uppercase;letter-spacing:.5px">{m_icon} {m_lbl}</span>')
            w(f'    <span style="font-size:10px;font-weight:700;background:{m_bg};color:{m_col};border:1px solid {m_bord};border-radius:4px;padding:2px 6px">{m_badge}</span>')
            w(f'  </div>')
            w(f'  <div style="display:flex;align-items:baseline;gap:10px;margin-bottom:10px">')
            w(f'    <span style="font-size:28px;font-weight:800;color:{m_col};line-height:1">{m_pct:.2f}%</span>')
            w(f'    <span style="font-size:13px;color:var(--text2);font-weight:600">{m_val:,} casos</span>')
            w(f'  </div>')
            w(f'  <div style="background:var(--surface2);border-radius:4px;height:6px;overflow:hidden;margin-bottom:6px">')
            w(f'    <div style="height:100%;width:{m_prog}%;background:{m_prog_col};border-radius:4px;transition:width .4s"></div>')
            w(f'  </div>')
            w(f'  <div style="font-size:10px;color:var(--text3)">{m_prog}% del límite · alerta {m_warn}% / crítico {m_bad}%</div>')
            w(f'</div>')
        w('</div>')
        w('</div></div>')
        issues = []
        if claims_pct > 2.0:   issues.append(f'🔴 Reclamos {claims_pct:.2f}% — supera límite ML. Revisar motivos y responder abiertos.')
        if delayed_pct > 10.0: issues.append(f'🔴 Tardíos {delayed_pct:.2f}% — impacto directo en MercadoLíder.')
        if canc_rep > 1.5:     issues.append(f'🟡 Cancelaciones hist. {canc_rep:.3f}% — revisar SKUs críticos.')
        if rat_neg > 5:        issues.append(f'🔴 Calificaciones negativas {rat_neg:.1f}% — revisar publicaciones problemáticas.')
        if not issues:
            issues.append(f'✅ Todos los indicadores dentro de umbrales. Nivel <b style="color:{rep_col}">{rep_lbl}</b> estable.')
        # ── Desempeño en envíos: Flex + Colecta (layout 2 filas) ──
        _flex = ship_perf.get('flex', {})
        _col  = ship_perf.get('colecta', {})
        _fetched_ship = ship_perf.get('fetched_at', '')[:10] if ship_perf else ''
        def _exp_col(exp):
            return {'Excelente':'var(--green)','Buena':'var(--blue)','Regular':'var(--yellow)','Sin exposición':'var(--red)'}.get(exp,'var(--text3)')
        def _exp_icon(exp):
            return {'Excelente':'🟢','Buena':'🔵','Regular':'🟡','Sin exposición':'🔴'}.get(exp,'⭕')
        # ── Desempeño en envíos: layout simétrico 2 columnas ──────────────────
        def _ship_col(d):
            if not d: return ''
            ea_c = _exp_col(d.get('exposure_actual','—'))
            ep_c = _exp_col(d.get('exposure_prevista','—'))
            ea_i = _exp_icon(d.get('exposure_actual','—'))
            ep_i = _exp_icon(d.get('exposure_prevista','—'))
            _pct = d.get('envios_correctos_pct','—')
            _pc  = 'var(--green)' if _pct == '100%' else ('var(--yellow)' if _pct.rstrip('%').isdigit() and int(_pct.rstrip('%')) >= 97 else 'var(--red)')
            _ead = (d.get('exposure_actual_detail','') or '')[:60]
            _epd = (d.get('exposure_prevista_detail','') or '')[:60]
            return (
                f'<div style="flex:1;min-width:200px;padding:14px 16px">'
                f'<div style="font-size:10px;font-weight:800;color:var(--text2);text-transform:uppercase;letter-spacing:.6px;margin-bottom:10px">{d.get("label","")}</div>'
                f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px">'
                f'<div><div style="font-size:9px;color:var(--text3);margin-bottom:3px">Exposición actual</div>'
                f'<div style="font-size:13px;font-weight:800;color:{ea_c}">{ea_i} {d.get("exposure_actual","—")}</div>'
                f'<div style="font-size:9px;color:var(--text3);margin-top:3px;line-height:1.4">{_ead}</div></div>'
                f'<div><div style="font-size:9px;color:var(--text3);margin-bottom:3px">Exposición prevista</div>'
                f'<div style="font-size:13px;font-weight:800;color:{ep_c}">{ep_i} {d.get("exposure_prevista","—")}</div>'
                f'<div style="font-size:9px;color:var(--text3);margin-top:3px;line-height:1.4">{_epd}</div></div>'
                f'</div>'
                f'<div style="background:var(--surface2);border-radius:8px;padding:8px 12px">'
                f'<div style="font-size:9px;color:var(--text3);margin-bottom:2px">{d.get("periodo","Esta semana")} · Envíos correctos</div>'
                f'<div style="font-size:22px;font-weight:800;color:{_pc}">{_pct}</div>'
                f'<div style="font-size:11px;color:var(--text2)">{d.get("envios_correctos_abs","")}</div>'
                f'</div></div>'
            )
        w('<div class="card" style="margin-top:14px">')
        w(f'<div class="card-title" style="display:flex;justify-content:space-between;align-items:center">')
        w(f'  <span>🚚 Desempeño en Envíos · Mercado Libre</span>')
        w(f'  <span style="font-size:10px;color:var(--text3);font-weight:400">ACTUALIZADO: {_fetched_ship}</span>')
        w('</div>')
        w('<div class="card-body" style="padding:0">')
        w('<div style="display:flex;flex-wrap:wrap;border-bottom:1px solid var(--border)">')
        w(_ship_col(_flex))
        w('<div style="width:1px;background:var(--border);align-self:stretch"></div>')
        w(_ship_col(_col))
        w('</div>')
        w('<div style="padding:10px 16px 14px">')
        w('  <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;flex-wrap:wrap">')
        w('    <span style="font-size:11px;font-weight:600;color:var(--text2)">Histórico · 4 semanas</span>')
        w('    <button id="ship-btn-flex" onclick="renderShipChart(\'flex\')" style="padding:2px 10px;font-size:10px;border-radius:12px;border:1px solid var(--blue);background:var(--blue);color:#fff;cursor:pointer">FLEX</button>')
        w('    <button id="ship-btn-col" onclick="renderShipChart(\'col\')" style="padding:2px 10px;font-size:10px;border-radius:12px;border:1px solid var(--border);background:transparent;color:var(--text2);cursor:pointer">COLECTA</button>')
        w('    <span style="font-size:10px;color:var(--text3);margin-left:4px"><span style="display:inline-block;width:8px;height:8px;border-radius:2px;background:#27AE60;margin-right:3px"></span>Entregados</span>')
        w('    <span style="font-size:10px;color:var(--text3)"><span style="display:inline-block;width:8px;height:8px;border-radius:2px;background:#E74C3C;margin-right:3px"></span>Fallidos</span>')
        w('  </div>')
        w('  <div class="chart-wrap" style="height:200px"><canvas id="ship-chart-unified"></canvas></div>')
        w('</div>')
        w('</div>')
        w('<script>')
        w('window._shipChartMode="flex";')
        w('window._shipChartInst=null;')
        w('window.renderShipChart=function(mode){')
        w('  window._shipChartMode=mode;')
        w('  var canvas=document.getElementById("ship-chart-unified");if(!canvas)return;')
        w('  ["flex","col"].forEach(function(k){var b=document.getElementById("ship-btn-"+k);if(!b)return;')
        w('    b.style.background=k===mode?"var(--blue)":"transparent";')
        w('    b.style.color=k===mode?"#fff":"var(--text2)";')
        w('    b.style.borderColor=k===mode?"var(--blue)":"var(--border)";')
        w('  });')
        w('  var today=new Date(),dow=today.getDay()||7,mon=new Date(today);')
        w('  mon.setDate(today.getDate()-dow+1);')
        w('  function iso(d){return d.toISOString().slice(0,10);}')
        w('  var DAYS=["L","M","X","J","V","S","D"];')
        w('  var labels=[],ok_data=[],bad_data=[];')
        w('  for(var wi=0;wi<4;wi++){var ws=new Date(mon);ws.setDate(mon.getDate()-21+wi*7);')
        w('    for(var di=0;di<7;di++){var d=new Date(ws);d.setDate(ws.getDate()+di);')
        w('      var ds=iso(d),row=(SHIP_DAILY||{})[ds];')
        w('      labels.push(DAYS[di]+" "+d.getDate()+"/"+(d.getMonth()+1));')
        w('      if(!row||(row[mode+"_ok"]||0)+(row[mode+"_bad"]||0)===0){ok_data.push(null);bad_data.push(null);}')
        w('      else{ok_data.push(row[mode+"_ok"]||0);bad_data.push(row[mode+"_bad"]||0);}')
        w('    }')
        w('  }')
        w('  if(window._shipChartInst){window._shipChartInst.destroy();window._shipChartInst=null;}')
        w('  var isDark=document.documentElement.getAttribute("data-theme")==="dark";')
        w('  var gc=isDark?"rgba(255,255,255,0.06)":"rgba(0,0,0,0.05)";')
        w('  var tc=isDark?"#888":"#999";')
        w('  window._shipChartInst=new Chart(canvas,{')
        w('    type:"bar",')
        w('    data:{labels:labels,datasets:[')
        w('      {label:"Entregados",data:ok_data,backgroundColor:"#27AE60",stack:"s"},')
        w('      {label:"Fallidos",data:bad_data,backgroundColor:"#E74C3C",stack:"s"}')
        w('    ]},')
        w('    options:{')
        w('      responsive:true,maintainAspectRatio:false,animation:false,')
        w('      interaction:{mode:"index",intersect:false},')
        w('      plugins:{')
        w('        legend:{display:false},')
        w('        tooltip:{callbacks:{')
        w('          title:function(items){return items[0].label;},')
        w('          label:function(item){var v=item.raw;if(v===null||v===undefined)return null;')
        w('            return item.dataset.label+": "+(v||0).toLocaleString("es-AR");},')
        w('          afterBody:function(items){var ok=0,bad=0;')
        w('            items.forEach(function(i){if(i.datasetIndex===0)ok=i.raw||0;else bad=i.raw||0;});')
        w('            if(ok+bad===0)return[];')
        w('            return["","Total: "+(ok+bad).toLocaleString("es-AR")+" — "+Math.round(ok/(ok+bad)*100)+"% correctos"];}')
        w('        }}')
        w('      },')
        w('      scales:{')
        w('        x:{stacked:true,grid:{color:gc},ticks:{color:tc,font:{size:9}}},')
        w('        y:{stacked:true,grid:{color:gc},ticks:{color:tc,font:{size:10},')
        w('          callback:function(v){return v>=1000?(v/1000).toFixed(1)+"k":v;}}}')
        w('      }')
        w('    }')
        w('  });')
        w('};')
        w('/* renderShipChart hook installed in main script after showTab is defined */')
        w('</script>')

        # ── Calificaciones (standalone card) ─────────────────────────────────
        rat_neu_val = max(0, round(100 - rat_pos - rat_neg, 1))
        w('<div class="card" style="margin-top:14px">')
        w('<div class="card-body">')
        w('<div style="background:var(--surface2);border-radius:10px;padding:14px 16px;border:1px solid var(--border)">')
        w('  <div style="font-size:11px;font-weight:700;color:var(--text2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:12px">⭐ Calificaciones de compradores</div>')
        w('  <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:12px">')
        w(f'    <div style="text-align:center"><div style="font-size:22px;font-weight:800;color:var(--green)">{rat_pos:.1f}%</div><div style="font-size:11px;color:var(--text3)">Positivas</div></div>')
        w(f'    <div style="text-align:center"><div style="font-size:22px;font-weight:800;color:var(--text2)">{rat_neu_val:.1f}%</div><div style="font-size:11px;color:var(--text3)">Neutras</div></div>')
        w(f'    <div style="text-align:center"><div style="font-size:22px;font-weight:800;color:var(--red)">{rat_neg:.1f}%</div><div style="font-size:11px;color:var(--text3)">Negativas</div></div>')
        w('  </div>')
        w(f'  <div style="display:flex;height:8px;border-radius:6px;overflow:hidden;gap:2px">')
        w(f'    <div style="flex:{rat_pos};background:var(--green);border-radius:6px 0 0 6px"></div>')
        w(f'    <div style="flex:{rat_neu_val};background:var(--text3)"></div>')
        w(f'    <div style="flex:{rat_neg};background:var(--red);border-radius:0 6px 6px 0"></div>')
        w('  </div>')
        w('</div>')
        w('</div></div>')
        w(f'<div class="card" style="border-left:3px solid var(--cyan)"><div class="card-title">Diagnóstico ejecutivo</div>')
        w('<div class="card-body" style="font-size:12px;line-height:2">')
        for msg in issues: w(f'<div>{msg}</div>')
        w(f'<div style="margin-top:8px;padding:6px 10px;background:rgba(52,131,250,.07);border-radius:6px;font-size:11px;color:var(--text3)">💡 Nota: ML Reclamos ({claims_pct:.2f}%) mide reclamos que <em>afectan reputación</em> en ventana 60 días (método ML). Mediaciones internas ({pv_cur.get("med",0)}/{pv_cur.get("paid",1)} = {pv_cur.get("rate_med",0):.1f}%) incluye <em>todos</em> los reclamos abiertos del período. Son métricas distintas.</div>')
        w('</div></div>')
        # ── Panel expandable calificaciones negativas ───────────────────────
        _neg_total = neg_revs.get('total_negative', 0)
        _neg_items = neg_revs.get('items_with_negative', 0)
        _neg_fetched = neg_revs.get('fetched_at', '')[:10]
        w('<div class="card" style="margin-top:14px">')
        w(f'<div class="card-title" style="display:flex;justify-content:space-between;align-items:center;cursor:pointer" onclick="(function(){{var p=document.getElementById(\'neg-reviews-panel\');if(p)p.style.display=p.style.display===\'none\'?\'\':p.style.display===\'\'?\'none\':\'none\';}})()">'
        )
        w(f'  <span>⭐ Calificaciones Negativas · {_neg_total} reseñas en {_neg_items} productos</span>')
        w(f'  <span style="font-size:10px;color:var(--text3)">Actualizado: {_neg_fetched} &nbsp; ▼ Ver detalle</span>')
        w('</div>')
        w('<div id="neg-reviews-panel" style="display:none"><div class="card-body">')
        # Per-item summary first
        _item_sum = neg_revs.get('item_summary', [])
        if _item_sum:
            w('<div style="margin-bottom:12px">')
            w('<div style="font-size:11px;font-weight:700;color:var(--text2);margin-bottom:8px;text-transform:uppercase;letter-spacing:.5px">Productos con más críticas</div>')
            w('<div style="display:flex;flex-wrap:wrap;gap:6px">')
            for _it in _item_sum[:10]:
                _it_id    = _it.get('item_id', '')
                _it_title = _it.get('item_title', '')[:40]
                _it_neg   = _it.get('neg_count', 0)
                w(f'<div style="background:rgba(239,68,68,.08);border:1px solid rgba(239,68,68,.25);border-radius:8px;padding:6px 10px;font-size:11px">')
                w(f'  <div style="font-weight:700;color:var(--red)">{_it_neg} neg.</div>')
                w(f'  <div style="color:var(--text2);margin-top:2px">{_it_title}</div>')
                w(f'  <div style="color:var(--text3);font-size:10px">{_it_id}</div>')
                w(f'</div>')
            w('</div></div>')
        # Reviews table
        _reviews = neg_revs.get('reviews', [])
        if _reviews:
            w('<div style="overflow-x:auto">')
            w('<table style="width:100%;border-collapse:collapse;font-size:11px">')
            w('<thead><tr>')
            w('<th style="text-align:left;padding:6px 8px;border-bottom:1px solid var(--border);color:var(--text3)">Producto</th>')
            w('<th style="text-align:center;padding:6px 8px;border-bottom:1px solid var(--border);color:var(--text3)">Nota</th>')
            w('<th style="text-align:left;padding:6px 8px;border-bottom:1px solid var(--border);color:var(--text3)">Título</th>')
            w('<th style="text-align:left;padding:6px 8px;border-bottom:1px solid var(--border);color:var(--text3)">Comentario</th>')
            w('<th style="text-align:right;padding:6px 8px;border-bottom:1px solid var(--border);color:var(--text3)">Fecha</th>')
            w('</thead><tbody>')
            for _rv in _reviews[:50]:
                _rv_rate    = _rv.get('rate', 0)
                _rv_stars   = '★' * _rv_rate + '☆' * (5 - _rv_rate)
                _rv_item    = _rv.get('item_id', '')
                _rv_title_p = _rv.get('item_title', '')[:35]
                _rv_rtitle  = _rv.get('review_title', '')[:40]
                _rv_content = _rv.get('content', '')[:120]
                _rv_date    = _rv.get('date', '')
                _rv_col     = 'var(--red)' if _rv_rate <= 1 else 'var(--yellow)'
                w(f'<tr style="border-bottom:1px solid var(--border)">')
                w(f'<td style="padding:6px 8px;color:var(--text3);font-size:10px">{_rv_item}<br><span style="color:var(--text2)">{_rv_title_p}</span></td>')
                w(f'<td style="padding:6px 8px;text-align:center;color:{_rv_col};white-space:nowrap">{_rv_stars}</td>')
                w(f'<td style="padding:6px 8px;font-weight:600;color:var(--text)">{_rv_rtitle}</td>')
                w(f'<td style="padding:6px 8px;color:var(--text2)">{_rv_content}</td>')
                w(f'<td style="padding:6px 8px;text-align:right;color:var(--text3);white-space:nowrap">{_rv_date}</td>')
                w(f'</tr>')
            w('</tbody></table></div>')
        w('</div></div>')
        w('</div>')
        w('</div>')
        w('</div>')  # close tab6

        # ── TAB 6: ADS ───────────────────────────────────────────────────────
        ads_per = ads.get('periodo','') or ads.get('fecha','') or period_cur
        w('<div id="tab7" class="tab">')
        w('<div class="tab-hdr">')
        w('  <div class="tab-hdr-top">')
        w('    <div class="tab-hdr-left">')
        w('      <span class="tab-hdr-icon">⚡</span>')
        w('      <div>')
        w('        <div class="tab-hdr-title">Ads &amp; Costos</div>')
        w('        <div class="tab-hdr-meta" id="ads-meta">Cargando...</div>')
        w('      </div>')
        w('    </div>')
        w('    <div><div class="tab-hdr-kpi-val" style="color:#9C27B0" id="hdr7-kpi">—</div><div class="tab-hdr-kpi-lbl">ROAS</div></div>')
        w('  </div>')
        w('</div>')
        w('<div class="kpi-grid" id="ads-kpis"></div>')
        w('<div class="card"><div class="card-title">Inversión vs Ingresos diarios</div>')
        w('<div class="card-body"><div class="chart-wrap" style="height:240px"><canvas id="c-ads-daily"></canvas></div></div></div>')
        w('<div class="card"><div class="card-title">ROAS diario</div>')
        w('<div class="card-body"><div class="chart-wrap" style="height:180px"><canvas id="c-ads-roas"></canvas></div></div></div>')
        w('<div class="card"><div class="card-title">Distribución de costos</div><div class="card-body" id="ads-costos"></div></div>')
        w('</div>')
        w('')



        # ── TAB 9: ANUAL ──────────────────────────────────────────────────
        w('<div id="tab9" class="tab">')
        w('<div class="tab-hdr alone">')
        w('  <div class="tab-hdr-top">')
        w('    <div class="tab-hdr-left">')
        w('      <span class="tab-hdr-icon">📅</span>')
        w('      <div>')
        w('        <div class="tab-hdr-title">Anual 2026</div>')
        w('        <div class="tab-hdr-meta">Evoluci&oacute;n mensual &middot; GMV y Ticket promedio</div>')
        w('      </div>')
        w('    </div>')
        w('    <div><div class="tab-hdr-kpi-val" style="color:var(--cyan)" id="hdr9-kpi">—</div><div class="tab-hdr-kpi-lbl">GMV acumulado</div></div>')
        w('  </div>')
        w('</div>')
        w('<div class="kpi-grid" id="anual-kpi-grid" style="margin-bottom:14px"></div>')
        w('<div class="card"><div class="card-title">GMV Mensual vs Ticket Promedio</div>')
        w('<div class="card-body"><div class="chart-wrap" style="height:320px"><canvas id="c-anual"></canvas></div></div></div>')
        w('<div class="card" style="margin-top:14px"><div class="card-title">Detalle mensual</div>')
        w('<div class="card-body tbl-scroll"><table style="width:100%;border-collapse:collapse;font-size:13px">')
        w('<thead><tr style="border-bottom:2px solid var(--border)">')
        w('<th style="text-align:left;padding:8px 10px;color:var(--text3);font-weight:600">Mes</th>')
        w('<th style="text-align:right;padding:8px 10px;color:var(--text3);font-weight:600">GMV</th>')
        w('<th style="text-align:right;padding:8px 10px;color:var(--text3);font-weight:600">Unidades</th>')
        w('<th style="text-align:right;padding:8px 10px;color:var(--text3);font-weight:600">Órdenes</th>')
        w('<th style="text-align:right;padding:8px 10px;color:var(--text3);font-weight:600">Ticket Prom.</th>')
        w('<th style="text-align:right;padding:8px 10px;color:var(--text3);font-weight:600">Cancel.</th>')
        w('<th style="text-align:right;padding:8px 10px;color:var(--text3);font-weight:600">Comisiones</th>')
        w('</tr></thead><tbody id="anual-tbody"></tbody></table></div></div>')
        w('</div>')  # close tab9
        w('')

        # ── TAB 10: LIQUIDACIONES ─────────────────────────────────────────
        w('<div class="tab" id="tab10">')
        w('<div id="liq-root" style="padding:0 4px"></div>')
        w('</div>')  # close tab10


        # Extra CSS for JS-rendered elements
        w('<style>')
        w(".badge-share{display:inline-block;background:#3483FA20;color:var(--blue);border:1px solid #3483FA44;border-radius:10px;padding:1px 6px;font-size:11px;font-weight:700} .badge-sku{display:inline-block;background:#00d4e418;color:var(--cyan);border:1px solid #00d4e433;border-radius:10px;padding:1px 6px;font-size:11px} .cat-toggle{display:inline-block;width:16px;margin-right:4px;font-size:11px} .sku-row td{background:var(--surface2)!important;font-size:12px;border-bottom:1px solid var(--border)} .heatmap-tbl{border-collapse:collapse;font-size:10px;width:100%} .heatmap-tbl th{color:var(--text3);padding:2px 3px;text-align:center;font-weight:500} .heatmap-lbl{color:var(--text2);font-weight:700;padding:2px 6px;white-space:nowrap} .hm-cell{width:3.8%;text-align:center;padding:3px 1px;border-radius:3px;cursor:default;font-size:9px} .period-bar{display:flex;align-items:center;gap:0;background:var(--surface2);border:1px solid var(--border);border-top:none;border-radius:0 0 12px 12px;margin-bottom:14px;overflow:hidden;flex-wrap:wrap} .period-presets{display:flex;align-items:center;gap:6px;padding:10px 14px;flex-wrap:wrap} .period-divider{width:1px;min-height:40px;background:var(--border);flex-shrink:0} .period-custom{display:flex;align-items:center;gap:8px;padding:10px 14px} .period-custom-lbl{font-size:11px;color:var(--text3);text-transform:uppercase;letter-spacing:.5px;white-space:nowrap;font-weight:600} .period-chip{height:30px;padding:0 13px;border-radius:20px;border:1px solid var(--border);background:transparent;color:var(--text2);font-size:12px;font-weight:600;cursor:pointer;white-space:nowrap;line-height:28px;transition:all .15s ease;font-family:inherit} .period-chip:hover{border-color:var(--blue);color:var(--blue);background:#3483FA10} .period-chip.active{background:var(--blue);color:#fff;border-color:var(--blue);box-shadow:0 2px 8px #3483FA44} .period-chip.hot{border-color:#FF6B3566;color:var(--orange)} .period-chip.hot:hover{background:#FF6B3510} .period-chip.hot.active{background:var(--orange);color:#fff;border-color:var(--orange);box-shadow:0 2px 8px rgba(255,107,53,.35)} .period-input{width:46px;height:30px;padding:0 6px;border:1px solid var(--border);border-radius:8px;background:var(--surface2);color:var(--text);font-size:13px;font-weight:700;text-align:center;-moz-appearance:textfield;appearance:textfield;font-family:inherit} .period-input::-webkit-inner-spin-button,.period-input::-webkit-outer-spin-button{-webkit-appearance:none} .period-input:focus{outline:none;border-color:var(--blue);box-shadow:0 0 0 3px #3483FA22} .period-sep{color:var(--text3);font-size:15px;font-weight:300;padding:0 2px} .period-info{font-size:12px;font-weight:700;color:var(--blue);background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:5px 11px;white-space:nowrap;min-width:90px;text-align:center}")
        w('</style>')

        # (Chart.js CDN already loaded in <head>)

        # Main JS block
        w('<script>')
        w("(function(){")
        w("  var t=localStorage.getItem(\"sp-theme\")||\"dark\";")
        w("  document.documentElement.setAttribute(\"data-theme\",t);")
        w("  var btn=document.getElementById(\"theme-icon\");")
        w("  if(btn)btn.textContent=t===\"dark\"?\"\u2600\":\"\u263e\";")
        w("})();")
        w("")
        w("function toggleTheme(){")
        w("  var cur=document.documentElement.getAttribute(\"data-theme\");")
        w("  var next=cur===\"dark\"?\"light\":\"dark\";")
        w("  document.documentElement.setAttribute(\"data-theme\",next);")
        w("  localStorage.setItem(\"sp-theme\",next);")
        w("  var btn=document.getElementById(\"theme-icon\");")
        w("  if(btn)btn.textContent=next===\"dark\"?\"\u2600\":\"\u263e\";")
        w("  Object.keys(_charts).forEach(function(k){if(_charts[k]){_charts[k].destroy();_charts[k]=null;}});")
        w("  _initTab(_activeTab);")
        w("}")
        w("")
        w("var _activeTab=0;")
        w("var _charts={};")
        w("")
        w("function showTab(n){")
        w("  document.querySelectorAll(\".tab\").forEach(function(t){t.classList.remove(\"active\");});") 
        w("  var el=document.getElementById(\"tab\"+n);if(el)el.classList.add(\"active\");")
        w("  document.querySelectorAll(\".nav-btn\").forEach(function(b){b.classList.toggle(\"active\",parseInt(b.dataset.tab)===n);});")
        w("  _activeTab=n;")
        w("  _initTab(n);")
        w("  var pb=document.querySelector('.period-bar');")
        w("  if(pb){")
        w("    if(n===9||n===5||n===6||n===20){pb.style.display='none';}")
        w("    else{")
        w("      pb.style.display='';")
        w("      var _hdr=document.querySelector('#tab'+n+' .tab-hdr');")
        w("      if(_hdr&&pb.parentElement!==_hdr){_hdr.appendChild(pb);}")
        w("    }")
        w("  }")
        w("}")
        w("")
        w("function _rng(){var fe=document.getElementById('day-from'),te=document.getElementById('day-to');return[parseInt((fe||{}).value)||1,parseInt((te||{}).value)||MAX_DAY];}")
        w("function _initTab(n){")
        w("  var r=_rng(),f=r[0],t=r[1];")
        w("  if(n===0)applyRange();")
        w("  if(n===1)renderVentas(f,t);")
        w("  if(n===2)renderHorarios(f,t);")
        w("  if(n===5)renderPostventa(f,t);")
        w("  if(n===6)renderReputacion(r[0],r[1]);")
        w("  if(n===7)renderAds(f,t);")
        w("  if(n===9)renderAnual();")
        w("  if(n===10)renderLiquidaciones(f,t);")
        w("  if(n===20&&typeof renderPhishing==='function')renderPhishing();")
        w("}")
        w("")
        w("function fmtM(n){")
        w("  n=Math.round(n||0);")
        w("  if(n>=1e9)return\"$\"+(n/1e9).toFixed(2)+\"B\";")
        w("  if(n>=1e6)return\"$\"+(n/1e6).toFixed(2)+\"M\";")
        w("  if(n>=1e3)return\"$\"+(n/1e3).toFixed(1)+\"K\";")
        w("  return\"$\"+n.toLocaleString(\"es-AR\");")
        w("}")
        w("function fmtN(n){return(n||0).toLocaleString(\"es-AR\");}")
        w("function dPct(a,b){return(b&&b!==0)?((a-b)/Math.abs(b))*100:null;}")
        w("function arrH(pct,inv){")
        w("  if(pct===null||pct===undefined)return'<span style=\"color:var(--text3)\">&#8212;</span>';")
        w("  var pos=inv?(pct<0):(pct>0);")
        w("  var col=pos?\"var(--green)\":\"var(--red)\";")
        w("  var sym=pct>0?\"&#9650;\":\"&#9660;\";")
        w("  return'<span style=\"color:'+col+'\">'+sym+\" \"+Math.abs(pct).toFixed(1)+\"%</span>\";")
        w("}")
        w("function cssV(v){return getComputedStyle(document.documentElement).getPropertyValue(v).trim();}")
        w("function cc(){return{")
        w("  blue:cssV(\"--blue\"),green:cssV(\"--green\"),red:cssV(\"--red\"),yellow:cssV(\"--yellow\"),")
        w("  cyan:cssV(\"--cyan\"),purple:cssV(\"--purple\"),text:cssV(\"--text\"),text2:cssV(\"--text2\"),")
        w("  text3:cssV(\"--text3\"),grid:cssV(\"--chart-grid\"),ctxt:cssV(\"--chart-text\"),")
        w("};}")
        w("function mkChart(id,cfg){")
        w("  if(_charts[id]){_charts[id].destroy();}")
        w("  var el=document.getElementById(id);if(!el)return null;")
        w("  if(!cfg.options)cfg.options={};")
        w("  cfg.options.interaction=Object.assign({mode:'index',intersect:false},cfg.options.interaction||{});")
        w("  if(!cfg.options.plugins)cfg.options.plugins={};")
        w("  if(!cfg.options.plugins.tooltip)cfg.options.plugins.tooltip={};")
        w("  if(!cfg.options.plugins.tooltip.callbacks){")
        w("    cfg.options.plugins.tooltip.callbacks={")
        w("      label:function(ctx){")
        w("        var lbl=(ctx.dataset.label||'');var ll=lbl.toLowerCase();var v=ctx.raw;")
        w("        if(v===null||v===undefined)return null;")
        w("        if(ll.includes('gmv')||ll.includes('invers')||ll.includes('ingres')||ll.includes('ticket')||ll.includes('neto')||ll.includes('cobrar')||ll.includes('pendiente'))return lbl+': '+fmtM(v);")
        w("        if(ll.includes('roas'))return lbl+': '+(+v).toFixed(2)+'x';")
        w("        if(ll.includes('unidad')||ll.includes('prom/dia')||ll.includes('prom/hora'))return lbl+': '+fmtN(v)+' u.';")
        w("        if(ll.includes('cpc'))return lbl+': '+fmtM(v);")
        w("        return lbl+': '+(typeof v==='number'?(+v).toLocaleString('es-AR'):v);")
        w("      }")
        w("    };")
        w("  }")
        w("  _charts[id]=new Chart(el,cfg);return _charts[id];")
        w("}")
        w("function bAxes(){")
        w("  var c=cc();")
        w("  return{")
        w("    x:{grid:{color:c.grid},ticks:{color:c.ctxt,font:{size:11}}},")
        w("    y:{grid:{color:c.grid},ticks:{color:c.ctxt,font:{size:11},callback:function(v){return fmtM(v);}}},")
        w("  };")
        w("}")
        w("")
        w("window._ltMode='gmv';window._logMode='gmv';window._gmvMode='gmv';")
        w("function setGmvMode(mode){")  
        w("  window._gmvMode=mode;")
        w("  var gb=document.getElementById('gmv-btn-gmv'),ub=document.getElementById('gmv-btn-u');")
        w("  if(gb){gb.style.background=mode==='gmv'?'var(--blue)':'transparent';gb.style.color=mode==='gmv'?'#fff':'var(--text2)';gb.style.borderColor=mode==='gmv'?'var(--blue)':'var(--border)';gb.style.fontWeight=mode==='gmv'?'700':'400';}")
        w("  if(ub){ub.style.background=mode==='u'?'var(--blue)':'transparent';ub.style.color=mode==='u'?'#fff':'var(--text2)';ub.style.borderColor=mode==='u'?'var(--blue)':'var(--border)';ub.style.fontWeight=mode==='u'?'700':'400';}")
        w("  var r=_rng();updateResumen(r[0],r[1]);")
        w("}")
        w("")
        w("function setShipMode(which,mode){")
        w("  if(which==='lt'){window._ltMode=mode;")
        w("    document.getElementById('lt-btn-gmv').style.background=mode==='gmv'?'var(--blue)':'transparent';")
        w("    document.getElementById('lt-btn-gmv').style.color=mode==='gmv'?'#fff':'var(--text2)';")
        w("    document.getElementById('lt-btn-gmv').style.borderColor=mode==='gmv'?'var(--blue)':'var(--border)';")
        w("    document.getElementById('lt-btn-gmv').style.fontWeight=mode==='gmv'?'700':'400';")
        w("    document.getElementById('lt-btn-u').style.background=mode==='u'?'var(--blue)':'transparent';")
        w("    document.getElementById('lt-btn-u').style.color=mode==='u'?'#fff':'var(--text2)';")
        w("    document.getElementById('lt-btn-u').style.borderColor=mode==='u'?'var(--blue)':'var(--border)';")
        w("    document.getElementById('lt-btn-u').style.fontWeight=mode==='u'?'700':'400';")
        w("  }else{window._logMode=mode;")
        w("    document.getElementById('log-btn-gmv').style.background=mode==='gmv'?'var(--cyan)':'transparent';")
        w("    document.getElementById('log-btn-gmv').style.color=mode==='gmv'?'#fff':'var(--text2)';")
        w("    document.getElementById('log-btn-gmv').style.borderColor=mode==='gmv'?'var(--cyan)':'var(--border)';")
        w("    document.getElementById('log-btn-gmv').style.fontWeight=mode==='gmv'?'700':'400';")
        w("    document.getElementById('log-btn-u').style.background=mode==='u'?'var(--cyan)':'transparent';")
        w("    document.getElementById('log-btn-u').style.color=mode==='u'?'#fff':'var(--text2)';")
        w("    document.getElementById('log-btn-u').style.borderColor=mode==='u'?'var(--cyan)':'var(--border)';")
        w("    document.getElementById('log-btn-u').style.fontWeight=mode==='u'?'700':'400';")
        w("  }")
        w("  var r=_rng();")
        w("  _pvUpdateShipBars(r[0],r[1]);")
        w("}")
        w("")
        w("function applyRange(){")
        w("  var fe=document.getElementById('day-from'),te=document.getElementById('day-to');")
        w("  var f=parseInt(fe?fe.value:1)||1;")
        w("  var t=parseInt(te?te.value:MAX_DAY)||MAX_DAY;")
        w("  f=Math.max(1,Math.min(f,MAX_DAY));t=Math.max(f,Math.min(t,MAX_DAY));")
        w("  if(fe)fe.value=f;if(te)te.value=t;")
        w("  var days=t-f+1;")
        w("  var lbl=document.getElementById('range-lbl');")
        w("  if(lbl)lbl.textContent=f+'\u2013'+t+' ('+days+' d\xedas)';")
        w("  document.querySelectorAll('.period-chip').forEach(function(b){b.classList.remove('active');});")
        w("  updateResumen(f,t);")
        w("  renderVentas(f,t);")
        w("  if(_activeTab===2)renderHorarios(f,t);")
        w("  // Postventa ignora filtro de período (datos siempre período completo)")
        w("  if(_activeTab===6)renderReputacion(f,t);")
        w("  if(_activeTab===7)renderAds(f,t);")
        w("  if(_activeTab===10)renderLiquidaciones(f,t);")
        w("}")
        w("function setRange(f,t){")
        w("  var fe=document.getElementById('day-from'),te=document.getElementById('day-to');")
        w("  if(fe)fe.value=f;if(te)te.value=t;")
        w("  applyRange();")
        w("}")
        w("function _pvBuildBreakdown(dailyData,f,t,skipLbls){")  
        w("  var merged={};")
        w("  for(var d=f;d<=t;d++){")
        w("    var dd=dailyData[String(d)]||{};")
        w("    Object.keys(dd).forEach(function(lbl){")
        w("      if(skipLbls&&skipLbls.indexOf(lbl)>=0)return;")
        w("      merged[lbl]=(merged[lbl]||0)+dd[lbl];")
        w("    });")
        w("  }")
        w("  return merged;")
        w("}")
        w("function _pvRenderBarList(elId,data,accent,isGmv){")  
        w("  var el=document.getElementById(elId);if(!el)return;")
        w("  var entries=Object.entries(data).filter(function(e){return e[1]>0;}).sort(function(a,b){return b[1]-a[1];});")
        w("  if(!entries.length){el.innerHTML='<div style=\"color:var(--text3);font-size:12px\">Sin datos</div>';return;}")
        w("  var tot=entries.reduce(function(s,e){return s+e[1];},0)||1;")
        w("  el.innerHTML=entries.map(function(e){")
        w("    var p=(e[1]/tot*100).toFixed(1);")
        w("    var valFmt=isGmv?fmtM(e[1]):fmtN(e[1])+' u.';")
        w("    return'<div class=\"pct-row\"><span class=\"pct-lbl\">'+e[0]+'</span>'")
        w("      +'<div class=\"pct-track\"><div class=\"pct-fill\" style=\"width:'+p+'%;background:'+accent+'\"></div></div>'")
        w("      +'<span class=\"pct-num\">'+valFmt+' ('+p+'%)</span></div>';")
        w("  }).join('');")
        w("}")
        w("function _pvUpdateShipBars(f,t){")  
        w("  var lm=window._ltMode||'gmv';")
        w("  var lgm=window._logMode||'gmv';")
        w("  var ltD=lm==='gmv'?_pvBuildBreakdown(DAILY_LT,f,t,['Sin clasificar']):_pvBuildBreakdown(DAILY_LT_U,f,t,['Sin clasificar']);")
        w("  var lgD=lgm==='gmv'?_pvBuildBreakdown(DAILY_LOG,f,t,[]):_pvBuildBreakdown(DAILY_LOG_U,f,t,[]);")
        w("  _pvRenderBarList('lt-section',ltD,'var(--blue)',lm==='gmv');")
        w("  _pvRenderBarList('log-section',lgD,'var(--cyan)',lgm==='gmv');")
        w("}")
        w("")
        w("function updateResumen(dayFrom,dayTo){")
        w("  var f=dayFrom||1,t=dayTo||MAX_DAY;")
        w("  var gmv=0,units=0,paid=0,canc=0,dArr=[],pArr=[];")
        w("  var gmvP=0,unitsP=0,paidP=0,cancP=0;")  
        w("  // canc tomado de DAILY_CUR/DAILY_PRI directamente")  
        w("  for(var day=1;day<=MAX_DAY;day++){")
        w("    var dc=DAILY_CUR[String(day)]||{};var dp=DAILY_PRI[String(day)]||{};")
        w("    if(day>=f&&day<=t){")
        w("      gmv+=dc.gmv||0;units+=dc.units||0;paid+=dc.paid||0;canc+=(dc.canc||0);")
        w("      gmvP+=dp.gmv||0;unitsP+=dp.units||0;paidP+=dp.paid||0;cancP+=(dp.canc||0);")
        w("    }")
        w("    dArr.push(dc.gmv||0);pArr.push(dp.gmv||0);")
        w("  }")
        w("  var cr=(paid+canc)>0?canc/(paid+canc)*100:0;")
        w("  var tk=paid>0?gmv/paid:0;")
        w("  var _h5=document.getElementById('hdr5-kpi');if(_h5){var _cp5=(CANCELLED_PROC&&CANCELLED_PROC.current)?CANCELLED_PROC.current.total:canc;_h5.textContent=fmtN(_cp5)+' cancel.';}")
        w("  var crP=(paidP+cancP)>0?cancP/(paidP+cancP)*100:0;")
        w("  var tkP=paidP>0?gmvP/paidP:0;")
        w("  var _h0=document.getElementById(\"hdr0-kpi\");if(_h0)_h0.textContent=fmtM(gmv);")
        w("  var kg=document.getElementById(\"kpi-grid\");")
        w("  if(kg){")
        w("    var rows=[")
        w("      {lbl:\"GMV Periodo\",val:fmtM(gmv),col:\"var(--blue)\",dp:dPct(gmv,gmvP),pri:\"Ant: \"+fmtM(gmvP)},")
        w("      {lbl:\"Unidades\",val:fmtN(units),col:\"var(--text)\",dp:dPct(units,unitsP),pri:\"Ant: \"+fmtN(unitsP)},")
        w("      {lbl:\"Ticket prom.\",val:fmtM(tk),col:\"var(--cyan)\",dp:dPct(tk,tkP),pri:\"Ant: \"+fmtM(tkP)},")
        w("      {lbl:\"Tasa cancel.\",val:cr.toFixed(1)+\"%\",col:cr>5?\"var(--red)\":\"var(--green)\",dp:dPct(cr,crP),inv:true,pri:\"Ant: \"+crP.toFixed(1)+\"%\"},")
        w("      {lbl:\"Pagadas\",val:fmtN(paid),col:\"var(--green)\",dp:dPct(paid,paidP),pri:\"Canc: \"+fmtN(canc)},")
        w("    ];")
        w("    kg.innerHTML=rows.map(function(k){")
        w("      var da=k.dp!==null?'<div class=\"kpi-delta\">'+arrH(k.dp,k.inv||false)+\"</div>\":\"\";")
        w("      return'<div class=\"kpi\"><div class=\"kpi-label\">'+k.lbl+'</div><div class=\"kpi-val\" style=\"color:'+k.col+'\">'+k.val+\"</div>\"+da+'<div class=\"kpi-pri\">'+k.pri+\"</div></div>\";")
        w("    }).join(\"\");")
        w("  }")
        w("  function renderPctList(elId,data,accent){")
        w("    var el=document.getElementById(elId);if(!el)return;")
        w("    var entries=Object.entries(data||{}).sort(function(a,b){return(b[1].gmv||0)-(a[1].gmv||0);});")
        w("    var tot=entries.reduce(function(s,e){return s+(e[1].gmv||0);},0)||1;")
        w("    el.innerHTML=entries.map(function(e){")
        w("      var p=(e[1].gmv/tot*100).toFixed(1);")
        w("      return'<div class=\"pct-row\"><span class=\"pct-lbl\">'+e[0]+'</span><div class=\"pct-track\"><div class=\"pct-fill\" style=\"width:'+p+'%;background:'+accent+'\"></div></div><span class=\"pct-num\">'+fmtM(e[1].gmv)+\" (\"+p+\"%)</span></div>\";")
        w("    }).join(\"\");")
        w("  }")
        w("  _pvUpdateShipBars(f,t);")
        w("  var c=cc();")
        w("  var lbl=Array.from({length:MAX_DAY},function(_,i){return i+1;});")
        w("  var bgCur=lbl.map(function(d){return(d>=f&&d<=t)?c.blue+'cc':c.text3+'33';});")
        w("  var bgPri=lbl.map(function(d){return(d>=f&&d<=t)?c.text2+'55':c.text3+'22';});")
        w("  var _gm=window._gmvMode||'gmv';")
        w("  var dArrG=Array.from({length:MAX_DAY},function(_,i){return(DAILY_CUR[String(i+1)]||{}).gmv||0;});")
        w("  var pArrG=Array.from({length:MAX_DAY},function(_,i){return(DAILY_PRI[String(i+1)]||{}).gmv||0;});")
        w("  var dArrU=Array.from({length:MAX_DAY},function(_,i){return(DAILY_CUR[String(i+1)]||{}).units||0;});")
        w("  var pArrU=Array.from({length:MAX_DAY},function(_,i){return(DAILY_PRI[String(i+1)]||{}).units||0;});")
        w("  var dArrF=_gm==='gmv'?dArrG:dArrU;")
        w("  var pArrF=_gm==='gmv'?pArrG:pArrU;")
        w("  mkChart(\"c-daily\",{type:\"bar\",")
        w("    data:{labels:lbl,datasets:[")
        w("      {label:\"Mes actual\",data:dArrF,backgroundColor:bgCur,borderRadius:4},")
        w("      {label:\"Mes anterior\",data:pArrF,backgroundColor:bgPri,borderRadius:4},")
        w("    ]},")
        w("    options:{responsive:true,maintainAspectRatio:false,")
        w("      plugins:{legend:{labels:{color:c.ctxt,font:{size:11}}},")
        w("        tooltip:{callbacks:{label:function(ctx){return _gm==='gmv'?fmtM(ctx.raw):fmtN(ctx.raw)+' u.';}}}},")
        w("      scales:bAxes()}")
        w("  });")
        # ── TOP 10 render inside updateResumen ──────────────────────────────
        w("  (function(){")
        w("    var rFiltSku={};")
        w("    for(var rd=f;rd<=t;rd++){var rds=DAILY_SKU[String(rd)]||{};Object.keys(rds).forEach(function(sk){if(!rFiltSku[sk])rFiltSku[sk]={gmv:0,units:0};rFiltSku[sk].gmv+=rds[sk].gmv||0;rFiltSku[sk].units+=rds[sk].units||0;});}")
        w("    var rAllItems=TOP_ITEMS.map(function(x){var fs=rFiltSku[x.id];return Object.assign({},x,{gmv:fs?fs.gmv:0,units:fs?fs.units:0});});")
        w("    var rSorted=rAllItems.slice().sort(function(a,b){return(b.gmv||0)-(a.gmv||0);});")
        w("    var rTop10=rSorted.slice(0,10);")
        w("    var rTotG=rAllItems.reduce(function(s,x){return s+(x.gmv||0);},0)||1;")
        w("    var rTotU=rAllItems.reduce(function(s,x){return s+(x.units||0);},0)||1;")
        w("    var rTop10G=rTop10.reduce(function(s,x){return s+(x.gmv||0);},0);")
        w("    var rTop10U=rTop10.reduce(function(s,x){return s+(x.units||0);},0);")
        w("    var rLeadGmv=rTop10[0]||{};")
        w("    var rLeadU=rTop10.slice().sort(function(a,b){return(b.units||0)-(a.units||0);})[0]||{};")
        w("    var rInsEl=document.getElementById('res-top10-insight');")
        w("    if(rInsEl){")
        w("      var ri1=(rTop10G/rTotG*100).toFixed(1);")
        w("      var ri2=fmtM(rTop10G);")
        w("      var ri3=(rTop10U/rTotU*100).toFixed(1);")
        w("      var ri4=fmtN(rTop10U);")
        w("      rInsEl.innerHTML='El Top 10 concentra <b>'+ri1+'%</b> del GMV (<b>'+ri2+'</b>) y <b>'+ri3+'%</b> de las unidades (<b>'+ri4+'</b>). Lidera en GMV: <b>'+(rLeadGmv.id||'')+'</b> · Lidera en unidades: <b>'+(rLeadU.id||'')+'</b>.';")
        w("    }")
        w("    var rTb=document.getElementById('res-top10-tbody');if(!rTb)return;")
        w("    rTb.innerHTML=rTop10.map(function(x,i){")
        w("      var pri=TOP_ITEMS_PRI[x.id]||{};")
        w("      var dpG=dPct(x.gmv,pri.gmv);")
        w("      var gSh=(x.gmv/rTotG*100).toFixed(1);")
        w("      var uSh=(x.units/rTotU*100).toFixed(1);")
        w("      var mCol=x.method==='master'?'var(--green)':x.method==='fallback'?'var(--yellow)':'var(--red)';")
        w("      var mIco=x.method==='master'?'OK':x.method==='fallback'?'~':'?';")
        w("      var rTkt=x.units>0?(x.gmv/x.units):0;")
        w("      return'<tr>'")
        w("        +'<td style=\"color:var(--text3);padding:6px 4px;text-align:center\">'+(i+1)+'</td>'")
        w("        +'<td style=\"padding:6px 8px\">'")  
        w("          +'<span style=\"color:'+mCol+';font-size:10px;font-weight:700;border:1px solid '+mCol+';border-radius:3px;padding:1px 3px\">'+mIco+'</span> '")
        w("          +'<span style=\"font-weight:700;color:var(--cyan);font-size:12px\">'+(x.id||'')+'</span> '")
        w("          +'<span style=\"color:var(--text2);font-size:11px\">'+(x.title||'').slice(0,38)+'</span>'")
        w("        +'</td>'")
        w("        +'<td style=\"font-weight:700;color:var(--blue);text-align:center;padding:6px 8px\">'+fmtM(x.gmv)+'</td>'")
        w("        +'<td style=\"text-align:center;padding:6px 4px\"><span class=\"badge-share\">'+gSh+'%</span></td>'")
        w("        +'<td style=\"text-align:center;padding:6px 8px;color:var(--text)\">'+fmtM(rTkt)+'</td>'")
        w("        +'<td style=\"text-align:center;padding:6px 8px\">'+fmtN(x.units||0)+'</td>'")
        w("        +'<td style=\"text-align:center;padding:6px 4px\"><span class=\"badge-share\">'+uSh+'%</span></td>'")
        w("        +'<td style=\"text-align:center;padding:6px 4px\">'+arrH(dpG)+'</td>'")
        w("        +'</tr>';")
        w("    }).join('');")
        w("  })();")
        w("}")
        w("")
        
        
        w("function renderReputacion(dayFrom,dayTo){")
        w("  var f=dayFrom||1;var t=dayTo||MAX_DAY;")
        w("  var el=document.getElementById('rep-period-kpis');")
        w("  if(el){")
        w("    var gmv=0,paid=0,canc=0;")
        w("    Object.keys(DAILY_CUR||{}).forEach(function(d){")
        w("      var day=parseInt(d);")
        w("      if(day>=f&&day<=t){var r=DAILY_CUR[d];gmv+=(r.gmv||0);paid+=(r.paid||0);canc+=(r.canc||0);}")
        w("    });")
        w("    el.innerHTML='<div class=\"kpi\"><div class=\"kpi-label\">Pagadas</div><div class=\"kpi-val\" style=\"color:var(--green)\">'+fmtN(paid)+'</div></div>'")
        w("     +'<div class=\"kpi\"><div class=\"kpi-label\">Canceladas</div><div class=\"kpi-val\" style=\"color:var(--red)\">'+fmtN(canc)+'</div></div>'")
        w("     +'<div class=\"kpi\"><div class=\"kpi-label\">GMV período</div><div class=\"kpi-val\">'+fmtM(gmv)+'</div></div>';")
        w("  }")
        w("  if(typeof renderShipChart==='function') renderShipChart('flex');")
        w("}")
        w("")
        w("function renderAnual(){")
        w("  var ytd=MONTHLY.filter(function(m){return m.year===new Date().getFullYear()||m.year===2026;});")
        w("  // Para meses sin fees (mes en curso), derivar de DAILY_LIQ")
        w("  ytd.forEach(function(m){")
        w("    if(m.fees===undefined||m.fees===null){")
        w("      var pfx=m.year+'-'+(m.month<10?'0':'')+m.month;")
        w("      m.fees=Object.keys(DAILY_LIQ||{}).filter(function(d){return d.startsWith(pfx);})")
        w("        .reduce(function(s,d){return s+(DAILY_LIQ[d].fee_ml||0);},0);")
        w("    }")
        w("  });")
        w("  var totalGmv=ytd.reduce(function(s,m){return s+m.gmv;},0);")
        w("  var totalU=ytd.reduce(function(s,m){return s+m.units;},0);")
        w("  var totalPaid=ytd.reduce(function(s,m){return s+m.paid;},0);")
        w("  var totalCanc=ytd.reduce(function(s,m){return s+m.cancelled;},0);")
        w("  var totalFees=ytd.reduce(function(s,m){return s+m.fees;},0);")
        w("  var avgTicket=totalPaid>0?totalGmv/totalPaid:0;")
        w("  var bestM=ytd.reduce(function(a,b){return b.gmv>a.gmv?b:a},{gmv:0,label:'-'});")
        w("  var grid=document.getElementById('anual-kpi-grid');")
        w("  if(grid){")
        w("    var cards=[")
        w("      {lbl:'GMV YTD',val:fmtM(totalGmv),sub:''},")
        w("      {lbl:'Unidades',val:fmtN(totalU),sub:''},")
        w("      {lbl:'Órdenes Pagas',val:fmtN(totalPaid),sub:''},")
        w("      {lbl:'Ticket Prom.',val:fmtM(avgTicket),sub:''},")
        w("      {lbl:'Cancelaciones',val:fmtN(totalCanc),sub:totalPaid+totalCanc>0?((totalCanc/(totalPaid+totalCanc)*100).toFixed(1)+'% tasa'):''},")
        w("      {lbl:'Comisiones',val:fmtM(totalFees),sub:totalGmv>0?((totalFees/totalGmv*100).toFixed(1)+'% del GMV'):''},")
        w("      {lbl:'Mejor Mes',val:bestM.label,sub:fmtM(bestM.gmv)},")
        w("    ];")
        w("    grid.innerHTML=cards.map(function(c){")
        w("      return '<div class=\"kpi\"><div class=\"kpi-label\">'+c.lbl+'</div>'")
        w("            +'<div class=\"kpi-val\">'+c.val+'</div>'")
        w("            +(c.sub?'<div class=\"kpi-sub\" style=\"font-size:11px;color:var(--text3)\">'+c.sub+'</div>':'')+'</div>';")
        w("    }).join('');")
        w("  }")
        w("  var hv=document.querySelector('#tab9 .tab-hdr-kpi-val');")
        w("  if(hv)hv.textContent=fmtM(totalGmv);")
        w("  var c=cc();")
        w("  mkChart('c-anual',{type:'bar',data:{")
        w("    labels:ytd.map(function(m){return m.label;}),")
        w("    datasets:[")
        w("      {type:'bar',label:'GMV',data:ytd.map(function(m){return m.gmv;}),backgroundColor:c.blue+'99',yAxisID:'y'},")
        w("      {type:'line',label:'Ticket Prom.',data:ytd.map(function(m){return m.avg_ticket;}),borderColor:c.yellow,backgroundColor:'transparent',yAxisID:'y2',tension:0.4,pointRadius:4},")
        w("    ]},options:{responsive:true,maintainAspectRatio:false,interaction:{mode:'index'},")
        w("    scales:{")
        w("      x:{grid:{color:c.grid},ticks:{color:c.ctxt,font:{size:11}}},")
        w("      y:{grid:{color:c.grid},ticks:{color:c.ctxt,font:{size:11},callback:function(v){return fmtM(v);}},title:{display:true,text:'GMV',color:c.text2}},")
        w("      y2:{position:'right',grid:{drawOnChartArea:false},ticks:{color:c.yellow,font:{size:11},callback:function(v){return fmtM(v);}},title:{display:true,text:'Ticket',color:c.yellow}},")
        w("    },plugins:{legend:{labels:{color:c.text}}}}});")
        w("  var tb=document.getElementById('anual-tbody');")
        w("  var _td='style=\\\"text-align:right;padding:8px 10px\\\"';")
        w("  var _tdl='style=\\\"text-align:left;padding:8px 10px\\\"';")
        w("  if(tb){tb.innerHTML=ytd.map(function(m){")
        w("    return '<tr style=\\\"border-bottom:1px solid var(--border)\\\">'")
        w("      +'<td '+_tdl+'>'+(m.complete?'':'<em style=\\\"color:var(--text3)\\\">')+(m.label||'')+(m.complete?'':'</em>')+'</td>'")
        w("      +'<td '+_td+'>'+fmtM(m.gmv)+'</td>'")
        w("      +'<td '+_td+'>'+fmtN(m.units)+'</td>'")
        w("      +'<td '+_td+'>'+fmtN(m.paid)+'</td>'")
        w("      +'<td '+_td+'>'+fmtM(m.avg_ticket)+'</td>'")
        w("      +'<td '+_td+'>'+fmtN(m.cancelled)+'</td>'")
        w("      +'<td '+_td+'>'+fmtM(m.fees||0)+'</td>'")
        w("      +'</tr>';")
        w("  }).join('');}")
        w("}")
        w("")
        w("var _expCats={};")
        w("function renderVentas(dayFrom,dayTo){")
        w("  var f=dayFrom||parseInt((document.getElementById('day-from')||{}).value)||1;")
        w("  var t=dayTo||parseInt((document.getElementById('day-to')||{}).value)||MAX_DAY;")
        w("  var tb=document.getElementById('cat-tbody');if(!tb)return;")
        w("  var stkMap={};ML_ITEMS_DATA.forEach(function(r){stkMap[r.codigo||r.sku||'']=r;});")
        w("  var filtSku={};")
        w("  for(var dsk=f;dsk<=t;dsk++){var dss=DAILY_SKU[String(dsk)]||{};Object.keys(dss).forEach(function(sk){if(!filtSku[sk])filtSku[sk]={gmv:0,units:0};filtSku[sk].gmv+=dss[sk].gmv||0;filtSku[sk].units+=dss[sk].units||0;});}")
        w("  var filtCat={},filtCatPri={};")
        w("  for(var d=f;d<=t;d++){")
        w("    var dc=DAILY_BY_CAT[String(d)]||{};")
        w("    var dp=DAILY_BY_CAT_PRI[String(d)]||{};")
        w("    Object.keys(dc).forEach(function(c){")
        w("      if(!filtCat[c])filtCat[c]={gmv:0,units:0};")
        w("      filtCat[c].gmv+=dc[c].gmv||0;filtCat[c].units+=dc[c].units||0;")
        w("    });")
        w("    Object.keys(dp).forEach(function(c){")
        w("      if(!filtCatPri[c])filtCatPri[c]={gmv:0,units:0};")
        w("      filtCatPri[c].gmv+=dp[c].gmv||0;filtCatPri[c].units+=dp[c].units||0;")
        w("    });")
        w("  }")
        w("  [filtCat,filtCatPri].forEach(function(fc){")
        w("    ['PEQUEÑO ELECTRO','PEQUEÑO ELECTROS'].forEach(function(b){")
        w("      if(fc[b]){if(!fc['PEQUEÑOS ELECTRO'])fc['PEQUEÑOS ELECTRO']={gmv:0,units:0};")
        w("        fc['PEQUEÑOS ELECTRO'].gmv+=fc[b].gmv;fc['PEQUEÑOS ELECTRO'].units+=fc[b].units;delete fc[b];}")
        w("    });")
        w("    if(fc['LUCES']){if(!fc['ILUMINACION'])fc['ILUMINACION']={gmv:0,units:0};")
        w("      fc['ILUMINACION'].gmv+=fc['LUCES'].gmv;fc['ILUMINACION'].units+=fc['LUCES'].units;delete fc['LUCES'];}")
        w("  });")
        w("  var totalGraw=Object.values(filtCat).reduce(function(s,c){return s+c.gmv;},0);")
        w("  var totalG=totalGraw||1;")
        w("  var totalU=Object.values(filtCat).reduce(function(s,c){return s+c.units;},0);")
        w("  var totalPriG=Object.values(filtCatPri).reduce(function(s,c){return s+c.gmv;},0);")
        w("  var _h1=document.getElementById('hdr1-kpi');if(_h1)_h1.textContent=fmtM(GMV_CUR);")
        w("  var vskKpis=document.getElementById('ventas-stock-kpis');")
        w("  if(vskKpis){")
        w("    var vskAll=ML_ITEMS_DATA;")
        w("    var vskZero=vskAll.filter(function(r){return((r.stock_dep||0)+(r.stock_full||0))===0;}).length;")
        w("    var vskDays=t-f+1;")
        w("    var vskCrit=vskAll.filter(function(r){var sk=r.codigo||r.sku||'';var pu=filtSku[sk]?filtSku[sk].units:0;var dr=pu>0?pu/vskDays:(r.vta_semana||0)/7;var stk=(r.stock_dep||0)+(r.stock_full||0);return stk>0&&dr>0&&(stk/dr)<7;}).length;")
        w("    var vskFull=vskAll.reduce(function(s,r){return s+(r.stock_full||0);},0);")
        w("    var vskDep=vskAll.reduce(function(s,r){return s+(r.stock_dep||0);},0);")
        w("    vskKpis.innerHTML=[")
        w("      {lbl:'SKUs con stock',val:fmtN(vskAll.length-vskZero)+' / '+fmtN(vskAll.length),col:'var(--text)'},")
        w("      {lbl:'Sin stock',val:vskZero,col:vskZero>0?'var(--red)':'var(--green)'},")
        w("      {lbl:'Cobertura &lt;7d',val:vskCrit,col:vskCrit>0?'var(--yellow)':'var(--green)'},")
        w("      {lbl:'Stock Dep.',val:fmtN(vskDep),col:'var(--text2)'},")
        w("      {lbl:'Stock Full',val:fmtN(vskFull),col:'var(--cyan)'},")
        w("    ].map(function(k){return'<div class=\"kpi\"><div class=\"kpi-label\">'+k.lbl+'</div><div class=\"kpi-val\" style=\"color:'+k.col+'\">'+k.val+'</div></div>';}).join('');")
        w("  }")
        w("  var cats=Object.entries(filtCat).filter(function(e){return e[1].gmv>0;}).sort(function(a,b){return b[1].gmv-a[1].gmv;});")
        w("  _CAT_KEYS=cats.map(function(e){return e[0];});")
        w("  var html='';")
        w("  var totStkDep=0,totStkFull=0,totStkAduana=0;")
        w("  cats.forEach(function(entry,idx){")
        w("    var cat=entry[0],fData=entry[1];")
        w("    var priData=filtCatPri[cat]||{};")
        w("    var fullData=BY_CAT_CUR[cat]||{gmv:0,units:0,skus:{}};")
        w("    var ratio=fullData.gmv>0?fData.gmv/fullData.gmv:1;")
        w("    var share=(fData.gmv/totalG*100).toFixed(1);")
        w("    var dpG=dPct(fData.gmv,priData.gmv);")
        w("    var tk=fData.units>0?fData.gmv/fData.units:0;")
        w("    var exp=!!_expCats[cat];")
        w("    var skuN=Object.keys(fullData.skus||{}).length;")
        w("    html+='<tr class=\"cat-row\" onclick=\"toggleCat('+idx+')\">'+''")
        w("    html+='<td><span class=\"cat-toggle\">'+(exp?'&#9660;':'&#9654;')+'</span><strong>'+cat+'</strong>'")
        w("    html+='<span style=\"color:var(--text3);font-size:11px;margin-left:6px\">'+skuN+' SKUs</span></td>'")
        w("    html+='<td style=\"font-weight:700;color:var(--blue);text-align:right\">'+fmtM(fData.gmv)+'</td>'")
        w("    html+='<td style=\"text-align:right\">'+fmtN(fData.units)+'</td>'")
        w("    html+='<td style=\"text-align:right\">'+fmtM(tk)+'</td>'")
        w("    html+='<td style=\"text-align:center\"><span class=\"badge-share\">'+share+'%</span></td>'")
        w("    var catStkDep=0,catStkFull=0,catStkAduana=0;")
        w("    Object.keys(fullData.skus||{}).forEach(function(sid){var sr=stkMap[sid]||{};catStkDep+=sr.stock_dep||0;catStkFull+=sr.stock_full||0;catStkAduana+=sr.stock_aduana||0;});")
        w("    totStkDep+=catStkDep;totStkFull+=catStkFull;totStkAduana+=catStkAduana;")
        w("    html+='<td style=\\\"text-align:center\\\">'+arrH(dpG)+'</td>'")
        w("    html+='<td style=\\\"text-align:right;color:var(--text2);font-size:12px\\\">'+fmtN(catStkDep)+'</td>'")
        w("    html+='<td style=\\\"text-align:right;color:var(--cyan);font-size:12px\\\">'+fmtN(catStkFull)+'</td>'")
        w("    html+='<td style=\\\"text-align:right;color:var(--text3);font-size:12px\\\">'+fmtN(catStkAduana)+'</td>'")
        w("    var catTotStk=catStkDep+catStkFull;")
        w("    var catUnits=fData.units||0;")
        w("    var catDays=t-f+1;")
        w("    var catDr=catUnits>0?catUnits/catDays:0;")
        w("    var catCovD=catDr>0?Math.round(catTotStk/catDr):null;")
        w("    var catCovStr=catCovD===null?'—':(catCovD>999?'>999d':catCovD+'d');")
        w("    var catCovC=catCovD===null?'var(--text3)':catCovD<7?'var(--red)':catCovD<15?'var(--yellow)':'var(--green)';")
        w("    html+='<td style=\\\"text-align:center;color:'+catCovC+';font-weight:700\\\">'+catCovStr+'</td></tr>'")
        w("    if(exp){")
        w("      var skus=Object.values(fullData.skus||{}).sort(function(a,b){return b.gmv-a.gmv;});")
        w("      skus.forEach(function(s){")
        w("        var sGmv=Math.round(s.gmv*ratio);")
        w("        var sU=Math.round(s.units*ratio);")
        w("        var sS=fData.gmv>0?(sGmv/fData.gmv*100).toFixed(1):'0.0';")
        w("        var sT=sU>0?sGmv/sU:0;")
        w("        var priSku=TOP_ITEMS_PRI[s.id]||{};")
        w("        var dpSku=priSku.gmv?dPct(sGmv,priSku.gmv):null;")
        w("        html+='<tr class=\"sku-row expanded\">';")
        w("        html+='<td style=\"padding-left:32px\"><span style=\"color:var(--cyan);font-weight:700;font-size:12px\">'+s.id+'</span> ';")
        w("        html+='<span style=\"color:var(--text2);font-size:11px\">'+(s.title||'').slice(0,45)+'</span></td>';")
        w("        html+='<td style=\"color:var(--cyan);text-align:right\">'+fmtM(sGmv)+'</td>';")
        w("        html+='<td style=\"text-align:right\">'+fmtN(sU)+'</td>';")
        w("        html+='<td style=\"text-align:right\">'+fmtM(sT)+'</td>';")
        w("        html+='<td style=\"text-align:center\"><span class=\"badge-sku\">'+sS+'%</span></td>';")
        w("        var stkR=stkMap[s.id]||{};")
        w("        var stkDep=stkR.stock_dep||0,stkFull=stkR.stock_full||0,stkAduana=stkR.stock_aduana||0;")
        w("        var stkDays2=t-f+1;")
        w("        var dr2=sU>0?sU/stkDays2:0;")
        w("        var covDays=dr2>0?Math.round((stkDep+stkFull)/dr2):null;")
        w("        var covStr=covDays===null?'—':(covDays>999?'&gt;999d':covDays+'d');")
        w("        var covC2=covDays===null?'var(--text3)':covDays<7?'var(--red)':covDays<15?'var(--yellow)':'var(--green)';")
        w("        html+='<td style=\"text-align:center\">'+arrH(dpSku)+'</td>';")
        w("        html+='<td style=\"text-align:right;color:var(--text2);font-size:11px\">'+fmtN(stkDep)+'</td>';")
        w("        html+='<td style=\"text-align:right;color:var(--cyan);font-size:11px\">'+fmtN(stkFull)+'</td>';")
        w("        html+='<td style=\"text-align:right;color:var(--text3);font-size:11px\">'+fmtN(stkAduana)+'</td>';")
        w("        html+='<td style=\"text-align:center;color:'+covC2+';font-weight:700;font-size:11px\">'+covStr+'</td></tr>';")
        w("      });")
        w("    }")
        w("  });")
        w("  var totTk=totalU>0?totalGraw/totalU:0;")
        w("  var dpTot=dPct(totalGraw,totalPriG);")
        w("  var totRow='<tr style=\"background:linear-gradient(90deg,rgba(52,131,250,.07),transparent);border-top:2px solid var(--blue);border-bottom:2px solid var(--blue)\">'+")
        w("    '<td style=\"padding:10px 10px 10px 12px\"><span style=\"font-size:12px;font-weight:800;color:var(--blue);text-transform:uppercase;letter-spacing:.3px\">📊 TOTAL PERÍODO</span>'+")
        w("    '<span style=\"color:var(--text3);font-size:11px;margin-left:8px;font-weight:400\">'+cats.length+' categorías</span></td>'+")
        w("    '<td style=\"font-weight:800;color:var(--blue);text-align:right;font-size:15px;padding-right:10px\">'+fmtM(totalGraw)+'</td>'+")
        w("    '<td style=\"font-weight:700;text-align:right;padding-right:10px\">'+fmtN(totalU)+'</td>'+")
        w("    '<td style=\"font-weight:700;text-align:right;padding-right:10px\">'+fmtM(totTk)+'</td>'+")
        w("    '<td style=\"text-align:center\"><span style=\"background:var(--blue);color:#fff;border-radius:4px;padding:2px 7px;font-size:11px;font-weight:700\">100%</span></td>'+")
        w("    '<td style=\\\"text-align:center\\\">'+arrH(dpTot)+'</td>';")
        w("  var _ts=totStkDep+totStkFull;var _td=totalU>0?totalU/(t-f+1):0;")
        w("  var _tc=_td>0?Math.round(_ts/_td):null;")
        w("  var _tcs=_tc===null?'—':(_tc>999?'>999d':_tc+'d');")
        w("  var _tcc=_tc===null?'var(--text3)':_tc<7?'var(--red)':_tc<15?'var(--yellow)':'var(--green)';")
        w('  totRow+="<td style=\\"text-align:right;color:var(--text2);font-weight:800;font-size:12px\\""+">" +fmtN(totStkDep)+"</td>";')
        w('  totRow+="<td style=\\"text-align:right;color:var(--cyan);font-weight:800;font-size:12px\\""+">" +fmtN(totStkFull)+"</td>";')
        w('  totRow+="<td style=\\"text-align:right;color:var(--text3);font-weight:800;font-size:12px\\""+">" +fmtN(totStkAduana)+"</td>";')
        w("  totRow+='<td style=\"text-align:center;font-weight:800;color:'+_tcc+'\">'+_tcs+'</td></tr>';")
        w("  tb.innerHTML=totRow+html;")
        w("}")
        w("var _CAT_KEYS=[];")
        w("function toggleCat(idx){var cat=_CAT_KEYS[idx];if(cat!==undefined){_expCats[cat]=!_expCats[cat];var fe=document.getElementById('day-from'),te=document.getElementById('day-to');renderVentas(parseInt((fe||{}).value)||1,parseInt((te||{}).value)||MAX_DAY);}}")
        w("")
        w("function renderHorarios(dayFrom,dayTo){")
        w("  var f=dayFrom||1,t=dayTo||MAX_DAY;")
        w("  var c=cc();")
        w("  var hourGmv=Array(24).fill(0),hourUnits=Array(24).fill(0),hourCnt=Array(24).fill(0);")
        w("  var dowGmv=Array(7).fill(0),dowUnits=Array(7).fill(0),dowCnt=Array(7).fill(0);")
        w("  var hmArr=Array(7).fill(null).map(function(){return Array(24).fill(0);});")
        w("  for(var d=f;d<=t;d++){")
        w("    var dh=DAILY_HOUR[String(d)]||{};")
        w("    var dayGmv=0,dayUnits=0;")
        w("    Object.keys(dh).forEach(function(h){")
        w("      var hh=parseInt(h),hv=dh[h]||{},gmv=hv.gmv||0,units=hv.units||0;")
        w("      hourGmv[hh]+=gmv;hourUnits[hh]+=units;hourCnt[hh]++;")
        w("      dayGmv+=gmv;dayUnits+=units;")
        w("    });")
        w("    var dow=DAY_TO_DOW[String(d)];")
        w("    if(dow!==undefined){dowGmv[dow]+=dayGmv;dowUnits[dow]+=dayUnits;dowCnt[dow]++;}")
        w("    Object.keys(dh).forEach(function(h){if(dow!==undefined)hmArr[dow][parseInt(h)]+=((dh[h]||{}).gmv||0);});")
        w("  }")
        w("  hourGmv=hourGmv.map(function(v,i){return hourCnt[i]?Math.round(v/hourCnt[i]):0;});")
        w("  var totUnitsH=hourUnits.reduce(function(s,v){return s+v;},0);")
        w("  var _h2=document.getElementById('hdr2-kpi');if(_h2)_h2.textContent=fmtN(UNITS_CUR)+' u.';")
        w("  hourUnits=hourUnits.map(function(v,i){return hourCnt[i]?parseFloat((v/hourCnt[i]).toFixed(1)):0;});")
        w("  dowGmv=dowGmv.map(function(v,i){return dowCnt[i]?Math.round(v/dowCnt[i]):0;});")
        w("  dowUnits=dowUnits.map(function(v,i){return dowCnt[i]?parseFloat((v/dowCnt[i]).toFixed(1)):0;});")
        w("  var l24=Array.from({length:24},function(_,i){return String(i).padStart(2,'00')+':00';});")
        w("  mkChart('c-hour',{type:'bar',")
        w("    data:{labels:l24,datasets:[")
        w("      {label:'GMV prom/hora',data:hourGmv,backgroundColor:c.blue+'cc',borderRadius:4,yAxisID:'y'},")
        w("      {label:'Unidades prom/hora',data:hourUnits,type:'line',borderColor:c.yellow,backgroundColor:c.yellow+'33',")
        w("       pointRadius:3,pointBackgroundColor:c.yellow,tension:0.35,yAxisID:'y2'}")
        w("    ]},")
        w("    options:{responsive:true,maintainAspectRatio:false,")
        w("      plugins:{legend:{labels:{color:c.ctxt,font:{size:11}}}},")
        w("      scales:{")
        w("        x:{grid:{color:c.grid},ticks:{color:c.ctxt,font:{size:11}}},")
        w("        y:{grid:{color:c.grid},ticks:{color:c.ctxt,font:{size:11},callback:function(v){return fmtM(v);}},")
        w("           title:{display:true,text:'GMV',color:c.ctxt}},")
        w("        y2:{position:'right',grid:{display:false},")
        w("            ticks:{color:c.yellow,font:{size:11}},")
        w("            title:{display:true,text:'Unidades',color:c.yellow}}")
        w("      }")
        w("    }")
        w("  });")
        w("  var DOW=['Lun','Mar','Mie','Jue','Vie','Sab','Dom'];")
        w("  mkChart('c-dow',{type:'bar',")
        w("    data:{labels:DOW,datasets:[")
        w("      {label:'GMV prom/dia',data:dowGmv,")
        w("       backgroundColor:DOW.map(function(_,i){return(i>=5?c.cyan:c.blue)+'cc';}),borderRadius:6,yAxisID:'y'},")
        w("      {label:'Unidades prom/dia',data:dowUnits,type:'line',borderColor:c.yellow,backgroundColor:c.yellow+'33',")
        w("       pointRadius:5,pointBackgroundColor:c.yellow,tension:0.2,yAxisID:'y2'}")
        w("    ]},")
        w("    options:{responsive:true,maintainAspectRatio:false,")
        w("      plugins:{legend:{labels:{color:c.ctxt,font:{size:11}}}},")
        w("      scales:{")
        w("        x:{grid:{color:c.grid},ticks:{color:c.ctxt,font:{size:11}}},")
        w("        y:{grid:{color:c.grid},ticks:{color:c.ctxt,font:{size:11},callback:function(v){return fmtM(v);}},")
        w("           title:{display:true,text:'GMV',color:c.ctxt}},")
        w("        y2:{position:'right',grid:{display:false},")
        w("            ticks:{color:c.yellow,font:{size:11}},")
        w("            title:{display:true,text:'Unidades',color:c.yellow}}")
        w("      }")
        w("    }")
        w("  });")
        w("  renderHeatmap(hmArr);")
        w("  var ins=document.getElementById('hora-insights');")
        w("  if(ins&&hourGmv.some(function(v){return v>0;})){")
        w("    var mxH=hourGmv.indexOf(Math.max.apply(null,hourGmv));")
        w("    var valH=hourGmv.filter(function(v){return v>0;});")
        w("    var mnH=valH.length?hourGmv.indexOf(Math.min.apply(null,valH)):0;")
        w("    var mxD=dowGmv.indexOf(Math.max.apply(null,dowGmv));")
        w("    var mxU=hourUnits.indexOf(Math.max.apply(null,hourUnits));")
        w("    ins.innerHTML=")
        w("      '<div>⚡ Hora pico GMV: <b>'+String(mxH).padStart(2,'0')+':00</b> &mdash; '+fmtM(hourGmv[mxH])+'/hora prom. | <b>'+hourUnits[mxH]+'</b> uds</div>'")
        w("      +'<div>🔴 Hora valle: <b>'+String(mnH).padStart(2,'0')+':00</b> &mdash; '+fmtM(hourGmv[mnH]||0)+'/hora</div>'")
        w("      +'<div>📅 Mejor d\xeda: <b>'+(DOW[mxD]||'?')+'</b> &mdash; '+fmtM(dowGmv[mxD]||0)+'/d\xeda prom. | <b>'+(dowUnits[mxD]||0).toFixed(1)+'</b> uds</div>'")
        w("      +'<div>🎯 Tip: Concentr\xe1 Ads entre '+String(Math.max(0,mxH-1)).padStart(2,'0')+':00–'+String(Math.min(23,mxH+2)).padStart(2,'0')+':00 los <b>'+(DOW[mxD]||'?')+'</b>.</div>';")
        w("  }")
        w("}")
        w("function renderHeatmap(hmData){")
        w("  var hm=hmData||(typeof HEATMAP!=='undefined'?HEATMAP:null);")
        w("  var wrap=document.getElementById('heatmap-wrap');if(!wrap||!hm||!hm.length)return;")
        w("  var DOW=['Lun','Mar','Mie','Jue','Vie','Sab','Dom'];")
        w("  var flat=hm.reduce(function(a,r){return a.concat(r.filter(function(v){return v>0;}));;},[]);")
        w("  var maxV=flat.length?Math.max.apply(null,flat):1;")
        w("  var html='<table class=\"heatmap-tbl\"><thead><tr><th></th>';")
        w("  for(var h=0;h<24;h++)html+='<th>'+String(h).padStart(2,'0')+'</th>';")
        w("  html+='</tr></thead><tbody>';")
        w("  hm.forEach(function(row,di){")
        w("    html+='<tr><td class=\"heatmap-lbl\">'+(DOW[di]||di)+'</td>';")
        w("    (row||[]).forEach(function(v){")
        w("      var p=v>0?v/maxV:0;")
        w("      var bg=v>0?'rgba(52,131,250,'+(p*0.8+0.12).toFixed(2)+')':'transparent';")
        w("      var tc=p>0.5?'#ffffff':'var(--text3)';")
        w("      html+='<td class=\"hm-cell\" style=\"background:'+bg+';color:'+tc+'\" title=\"'+fmtM(v)+'\">'+(v>0?fmtM(v).replace('$',''):'')+'</td>';")
        w("    });")
        w("    html+='</tr>';")
        w("  });")
        w("  html+='</tbody></table>';")
        w("  wrap.innerHTML=html;")
        w("}")
        w("")
        w("function renderCancChart(dayFrom,dayTo){")
        w("  var f=dayFrom||1,t=dayTo||MAX_DAY;")
        w("  var c=cc();")
        w("  var totPaid=0,totCanc=0,totGmv=0,totPaidP=0,totCancP=0;")
        w("  for(var d=f;d<=t;d++){")
        w("    var dc=DAILY_CUR[String(d)]||{},dp=DAILY_PRI[String(d)]||{};")
        w("    totPaid+=dc.paid||0;totCanc+=dc.canc||0;totGmv+=dc.gmv||0;")
        w("    totPaidP+=dp.paid||0;totCancP+=dp.canc||0;")
        w("  }")
        w("  var cr=(totPaid+totCanc)>0?totCanc/(totPaid+totCanc)*100:0;")
        w("  var crP=(totPaidP+totCancP)>0?totCancP/(totPaidP+totCancP)*100:0;")
        w("  var tk=totPaid>0?totGmv/totPaid:0;")
        w("  var gmvPerd=totCanc*tk;")
        w("  var kg=document.getElementById('canc-kpi-grid');")
        w("  if(kg){")
        w("    var rows=[")
        w("      {lbl:'Tasa cancelación',val:cr.toFixed(1)+'%',col:'var(--text)',dp:dPct(cr,crP),pri:'Ant: '+crP.toFixed(1)+'%'},")
        w("      {lbl:'Unidades canceladas',val:fmtN(totCanc),col:'var(--red)',pri:'Ant: '+fmtN(totCancP)},")
        w("      {lbl:'Órdenes pagadas',val:fmtN(totPaid),col:'var(--green)',pri:'Ant: '+fmtN(totPaidP)},")
        w("      {lbl:'GMV perdido est.',val:fmtM(gmvPerd),col:'var(--red)',pri:'Ticket: '+fmtM(tk)},")
        w("    ];")
        w("    kg.innerHTML=rows.map(function(r){")
        w("      return '<div class=\"kpi\"><div class=\"kpi-label\">'+r.lbl+'</div>'")
        w("             +'<div class=\"kpi-val\" style=\"color:'+r.col+'\">'+r.val+'</div>'")
        w("             +(r.dp!==null&&r.dp!==undefined?'<div class=\"kpi-delta '+(r.dp>=0?'good':'bad')+'\">'+arrH(r.dp)+'</div>':'')") 
        w("             +'<div class=\"kpi-pri\">'+(r.pri||'')+'</div></div>';")
        w("    }).join('');")
        w("  }")
        w("  var lbl=Array.from({length:t-f+1},function(_,i){return f+i;});")
        w("  var cArr=lbl.map(function(d){return(DAILY_CUR[String(d)]||{}).canc||0;});")
        w("  var pArr=lbl.map(function(d){return(DAILY_CUR[String(d)]||{}).paid||0;});")
        w("  mkChart('c-canc-daily',{type:'bar',")
        w("    data:{labels:lbl,datasets:[")
        w("      {label:'Pagadas',data:pArr,backgroundColor:c.green+'99',borderRadius:3},")
        w("      {label:'Canceladas',data:cArr,backgroundColor:c.red+'cc',borderRadius:3},")
        w("    ]},")
        w("    options:{responsive:true,maintainAspectRatio:false,")
        w("      plugins:{legend:{labels:{color:c.ctxt}}},")
        w("      scales:{")
        w("        x:{grid:{color:c.grid},ticks:{color:c.ctxt,font:{size:10}}},")
        w("        y:{grid:{color:c.grid},ticks:{color:c.ctxt,font:{size:10}}},")
        w("      }}")
        w("  });")
        w("}")

        w("function renderAds(dayFrom,dayTo){")
        w("  var f=dayFrom||1,t=dayTo||MAX_DAY;")
        w("  var c=cc();")
        w("  function d2s(d){var m=String(ADS_MONTH).padStart(2,'0');return ADS_YEAR+'-'+m+'-'+String(d).padStart(2,'0');}")
        w("  var cost=0,totAmt=0,dirAmt=0,indAmt=0,clicks=0,prints=0,totU=0;")
        w("  var days=[],dCost=[],dRev=[],dRoas=[];")
        w("  for(var d=f;d<=t;d++){")
        w("    var v=ADS_DAILY[d2s(d)]||{};")
        w("    var dc=v.cost||0, dr=v.total_amount||0;")
        w("    cost+=dc; totAmt+=dr;")
        w("    dirAmt+=v.direct_amount||0; indAmt+=v.indirect_amount||0;")
        w("    clicks+=v.clicks||0; prints+=v.prints||0;")
        w("    totU+=(v.direct_units_quantity||0)+(v.indirect_units_quantity||0);")
        w("    days.push(d); dCost.push(Math.round(dc)); dRev.push(Math.round(dr));")
        w("    dRoas.push(dc>0?Math.round(dr/dc*100)/100:0);")
        w("  }")
        w("  var roas=cost>0?totAmt/cost:0;")
        w("  var ctr=prints>0?clicks/prints*100:0;")
        w("  var cpc=clicks>0?cost/clicks:0;")
        w("  // TACOS = costo / GMV total período; ACOS = costo / ingresos ads")
        w("  var gmvTotP=0,unitsTotP=0;")
        w("  for(var dd=f;dd<=t;dd++){var dv=DAILY_CUR[String(dd)]||{};gmvTotP+=dv.gmv||0;unitsTotP+=dv.units||0;}")
        w("  var acos=totAmt>0?cost/totAmt*100:0;")
        w("  var tacos=gmvTotP>0?cost/gmvTotP*100:0;")
        w("  var pctIngGmv=gmvTotP>0?totAmt/gmvTotP*100:0;")
        w("  var pctVentas=unitsTotP>0?totU/unitsTotP*100:0;")
        w("  // Actualizar meta")
        w("  var _h7=document.getElementById('hdr7-kpi');if(_h7)_h7.textContent=roas.toFixed(2)+'x';")
        w("  var meta=document.getElementById('ads-meta');")
        w("  if(meta)meta.textContent='D\u00edas '+f+'\u2013'+t+' \u00b7 Datos v\u00eda API Product Ads';")
        w("  // KPIs fila única")
        w("  var kg=document.getElementById('ads-kpis');")
        w("  if(kg){")
        w("    kg.style.cssText='display:grid;grid-template-columns:repeat(11,1fr);gap:8px;';")
        w("    var kpis=[")
        w("      {lbl:'Inversi\u00f3n',val:fmtM(cost),col:'var(--purple)'},")
        w("      {lbl:'Ingresos Ads',val:fmtM(totAmt),sub:pctIngGmv.toFixed(1)+'% del GMV',col:'var(--green)'},")
        w("      {lbl:'ROAS',val:roas.toFixed(2)+'x',col:roas>=10?'var(--green)':'var(--yellow)'},")
        w("      {lbl:'ACOS',val:acos.toFixed(1)+'%',sub:'Inv/Ing Ads',col:acos<10?'var(--green)':'var(--yellow)'},")
        w("      {lbl:'TACOS',val:tacos.toFixed(1)+'%',sub:'Inv/GMV total',col:tacos<10?'var(--green)':'var(--yellow)'},")
        w("      {lbl:'CTR',val:ctr.toFixed(2)+'%',col:'var(--blue)'},")
        w("      {lbl:'CPC',val:fmtM(cpc),col:'var(--cyan)'},")
        w("      {lbl:'Clics',val:fmtN(clicks),col:'var(--text)'},")
        w("      {lbl:'Impresiones',val:fmtN(prints),col:'var(--text3)'},")
        w("      {lbl:'Ventas Ads',val:fmtN(totU),sub:pctVentas.toFixed(1)+'% del total',col:'var(--green)'},")
        w("      {lbl:'Dir. / Ind.',val:fmtM(dirAmt)+' / '+fmtM(indAmt),col:'var(--text2)'},")
        w("    ];")
        w("    kg.innerHTML=kpis.map(function(k){")
        w("      var sub=k.sub?'<div class=\"kpi-pri\">'+k.sub+'</div>':'';")
        w("      return '<div class=\"kpi\" ><div class=\"kpi-label\">'+k.lbl+'</div><div class=\"kpi-val\" style=\"color:'+k.col+'\">'+k.val+'</div>'+sub+'</div>';")
        w("    }).join('');")
        w("  }")
        w("  // Gráfico inversión vs ingresos")
        w("  mkChart('c-ads-daily',{type:'bar',data:{labels:days,datasets:[")
        w("    {label:'Inversión',data:dCost,yAxisID:'y',backgroundColor:c.purple+'99',borderRadius:4,order:2},")
        w("    {label:'Ingresos',data:dRev,yAxisID:'y2',type:'line',borderColor:c.green,")
        w("     backgroundColor:'transparent',borderWidth:2,pointRadius:3,pointBackgroundColor:c.green,order:1},")
        w("  ]},options:{responsive:true,maintainAspectRatio:false,")
        w("    plugins:{legend:{labels:{color:c.ctxt,font:{size:11}}}},")
        w("    scales:{")
        w("      x:{grid:{color:c.grid},ticks:{color:c.ctxt}},")
        w("      y:{grid:{color:c.grid},ticks:{color:c.purple,callback:function(v){return fmtM(v);}},")
        w("         title:{display:true,text:'Inversión',color:c.purple}},")
        w("      y2:{position:'right',grid:{display:false},")
        w("          ticks:{color:c.green,callback:function(v){return fmtM(v);}},")
        w("          title:{display:true,text:'Ingresos',color:c.green}},")
        w("    }}});")
        w("  // Gráfico ROAS diario")
        w("  mkChart('c-ads-roas',{type:'line',data:{labels:days,datasets:[")
        w("    {label:'ROAS',data:dRoas,borderColor:c.yellow,backgroundColor:c.yellow+'22',")
        w("     fill:true,borderWidth:2,pointRadius:3,tension:0.3},")
        w("  ]},options:{responsive:true,maintainAspectRatio:false,")
        w("    plugins:{legend:{labels:{color:c.ctxt,font:{size:11}}}},")
        w("    scales:{")
        w("      x:{grid:{color:c.grid},ticks:{color:c.ctxt}},")
        w("      y:{grid:{color:c.grid},ticks:{color:c.yellow,callback:function(v){return v.toFixed(1)+'x';}},")
        w("         title:{display:true,text:'ROAS',color:c.yellow}},")
        w("    }}});")
        w("  // Distribución costos")
        w("  var cc2=document.getElementById('ads-costos');")
        w("  if(cc2&&cost>0){")
        w("    var gmvTot=0;for(var dd=f;dd<=t;dd++){var dv=DAILY_CUR[String(dd)]||{};gmvTot+=dv.gmv||0;}")
        w("    var tacos2=gmvTot>0?cost/gmvTot*100:0;")
        w("    var rows=[")
        w("      {lbl:'Inversión Ads',val:cost,tot:gmvTot,col:'var(--purple)'},")
        w("      {lbl:'Ingresos generados',val:totAmt,tot:gmvTot,col:'var(--green)'},")
        w("      {lbl:'TACOS (Ads/GMV)',val:null,pct:tacos2,col:tacos2<10?'var(--green)':'var(--yellow)',txt:tacos2.toFixed(1)+'%'},")
        w("    ];")
        w("    cc2.innerHTML=rows.map(function(r){")
        w("      var pct=r.pct!=null?r.pct:(r.tot>0?r.val/r.tot*100:0);")
        w("      var txt=r.txt||fmtM(r.val)+' ('+pct.toFixed(1)+'%)';")
        w("      return '<div class=\"pct-row\"><span class=\"pct-lbl\">'+r.lbl+'</span>'")
        w("        +'<div class=\"pct-track\"><div class=\"pct-fill\" style=\"width:'+Math.min(100,Math.round(pct))+'%;background:'+r.col+'\"></div></div>'")
        w("        +'<span class=\"pct-num\" style=\"color:'+r.col+'\">'+txt+'</span></div>';")
        w("    }).join('');")
        w("  }")
        w("}")
        # -- renderLiquidaciones JS v5 — Executive Redesign
        w("function _fM(v){v=Math.round(v||0);if(v>=1e9)return'$'+(v/1e9).toFixed(2)+'B';if(v>=1e6)return'$'+(v/1e6).toFixed(1)+'M';if(v>=1e3)return'$'+(v/1e3).toFixed(0)+'K';return'$'+v.toLocaleString('es-AR');}")
        w("function _pct(v,t){return t>0?(v/t*100).toFixed(1)+'%':'–';}")
        w("function _renderWfPanel(key,col){")
        w("  var p=document.getElementById('wf-panel');")
        w("  if(!p)return;")
        w("  if(!key){p.innerHTML='';return;}")
        w("  var DL=DAILY_LIQ||{};")
        w("  var dates=Object.keys(DL).sort();")
        w("  // Aggregate breakdown across all visible dates")
        w("  var shipBk={},taxBk={},totShip=0,totTax=0,totGmv=0,totFee=0,totDev=0,totNeto=0;")
        w("  dates.forEach(function(d){")
        w("    var r=DL[d];")
        w("    totGmv+=r.gmv||0; totFee+=r.fee_ml||0;")
        w("    totShip+=((r.shipping_real||r.shipping))||0;")
        w("    totTax+=r.taxes_real||0; totDev+=r.devuelto||0; totNeto+=r.neto||0;")
        w("    Object.entries(r.ship_breakdown||{}).forEach(function(e){shipBk[e[0]]=(shipBk[e[0]]||0)+e[1];});")
        w("    Object.entries(r.tax_breakdown||{}).forEach(function(e){taxBk[e[0]]=(taxBk[e[0]]||0)+e[1];});")
        w("  });")
        w("  // Name maps")
        w("  var shipNames={'shp_fulfillment':'Mercado Envíos Full','shp_cross_docking':'Cross-docking','shp_colecta':'Colecta','shp_agency':'Agencia'};")
        w("  var taxNames={'IIBB Córdoba':'IIBB Córdoba','IIBB Buenos Aires':'IIBB Buenos Aires','IDB':'Débitos y Créditos'};")
        w("  function bkRows(data,total,nameMap){")
        w("    var entries=Object.entries(data).sort(function(a,b){return b[1]-a[1];});")
        w("    if(!entries.length) return '<div style=\"color:var(--text3);font-size:12px\">Sin detalle disponible</div>';")
        w("    return entries.map(function(e){")
        w("      var nm=nameMap[e[0]]||e[0].replace(/_/g,' ');")
        w("      var pct=total>0?Math.round(e[1]/total*100):0;")
        w("      return '<div style=\"margin-bottom:8px\">'+")
        w("        '<div style=\"display:flex;justify-content:space-between;margin-bottom:3px\">'+")
        w("          '<span style=\"font-size:11px;color:var(--text2)\">'+nm+'</span>'+")
        w("          '<span style=\"font-size:11px;font-weight:700;color:'+col+'\">'+_fM(e[1])+' <span style=\"color:var(--text3);font-weight:400\">('+pct+'%)</span></span>'+")
        w("        '</div>'+")
        w("        '<div style=\"background:var(--surface2);border-radius:3px;height:6px\">'+")
        w("          '<div style=\"background:'+col+';height:100%;width:'+Math.min(100,pct)+'%;border-radius:3px;transition:width .3s\"></div>'+")
        w("        '</div></div>';")
        w("    }).join('');")
        w("  }")
        w("  function dayRows(getVal,getCol){")
        w("    return '<table style=\"width:100%;border-collapse:collapse;font-size:11px\">'+")
        w("      dates.map(function(d){")
        w("        var r=DL[d];")
        w("        var v=getVal(r); var vc=getCol?getCol(r,v):'var(--text)';")
        w("        return '<tr><td style=\"padding:3px 0;color:var(--text3)\">'+d.slice(8)+'/'+d.slice(5,7)+'</td>'+")
        w("               '<td style=\"text-align:right;font-weight:700;color:'+vc+'\">'+_fM(v)+'</td></tr>';")
        w("      }).join('')+")
        w("    '</table>';")
        w("  }")
        w("  var content='';")
        w("  if(key==='gmv'){")
        w("    var best=dates.reduce(function(a,d){return(DL[d].gmv||0)>(DL[a]&&DL[a].gmv||0)?d:a;},dates[0]||'');")
        w("    content='<div style=\"display:grid;grid-template-columns:1fr 1fr;gap:16px\">'+")
        w("      '<div><div style=\"font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px\">GMV por día</div>'+dayRows(function(r){return r.gmv||0;})+'</div>'+")
        w("      '<div><div style=\"font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px\">Destaques</div>'+")
        w("        '<div style=\"font-size:12px;color:var(--text2)\">🏆 Mejor día: <b style=\"color:'+col+'\">'+best.slice(8)+'/'+best.slice(5,7)+'</b> · '+_fM((DL[best]||{}).gmv||0)+'</div>'+")
        w("        '<div style=\"font-size:12px;color:var(--text2);margin-top:8px\">📦 Promedio diario: <b style=\"color:'+col+'\">'+_fM(dates.length?totGmv/dates.length:0)+'</b></div>'+")
        w("        '<div style=\"font-size:12px;color:var(--text2);margin-top:8px\">🔢 Órdenes total: <b>'+dates.reduce(function(s,d){return s+(DL[d].n_orders||0);},0).toLocaleString('es-AR')+'</b></div>'+")
        w("      '</div>'+")
        w("    '</div>';")
        w("  } else if(key==='fee'){")
        w('    var _pend=[],_est=[];')
        w('    dates.forEach(function(d){')
        w('      var r=DL[d];')
        w('      if((!r.fee_ml||r.fee_ml===0)&&r.gmv>0){')
        w('        if(r.n_fee_derived>0) _est.push(d); else _pend.push(d);')
        w('      }')
        w('    });')
        w("    var _note='';")
        w('    if(_pend.length>0) _note+=\'<div style="margin-top:12px;padding:10px 12px;background:var(--surface2);border-left:3px solid #FF9800;border-radius:6px;font-size:11px;color:var(--text2)">\'+')
        w("      '⚠️ ML reporta comisiones con 5–7 días de retraso. Días en gris: comisión pendiente de informar.</div>';")
        w('    if(_est.length>0) _note+=\'<div style="margin-top:8px;padding:8px 12px;background:var(--surface2);border-left:3px solid #FF9800;border-radius:6px;font-size:11px;color:var(--text2)">\'+')
        w("      '≈ Días con ~ son estimados desde balance (GMV − neto − envíos − retenciones). Se actualizarán cuando ML informe.</div>';")
        w('    var _fR=\'<table style="width:100%;border-collapse:collapse;font-size:11px">\';')
        w('    dates.forEach(function(d){')
        w('      var r=DL[d];')
        w('      var hasReal=r.fee_ml>0&&r.n_fee_derived===0;')
        w('      var hasEst =r.fee_ml>0&&r.n_fee_derived>0;')
        w('      var isPend =(!r.fee_ml||r.fee_ml===0)&&r.gmv>0;')
        w("      var dc=isPend?'#BDBDBD':'var(--text3)';")
        w('      var fv=isPend?\'<span style="color:#BDBDBD">— pend.</span>\'')
        w('          :hasEst?\'<span style="color:#FF9800">≈\'+_fM(r.fee_ml)+\'</span>\'')
        w('          :_fM(r.fee_ml);')
        w('      _fR+=\'<tr><td style="padding:3px 0;color:\'+dc+\'">\'+d.slice(8)+\'/\'+d.slice(5,7)+\'</td>\'')
        w('           +\'<td style="text-align:right;font-weight:700">\'+fv+\'</td></tr>\';')
        w('    });')
        w("    _fR+='</table>';")
        w('    var _pR=\'<table style="width:100%;border-collapse:collapse;font-size:11px">\';')
        w('    dates.forEach(function(d){')
        w('      var r=DL[d];')
        w('      var hasEst=r.fee_ml>0&&r.n_fee_derived>0;')
        w('      var isPend=(!r.fee_ml||r.fee_ml===0)&&r.gmv>0;')
        w('      var fp=r.gmv>0&&!isPend?r.fee_ml/r.gmv*100:0;')
        w("      var fc=isPend?'#BDBDBD':hasEst?'#FF9800':fp>18?'#EF5350':fp>15?'var(--text)':'#00C853';")
        w("      var dc=isPend?'#BDBDBD':'var(--text3)';")
        w('      var pv=isPend?\'<span style="color:#BDBDBD">— pend.</span>\'')
        w('          :hasEst?\'<span style="color:#FF9800">≈\'+fp.toFixed(1)+\'%</span>\'')
        w("          :fp.toFixed(1)+'%';")
        w('      _pR+=\'<tr><td style="padding:3px 0;color:\'+dc+\'">\'+d.slice(8)+\'/\'+d.slice(5,7)+\'</td>\'')
        w('           +\'<td style="text-align:right;font-weight:700;color:\'+fc+\'">\'+pv+\'</td></tr>\';')
        w('    });')
        w("    _pR+='</table>';")
        w('    content=\'<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">\'')
        w('      +\'<div><div style="font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px">Comisiones por día</div>\'+_fR+\'</div>\'')
        w('      +\'<div><div style="font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px">Fee % del GMV por día</div>\'+_pR+\'</div>\'')
        w("      +'</div>'+_note;")
        w("  } else if(key==='ship'){")
        w("    content='<div style=\"display:grid;grid-template-columns:1fr 1fr;gap:16px\">'+")
        w("      '<div><div style=\"font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px\">Desglose por tipo de envío</div>'+")
        w("        bkRows(shipBk,totShip,shipNames)+")
        w("      '</div>'+")
        w("      '<div><div style=\"font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px\">Envíos por día</div>'+")
        w("        dayRows(function(r){return((r.shipping_real||r.shipping))||0;})+'</div>'+")
        w("    '</div>';")
        w("  } else if(key==='tax'){")
        w("    content='<div style=\"display:grid;grid-template-columns:1fr 1fr;gap:16px\">'+")
        w("      '<div><div style=\"font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px\">Desglose de retenciones</div>'+")
        w("        bkRows(taxBk,totTax,taxNames)+")
        w("      '</div>'+")
        w("      '<div><div style=\"font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px\">Retenciones por día</div>'+")
        w("        dayRows(function(r){return r.taxes_real||0;})+'</div>'+")
        w("    '</div>';")
        w("  } else if(key==='dev'){")
        w("    content='<div style=\"display:grid;grid-template-columns:1fr 1fr;gap:16px\">'+")
        w("      '<div><div style=\"font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px\">Devoluciones por día</div>'+dayRows(function(r){return r.devuelto||0;})+'</div>'+")
        w("      '<div><div style=\"font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px\">Destaques</div>'+")
        w("        '<div style=\"font-size:12px;color:var(--text2)\">💸 Total devuelto: <b style=\"color:'+col+'\">'+_fM(totDev)+'</b></div>'+")
        w("        '<div style=\"font-size:12px;color:var(--text2);margin-top:8px\">📉 % sobre GMV: <b style=\"color:'+col+'\">'+_pct(totDev,totGmv)+'</b></div>'+")
        w("      '</div></div>';")
        w("  } else if(key==='neto'){")
        w("    content='<div style=\"display:grid;grid-template-columns:1fr 1fr;gap:16px\">'+")
        w("      '<div><div style=\"font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px\">Neto por día</div>'+dayRows(function(r){return r.neto||0;},function(r,v){var m=r.gmv>0?v/r.gmv*100:0;return m>=70?'#00C853':m>=55?'var(--text)':'#EF5350';})+'</div>'+")
        w("      '<div><div style=\"font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;margin-bottom:8px\">Margen % por día</div>'+")
        w("        '<table style=\"width:100%;border-collapse:collapse;font-size:11px\">'+")
        w("        dates.map(function(d){var r=DL[d];var mg=r.gmv>0?r.neto/r.gmv*100:0;var mc=mg>=70?'#00C853':mg>=55?'var(--text)':'#EF5350';")
        w("          return'<tr><td style=\"padding:3px 0;color:var(--text3)\">'+d.slice(8)+'/'+d.slice(5,7)+'</td><td style=\"text-align:right;font-weight:700;color:'+mc+'\">'+mg.toFixed(1)+'%</td></tr>';")
        w("        }).join('')+'</table>'+")
        w("      '</div></div>';")
        w("  }")
        w("  p.innerHTML='<div style=\"margin-top:-4px;margin-bottom:14px;background:var(--surface);border:1px solid var(--border);border-top:3px solid '+col+';border-radius:12px;padding:16px 18px;animation:fadeIn .2s\">'+")
        w("    '<div style=\"display:flex;justify-content:space-between;align-items:center;margin-bottom:12px\">'+")
        w("      '<div style=\"font-size:10px;font-weight:700;color:'+col+';text-transform:uppercase;letter-spacing:.5px\">🔍 Detalle: '+")
        w("        {gmv:'Total vendido',fee:'Comisiones ML',ship:'Envíos Full/Colecta',tax:'Retenciones',dev:'Devoluciones',neto:'Neto final'}[key]+")
        w("      '</div>'+")
        w("      '<button onclick=\"_renderWfPanel(null);if(_wfActive){_wfActive.style.transform=\\'\\';_wfActive.style.boxShadow=\\'\\';_wfActive.style.outline=\\'\\';_wfActive=null;}\" "+
          "style=\"border:none;background:none;color:var(--text3);cursor:pointer;font-size:16px;line-height:1\">✕</button>'+")
        w("    '</div>'+")
        w("    content+")
        w("  '</div>';")
        w("}")
        w('function _initLiqChart(){')
        w("  var canvas=document.getElementById('liq-cf-chart');")
        w('  if(!canvas||!window.Chart)return;')
        w("  if(_charts['liq-cf']){_charts['liq-cf'].destroy();_charts['liq-cf']=null;}")
        w('  // Only future (pending to release) — past already collected')
        w('  var CF=(FACT_CASHFLOW||[]).filter(function(c){return c.is_future&&c.pending>0;});')
        w('  var cfLabels=CF.map(function(c){var p=c.date.split("-");return p[2]+"/"+p[1];});')
        w('  var cfFuture=CF.map(function(c){return c.pending;});')
        w('  var totalPending=CF.reduce(function(s,c){return s+c.pending;},0);')
        w("  // Show authoritative total from MP_BALANCE; note delta as 'en proceso'")
        w("  var mpTotal=MP_BALANCE?MP_BALANCE.a_cobrar:totalPending;")
        w("  var delta=mpTotal>totalPending?Math.round(mpTotal-totalPending):0;")
        w("  var isDark=document.documentElement.getAttribute('data-theme')==='dark';")
        w("  var gc=isDark?'rgba(255,255,255,0.08)':'rgba(0,0,0,0.06)';")
        w("  var lc=isDark?'#bbb':'#555';")
        w("  var datasets=[")
        w("    {label:'\u23f3 Por liberar',data:cfFuture,backgroundColor:'rgba(52,131,250,0.75)',borderColor:'#3483FA',borderWidth:1,borderRadius:4}];")
        w("  var lbl=document.getElementById('cf-total-lbl');")
        w("  var deltaNote=delta>0?' · +'+_fM(delta)+' en proceso':'';")
        w("  if(lbl)lbl.textContent='\u00b7 '+_fM(mpTotal)+' en '+CF.length+' fechas'+deltaNote;")
        w("  _charts['liq-cf']=new Chart(canvas,{")
        w("    type:'bar',")
        w('    data:{labels:cfLabels,datasets:datasets},')
        w('    options:{')
        w('      responsive:true,')
        w('      maintainAspectRatio:false,')
        w('      plugins:{')
        w("        legend:{display:false},")
        w("        tooltip:{callbacks:{label:function(ctx){return _fM(ctx.raw);},title:function(items){return items[0].label==='Sin fecha'?'Sin fecha de liberación asignada':'Libera el '+items[0].label;}}},")
        w('      },')
        w('      scales:{')
        w('        x:{ticks:{color:lc,font:{size:10},maxRotation:45},grid:{color:gc}},')
        w('        y:{ticks:{color:lc,font:{size:10},callback:function(v){return _fM(v);}},grid:{color:gc},beginAtZero:true}')
        w('      },')
        w('    }')
        w('  });')
        w('}')
        w('')
        w("var _wfActive=null;")
        w("function _wfHover(el,col,on){")
        w("  if(_wfActive===el)return;")
        w("  el.style.transform=on?'translateY(-5px)':'';")
        w("  el.style.boxShadow=on?'0 10px 28px '+col+'44':'';")
        w("  el.style.borderColor=on?col:'';")
        w("  el.style.position=on?'relative':'';")
        w("  el.style.zIndex=on?'10':'';")
        w("}")
        w("function _wfClick(el,key,col){")
        w("  if(_wfActive&&_wfActive!==el){")
        w("    _wfActive.style.transform='';_wfActive.style.boxShadow='';")
        w("    _wfActive.style.outline='';_wfActive.style.borderColor='';")
        w("  }")
        w("  var isDesel=_wfActive===el;")
        w("  _wfActive=isDesel?null:el;")
        w("  if(!isDesel){")
        w("    el.style.transform='translateY(-5px)';")
        w("    el.style.boxShadow='0 10px 28px '+col+'55';")
        w("    el.style.outline='2px solid '+col;")
        w("  } else {")
        w("    el.style.transform='';el.style.boxShadow='';el.style.outline='';el.style.borderColor='';")
        w("  }")
        w("  _renderWfPanel(isDesel?null:key,col);")
        w("}")
        w('function renderLiquidaciones(dayFrom,dayTo){')
        w("  var root=document.getElementById('liq-root');")
        w('  if(!root)return;')
        w('  var f=dayFrom||1,t=dayTo||MAX_DAY;')
        w('  var S=LIQ_SUMMARY||{},P=LIQ_PRIOR||{},DL=DAILY_LIQ||{};')
        w('  var CF=FACT_CASHFLOW||[];')
        w('  // ── Aggregate filtered daily data ─────────────────────────────────────')
        w('  var totGmv=0,totFee=0,totShip=0,totTax=0,totDev=0,totNeto=0,totOrders=0;')
        w('  var allDates=Object.keys(DL).sort();')
        w("  var mainMonth=allDates.length?allDates[allDates.length-1].slice(0,7):'';")
        w('  var filtDates=allDates.filter(function(d){')
        w("    if(d.slice(0,7)!==mainMonth)return false;")
        w('    var day=parseInt(d.slice(-2),10);')
        w('    return day>=f&&day<=t;')
        w('  });')
        w('  filtDates.forEach(function(d){')
        w('    var r=DL[d];')
        w('    totGmv+=r.gmv||0;totFee+=r.fee_ml||0;')
        w('    totShip+=((r.shipping_real||r.shipping))||0;')
        w('    totTax+=r.taxes_real||0;')
        w('    totDev+=r.devuelto||0;totNeto+=r.neto||0;totOrders+=r.n_orders||0;')
        w('  });')
        w('  // Si taxes_real=0 per-orden (charges vacíos), usar retenciones_real.json (payments/search)')
        w('  var _retEsReal=false;')
        w('  if(totTax===0&&S.retenciones_total&&S.retenciones_total>0){')
        w("    totTax=S.retenciones_total; _retEsReal=S.retenciones_es_real!==false;")
        w('  }')
        w('  // Fallback final: derivar del residual (estimado)')
        w("  if(totTax===0&&totNeto>0&&totGmv>0){var _dt=totGmv-totFee-totShip-totDev-totNeto;if(_dt>0){totTax=_dt;_retEsReal=false;}}")
        w('  // Si shipping DL es estimado, usar S.shipping_total real (payments/search)')
        w('  var _shipEsReal=false;')
        w('  if(S.envios_es_real&&S.shipping_total&&S.shipping_total>0){totShip=S.shipping_total;_shipEsReal=true;}')
        w("  function fP(v){return v.toFixed(1)+'%';}")
        w("  function fO(v){return(v||0).toLocaleString('es-AR');}")
        w('  var feeRate=totGmv>0?totFee/totGmv*100:0;')
        w('  var shipRate=totGmv>0?totShip/totGmv*100:0;')
        w('  var taxRate=totGmv>0?totTax/totGmv*100:0;')
        w('  var devRate=totGmv>0?totDev/totGmv*100:0;')
        w('  var margen=totGmv>0?totNeto/totGmv*100:0;')
        w('  var pGmv=P.gmv||0,pNeto=P.neto||0,pFeeRate=P.fee_rate||0;')
        w('  // ── Delta helper ──────────────────────────────────────────────────────')
        w('  function dlt(cur,pri,inv,isFeeRate){')
        w('    if(!pri||!cur)return \'\';')
        w('    var p=(cur-pri)/Math.abs(pri)*100;')
        w("    var col=inv?(p>0?'#EF5350':'#00C853'):(p>0?'#00C853':'#EF5350');")
        w("    var sym=p>0?'▲':'▼';")
        w("    var label=isFeeRate?Math.abs(p).toFixed(1)+'pp':Math.abs(p).toFixed(1)+'%';")
        w("    return '<span style=\"font-size:10px;color:'+col+'\">'+sym+' '+label+' vs ant.</span>';")
        w('  }')
        w("  var dltGmv=dlt(totGmv,pGmv,false,false);")
        w("  var dltNeto=dlt(totNeto,pNeto,false,false);")
        w("  var dltFee=dlt(feeRate,pFeeRate,true,true);")
        w('  // ── Cash flow state ───────────────────────────────────────────────────')
        w('  var libEst=S.liberado_estimado||0,pendReal=S.pendiente_real||0;')
        w('  var futureRel=CF.filter(function(c){return c.is_future&&c.pending>0;});')
        w("  var periodLabel=filtDates.length?filtDates[0].slice(8)+'/'+filtDates[0].slice(5,7)+' – '+filtDates[filtDates.length-1].slice(8)+'/'+filtDates[filtDates.length-1].slice(5,7):'–';")
        w("  var updAt=(S.updated_at||'').slice(0,16).replace('T',' ');")
        w("  function wfCard(icon,lbl,val,col,sub,d,isRes,key){")
        w("    var bd=isRes?'border:2px solid '+col+';':'border:1px solid var(--border);border-top:3px solid '+col+';';")
        w("    var clickAttr=key?'onclick=\"_wfClick(this,\\''+key+'\\',\\''+col+'\\')\" '")
        w("                    +'onmouseenter=\"_wfHover(this,\\''+col+'\\',true)\" '")
        w("                    +'onmouseleave=\"_wfHover(this,\\''+col+'\\',false)\" '")
        w("                    +'style=\"flex:1;min-width:130px;background:var(--surface);'+bd+'border-radius:12px;padding:16px 14px;text-align:center;cursor:pointer;transition:transform .18s,box-shadow .18s,border-color .18s\"'")
        w("                   :'style=\"flex:1;min-width:130px;background:var(--surface);'+bd+'border-radius:12px;padding:16px 14px;text-align:center\"';")
        w("    return '<div '+clickAttr+'>'+")
        w("      '<div style=\"font-size:20px;margin-bottom:4px\">'+icon+'</div>'+")
        w("      '<div style=\"font-size:9px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px\">'+lbl+'</div>'+")
        w("      '<div style=\"font-size:22px;font-weight:800;color:'+col+';line-height:1.1\">'+val+'</div>'+")
        w("      '<div style=\"font-size:10px;color:var(--text3);margin-top:5px\">'+sub+'</div>'+")
        w("      (d?'<div style=\"margin-top:5px\">'+d+'</div>':'')+'</div>';")
        w('  }')
        w("  function arrow(s){return '<div style=\"display:flex;align-items:center;justify-content:center;font-size:18px;color:var(--text3);padding:0 2px;align-self:center\">'+s+'</div>';}")
        w("  function distBar(lbl,val,pct,col){")
        w("    return '<div style=\"display:flex;align-items:center;gap:8px;padding:7px 0;border-bottom:1px solid var(--border)\">'+")
        w("      '<div style=\"font-size:11px;color:var(--text2);width:120px;flex-shrink:0\">'+lbl+'</div>'+")
        w("      '<div style=\"flex:1;background:var(--surface2);border-radius:3px;height:9px;overflow:hidden\">'+")
        w("        '<div style=\"background:'+col+';height:100%;width:'+pct.toFixed(1)+'%;border-radius:3px\"></div>'+")
        w("      '</div>'+")
        w("      '<div style=\"font-size:11px;font-weight:700;color:'+col+';width:60px;text-align:right\">'+_fM(val)+'</div>'+")
        w("      '<div style=\"font-size:10px;color:var(--text3);width:40px;text-align:right\">'+pct.toFixed(1)+'%</div>'+")
        w("    '</div>';")
        w('  }')
        w('  // ── Cobro cards (replace daily table) ────────────────────────────────')
        w("  var CB=LIQ_COBRO||{};")
        w("  var cbPend=(MP_BALANCE&&MP_BALANCE.a_cobrar)||CB.pendiente||0,cbCob=CB.cobrado||S.liberado_estimado||0,cbTrans=CB.transito||0,cbPlazo=CB.plazo_promedio||0;")
        w("  var cbTot=cbCob+cbPend+cbTrans;")
        w("  function cobCard(icon,lbl,val,col,sub){")
        w("    return '<div style=\"flex:1;min-width:160px;background:var(--surface);border:1px solid var(--border);border-top:3px solid '+col+';border-radius:12px;padding:18px 16px;text-align:center\">'+")
        w("      '<div style=\"font-size:22px;margin-bottom:6px\">'+icon+'</div>'+")
        w("      '<div style=\"font-size:9px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.6px;margin-bottom:10px\">'+lbl+'</div>'+")
        w("      '<div style=\"font-size:26px;font-weight:800;color:'+col+';line-height:1.1\">'+val+'</div>'+")
        w("      '<div style=\"font-size:10px;color:var(--text3);margin-top:6px\">'+sub+'</div>'+")
        w("    '</div>';")
        w("  }")
        w('  // ── Render root ──────────────────────────────────────────────────────')
        w("  var _liqPb=document.querySelector('.period-bar');")
        w("  if(_liqPb&&root.contains(_liqPb)){document.getElementById('tab10').insertBefore(_liqPb,root);}")
        w("  root.innerHTML=")
        # ── Header (Option C — tab-hdr connects to period-bar, not .alone) ──
        w("    '<div class=\"tab-hdr\">'+")
        w("      '<div class=\"tab-hdr-top\">'+")
        w("        '<div class=\"tab-hdr-left\">'+")
        w("          '<span class=\"tab-hdr-icon\">💰</span>'+")
        w("          '<div>'+")
        w("            '<div class=\"tab-hdr-title\">Liquidaciones Mercado Pago</div>'+")
        w("            '<div class=\"tab-hdr-meta\">Días '+f+'–'+t+' · '+fO(totOrders)+' órdenes'+(updAt?' · actualizado '+updAt:'')+'</div>'+")
        w("          '</div>'+")
        w("        '</div>'+")
        w("        '<div>'+")
        w("          '<div class=\"tab-hdr-kpi-val\" style=\"color:#00C853\">'+_fM(totNeto)+'</div>'+")
        w("          '<div class=\"tab-hdr-kpi-lbl\">neto del período</div>'+")
        w("        '</div>'+")
        w("      '</div>'+")
        w("    '</div>'+")
        # ── Waterfall cards ──
        w("    '<div class=\"card\" style=\"margin-bottom:14px\">'+")
        w("      '<div class=\"card-title\">RESUMEN FINANCIERO DEL PERÍODO</div>'+")
        w("      '<div class=\"card-body\">'+")
        w("        '<div style=\"display:flex;flex-wrap:wrap;gap:6px;align-items:stretch\">'+")
        w("          wfCard('💰','Total vendido',_fM(totGmv),'#3483FA','GMV bruto del período',dltGmv,false,'gmv')+")
        w("          arrow('→')+")
        w("          wfCard('📋','Comisiones ML','−'+_fM(totFee),'#FF7043',fP(feeRate)+' del GMV',dltFee,false,'fee')+")
        w("          arrow('+')+")
        w("          wfCard('🚚','Envíos Full/Colecta','−'+_fM(totShip),'#78909C',fP(shipRate)+' del GMV',(_shipEsReal?'Full + Cross-docking (real)':'estimado'),false,'ship')+")
        w("          arrow('+')+")
        w("          wfCard('🏛','Retenciones','−'+_fM(totTax),'#9C27B0',fP(taxRate)+' del GMV',_retEsReal?'IDB + IIBB (real)':'⚠ estimado — IIBB + IDB',false,'tax')+")
        w("          arrow('+')+")
        w("          wfCard('↩','Devoluciones','−'+_fM(totDev),'#EF5350',fP(devRate)+' del GMV','',false,'dev')+")
        w("          arrow('=')+")
        w("          wfCard('✅','Neto final',_fM(totNeto),'#00C853',fP(margen)+' del GMV',dltNeto,true,'neto')+")
        w("        '</div>'+")
        w("      '</div>'+")
        w("      '<div id=\"wf-panel\"></div>'+")
        w("    '</div>'+")
        # ── Cash flow state cards ──
        w("    '<div class=\"liq-cash-grid\" style=\"display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:14px\">'+")
        # Card 1: Dinero en cuenta
        w("      '<div class=\"liq-kpi\" style=\"background:linear-gradient(135deg,rgba(76,175,80,.12),var(--surface));border:1px solid rgba(76,175,80,.35);border-radius:14px;padding:18px 20px;text-align:center\">'+")
        w("        '<div style=\"font-size:10px;font-weight:700;color:#4CAF50;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px\">💵 Dinero en cuenta</div>'+")
        w("        '<div style=\"font-size:30px;font-weight:800;color:#4CAF50;line-height:1.1\">'+(MP_BALANCE?_fM(MP_BALANCE.disponible):'-')+'</div>'+")
        w("        '<div style=\"font-size:10px;color:var(--text3);margin-top:6px\">'+(MP_BALANCE?'Saldo MP · '+MP_BALANCE.snapshot_at:'')+'</div>'+")
        w("      '</div>'+")
        # Card 2: Dinero a cobrar
        w("      '<div class=\"liq-kpi\" style=\"background:linear-gradient(135deg,rgba(0,200,83,.12),var(--surface));border:1px solid rgba(0,200,83,.3);border-radius:14px;padding:18px 20px;text-align:center\">'+")
        w("        '<div style=\"font-size:10px;font-weight:700;color:#00C853;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px\">💰 Dinero a cobrar</div>'+")
        w("        '<div style=\"font-size:30px;font-weight:800;color:#00C853;line-height:1.1\">'+(MP_BALANCE?_fM(MP_BALANCE.a_cobrar):_fM(libEst))+'</div>'+")
        w("        '<div style=\"font-size:10px;color:var(--text3);margin-top:6px\">'+(MP_BALANCE?'Todos los períodos':'Estimado del período')+'</div>'+")
        w("      '</div>'+")
        # Card 3: Dinero retenido (resta)
        w("      '<div class=\"liq-kpi\" style=\"background:linear-gradient(135deg,rgba(239,83,80,.1),var(--surface));border:1px solid rgba(239,83,80,.3);border-radius:14px;padding:18px 20px;text-align:center\">'+")
        w("        '<div style=\"font-size:10px;font-weight:700;color:#EF5350;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px\">🔒 Dinero retenido</div>'+")
        w("        '<div style=\"font-size:30px;font-weight:800;color:#EF5350;line-height:1.1\">'+(MP_BALANCE?_fM(MP_BALANCE.retenido):'-')+'</div>'+")
        w("        '<div style=\"font-size:10px;color:var(--text3);margin-top:6px\">Resta del total disponible</div>'+")
        w("      '</div>'+")
        # Card 4: Total disponible = cuenta + cobrar - retenido
        w("      '<div class=\"liq-kpi\" style=\"background:linear-gradient(135deg,rgba(52,131,250,.1),var(--surface));border:1px solid rgba(52,131,250,.3);border-radius:14px;padding:18px 20px;text-align:center\">'+")
        w("        '<div style=\"font-size:10px;font-weight:700;color:#3483FA;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px\">🏦 Total disponible</div>'+")
        w("        '<div style=\"font-size:30px;font-weight:800;color:#3483FA;line-height:1.1\">'+(MP_BALANCE?_fM(MP_BALANCE.disponible+MP_BALANCE.a_cobrar-MP_BALANCE.retenido):'-')+'</div>'+")
        w("        '<div style=\"font-size:10px;color:var(--text3);margin-top:6px\">Cuenta + A cobrar − Retenido</div>'+")
        w("      '</div>'+")
        w("    '</div>'+")
        # ── Chart + Distribution ──
        w("    '<div class=\"liq-charts-grid\" style=\"display:grid;grid-template-columns:3fr 2fr;gap:12px;margin-bottom:14px\">'+")
        w("      '<div class=\"card\">'+")
        w("        '<div class=\"card-title\">PRÓXIMAS LIBERACIONES <span id=\"cf-total-lbl\" style=\"font-size:10px;font-weight:400;color:var(--text3)\"></span></div>'+")
        w("        '<div class=\"card-body\">'+")
        w("          '<div style=\"position:relative;height:210px\"><canvas id=\"liq-cf-chart\"></canvas></div>'+")
        w("        '</div>'+")
        w("      '</div>'+")
        w("      '<div class=\"card\">'+")
        w("        '<div class=\"card-title\">DISTRIBUCIÓN DE COSTOS</div>'+")
        w("        '<div class=\"card-body\" style=\"padding-top:10px\">'+")
        w("          distBar('💰 Ventas brutas',totGmv,100,'#3483FA')+")
        w("          distBar('📋 Comisiones ML',totFee,feeRate,'#FF7043')+")
        w("          distBar((_shipEsReal?'🚚 Envíos Full/Colecta':'🚚 Envíos (est.)'),totShip,shipRate,'#78909C')+")
        w("          distBar('🏛 Retenciones',totTax,taxRate,'#9C27B0')+")
        w("          distBar('↩ Devoluciones',totDev,devRate,'#EF5350')+")
        w("          distBar('✅ Neto final',totNeto,margen,'#00C853')+")
        w("        '</div>'+")
        w("      '</div>'+")
        w("    '</div>'+")
        # ── Cobro state cards ──
        w("    '<div class=\"card\" style=\"margin-bottom:14px\">'+")
        w("      '<div class=\"card-title\">ESTADO DE COBROS <span style=\"font-size:10px;font-weight:400;color:var(--text3)\">· saldo MP autoritativo · actualiza cada día con el pipeline</span></div>'+")
        w("      '<div class=\"card-body\">'+")
        w("        '<div style=\"display:flex;flex-wrap:wrap;gap:12px;justify-content:stretch\">'+")
        w("          cobCard('✅','Cobrado (liberado)',_fM(cbCob),'#00C853','Neto liberado en el período')+")
        w("          cobCard('⏳','Pendiente de cobro',_fM(cbPend),'#3483FA','Entregado + en tránsito · todos los períodos')+")
        w("          cobCard('📅','Plazo promedio real',cbPlazo+' días','#9C27B0','Días corridos venta→cobro · ponderado por $')+")
        w("        '</div>'+")
        w("      '</div>'+")
        w("    '</div>';")
        w("  var _liqHdr=document.querySelector('#tab10 .tab-hdr');")
        w("  if(_liqPb&&_liqHdr&&_liqPb.parentElement!==_liqHdr){_liqHdr.appendChild(_liqPb);}")
        w('  setTimeout(_initLiqChart,80);')
        w('}')
        w('')

        w("function renderPostventa(dayFrom,dayTo){")
        w("  var f=dayFrom||1,t=dayTo||MAX_DAY;")
        w("  var c=PV_CUR||{},p=PV_PRI||{};")
        w("  var cc=CANCELLED_PROC||{};")
        w("  var ccCur=cc.current||{},ccPri=cc.prior||{};")
        w("  var canc=ccCur.total||0;")
        w("  var cancCodes=ccCur.by_code||{};")
        w("  var cancTot=Object.values(cancCodes).reduce(function(s,v){return s+(v.n||0);},0)||1;")
        w("  var actors=ccCur.by_actor||{};")
        w("  var ca=typeof CLAIMS_ANALYSIS!=='undefined'?CLAIMS_ANALYSIS:{};")
        w("  var caS=ca.summary||{};")
        w("  // ── KPI Cards: siempre período completo ──")
        w("  var nBuyer=(cancCodes.buyer_cancel_express||{}).n||0;")
        w("  var nPack=(cancCodes.pack_splitted||{}).n||0;")
        w("  var nMedC=(cancCodes.mediations||{}).n||0;")
        w("  var nLog=((cancCodes.shipment_not_delivered||{}).n||0)+((cancCodes.shipment_unfulfilled||{}).n||0);")
        w("  var nOther=canc-nBuyer-nPack-nMedC-nLog;")
        w("  var pillsEl=document.getElementById('pv-pills-wrap');")
        w("  if(pillsEl){")
        w("    function _kpiCard(n,lbl,pct,color,bg,border){")
        w("      return '<div class=\"pv-card\" style=\"padding:14px 16px;border-radius:10px;background:'+bg+';border:1px solid '+border+';min-width:100px;flex:1\">'+")
        w("        '<div style=\"font-size:26px;font-weight:800;color:'+color+';line-height:1\">'+n+'</div>'+")
        w("        '<div style=\"font-size:11px;color:var(--text2);margin-top:4px;font-weight:600\">'+lbl+'</div>'+")
        w("        (pct?'<div style=\"font-size:10px;color:'+color+';margin-top:2px;font-weight:700\">'+pct+'</div>':'')+'</div>';")
        w("    }")
        w("    var totalOrds=(PAID_CUR||0)+canc;")
        w("    var pBuyer=totalOrds>0?(nBuyer/totalOrds*100).toFixed(1)+'%':'';")
        w("    var pPack=totalOrds>0?(nPack/totalOrds*100).toFixed(1)+'%':'';")
        w("    var pMedC=totalOrds>0?(nMedC/totalOrds*100).toFixed(1)+'%':'';")
        w("    var pLog=totalOrds>0?(nLog/totalOrds*100).toFixed(1)+'%':'';")
        w("    var pOther=totalOrds>0?(nOther/totalOrds*100).toFixed(1)+'%':'';")
        w("    pillsEl.innerHTML=")
        w("      '<div style=\"display:flex;gap:10px;flex-wrap:wrap;align-items:stretch\">'+")
        w("      _kpiCard(canc,'📅 Total Mayo','','var(--purple)','rgba(156,39,176,.08)','rgba(156,39,176,.25)')+''+")
        w("      _kpiCard(nBuyer,'🛒 Comprador',pBuyer,'#3483FA','rgba(52,131,250,.08)','rgba(52,131,250,.25)')+''+")
        w("      _kpiCard(nPack,'📦 Pack ML',pPack,'var(--purple)','rgba(156,39,176,.06)','rgba(156,39,176,.2)')+''+")
        w("      _kpiCard(nMedC,'⚖️ Por reclamo',pMedC,'var(--red)','rgba(239,68,68,.08)','rgba(239,68,68,.25)')+''+")
        w("      _kpiCard(nLog,'🚚 Logística',pLog,'var(--orange)','rgba(251,146,60,.08)','rgba(251,146,60,.25)')+''+")
        w("      _kpiCard(nOther,'❓ Otros',pOther,'var(--text3)','rgba(100,100,100,.06)','rgba(100,100,100,.15)')+''+")
        w("      '</div>';")
        w("  }")

        # ── Cancelaciones por motivo (50/50 unified rows) ──
        w("  var wrapEl=document.getElementById('pv-cancel-wrap');")
        w("  if(wrapEl){")
        w("    var codeMap={'buyer_cancel_express':'🛒 Comprador canceló','pack_splitted':'📦 División de pack (ML)','mediations':'⚖️ Cancelado por reclamo','shipment_not_delivered':'🚚 No entregado por logística','fraud':'🚫 Fraude detectado','shipment_unfulfilled':'📭 Envío no cumplido','unknown':'❓ Sin especificar'};")
        w("    var accentMap={'buyer_cancel_express':'var(--blue)','pack_splitted':'var(--purple)','mediations':'var(--red)','shipment_not_delivered':'var(--orange)','fraud':'#888','shipment_unfulfilled':'var(--yellow)','unknown':'var(--text3)'};")
        w("    var glossDefs={")
        w("      buyer_cancel_express:'Comprador canceló antes del despacho. No penaliza reputación. Proceso automático.',")
        w("      pack_splitted:'ML cancela cuando no puede enviar el pack completo. Decisión de ML, no del vendedor.',")
        w("      mediations:'Orden cancelada por reclamo activo del comprador. Sí impacta reputación.',")
        w("      shipment_not_delivered:'Paquete despachado pero el carrier no pudo entregarlo.',")
        w("      shipment_unfulfilled:'Vendedor no despachó a tiempo. ML cancela y penaliza reputación.',")
        w("      fraud:'ML detectó fraude del comprador. Cancela preventivamente, no penaliza al vendedor.',")
        w("      unknown:'Sin motivo registrado en la API.'")
        w("    };")
        w("    var actorHtml='<div style=\"display:flex;gap:10px;margin-bottom:10px;font-size:11px\">'")
        w("      +'<span style=\"padding:3px 10px;border-radius:20px;background:rgba(52,131,250,.12);color:#3483FA\">🛒 Comprador: '+(actors.buyer||0)+'</span>'")
        w("      +'<span style=\"padding:3px 10px;border-radius:20px;background:rgba(156,39,176,.12);color:#9C27B0\">🤖 ML: '+(actors.meli||0)+'</span></div>';")
        w("    var glHeader='<div style=\"font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--text3);padding-bottom:4px\">Glosario</div>';")
        w("    var rowsHtml=Object.entries(cancCodes).sort(function(a,b){return b[1].n-a[1].n;}).map(function(e,idx){")
        w("      var pct=cancTot>0?Math.round(e[1].n/cancTot*100):0;")
        w("      var lbl=codeMap[e[0]]||e[0];")
        w("      var acc=accentMap[e[0]]||'var(--text3)';")
        w("      var gmvM=(e[1].gmv||0)>=1e6?'$'+((e[1].gmv||0)/1e6).toFixed(1)+'M':'$'+Math.round((e[1].gmv||0)/1e3)+'K';")
        w("      var desc=glossDefs[e[0]]||'';")
        w("      return '<div style=\"display:flex;align-items:center;min-height:28px;border-bottom:1px solid var(--border);padding:2px 0\">'")
        w("        +'<div style=\"flex:1;padding-right:12px\">'")
        w("        +'<div class=\"pct-row\" style=\"margin:0\"><span class=\"pct-lbl\" style=\"font-size:11px\">'+lbl+'</span>'")
        w("        +'<div class=\"pct-track\"><div class=\"pct-fill\" style=\"width:'+pct+'%;background:'+acc+'\"></div></div>'")
        w("        +'<span class=\"pct-num\">'+e[1].n+' · '+gmvM+' ('+pct+'%)</span></div></div>'")
        w("        +'<div style=\"width:1px;background:var(--border);align-self:stretch;margin:2px 0\"></div>'")
        w("        +'<div style=\"flex:1;padding-left:12px\">'+(idx===0?glHeader:'')")
        w("        +'<span style=\"font-size:10px;color:var(--text3)\">'+desc+'</span></div></div>';")
        w("    }).join('');")
        w("    wrapEl.innerHTML=actorHtml+rowsHtml;")
        w("  }")

        # ── Top 10 productos por mediaciones (excluye pre-despacho) ──────────
        w("  var tbody=document.getElementById('pv-top10-tbody');")
        w("  var activosBySku=(typeof CLAIMS_ANALYSIS!=='undefined'&&CLAIMS_ANALYSIS.activos_by_sku)||{};")
        w("  if(tbody){")        
        w("    var topRaw=(c.top_dev&&c.top_dev.length?c.top_dev:c.top_med||[]);topRaw.sort(function(a,b){var rA=a.paid>0?(a.dev/a.paid):0,rB=b.paid>0?(b.dev/b.paid):0;return rB-rA;});var topMed=topRaw.slice(0,10);")
        w("    tbody.innerHTML=topMed.map(function(x,i){")
        w("      var rM=x.paid>0?((x.med/x.paid)*100).toFixed(1):'0.0';")
        w("      var rD=x.paid>0?((x.dev/x.paid)*100).toFixed(1):'0.0';")
        w("      var cM=parseFloat(rM)>=10?'var(--red)':parseFloat(rM)>=5?'var(--orange)':'var(--green)';")
        w("      var rDn=parseFloat(rD);var cD=rDn>=10?'var(--red)':rDn>=5?'var(--orange)':'var(--text2)';")
        w("      var activos=activosBySku[x.id]||0;")
        w("      var cA=activos>0?'var(--red)':'var(--text3)';")
        w("      var gRisk=(x.gmv_med||0)+(x.gmv_dev||0);")
        w("      var mlUrl='https://articulo.mercadolibre.com.ar/'+x.id;")
        w("      var skuR=(typeof SKU_CANCEL_REASONS!=='undefined'&&SKU_CANCEL_REASONS[x.id])||{};")
        w("      var pills=[];")
        w("      if(skuR.buyer>0)pills.push('\U0001f6d2 Comprador: <b>'+skuR.buyer+'</b>');")
        w("      if(skuR.med>0)pills.push('\u2696\ufe0f Por reclamo: <b>'+skuR.med+'</b>');")
        w("      if(skuR.pack>0)pills.push('\U0001f4e6 Pack dividido: <b>'+skuR.pack+'</b>');")
        w("      if(skuR.noent>0)pills.push('\U0001f69a No entregado: <b>'+skuR.noent+'</b>');")
        w("      if(skuR.fraud>0)pills.push('\U0001f6ab Fraude: <b>'+skuR.fraud+'</b>');")
        w("      if(skuR.other>0)pills.push('\u2753 Otro: <b>'+skuR.other+'</b>');")
        w("      var pillsHtml=pills.map(function(p){return '<span style=\"display:inline-flex;gap:3px;margin:2px 4px 2px 0;padding:2px 8px;background:rgba(52,131,250,.08);border-radius:4px;font-size:11px;color:var(--text2)\">'+p+'</span>';}).join('');")
        w("      var hasDet=pills.length>0||activos>0;")
        w("      var detBody='<div style=\"font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.4px;margin-bottom:5px\">Cancelaciones del mes</div>'+(pillsHtml||'<span style=\"font-size:11px;color:var(--text3)\">Sin cancelaciones este mes</span>');")
        w("      if(activos>0)detBody+='<div style=\"margin-top:6px;font-size:11px\"><span style=\"color:var(--red);font-weight:700\">\U0001f534 '+activos+' reclamo(s) activo(s)</span> <a href=\"https://www.mercadolibre.com.ar/reclamaciones\" target=\"_blank\" style=\"color:var(--blue)\">Ver en ML</a></div>';")
        w("      var detRow=hasDet?'<tr class=\"pvdet\" style=\"display:none\"><td></td><td colspan=\"4\" style=\"padding:7px 10px 10px 4px;background:rgba(0,0,0,.025);border-bottom:1px solid var(--border)\">'+detBody+'</td></tr>':'';")
        w("      return '<tr'+(hasDet?' style=\"cursor:pointer\" onclick=\"pvToggle(this)\"':'')+'>'")
        w("        +'<td style=\"color:var(--text3);text-align:center;font-size:11px\">'+(i+1)+'</td>'")
        w("        +'<td style=\"font-size:12px\"><a href=\"'+mlUrl+'\" target=\"_blank\" style=\"color:var(--text1);text-decoration:none;font-weight:600\" onclick=\"event.stopPropagation()\">'+x.title+'</a>'")
        w("        +(rDn>=10?'<span style=\"font-size:9px;font-weight:700;background:rgba(239,68,68,.12);color:var(--red);border:1px solid rgba(239,68,68,.3);border-radius:3px;padding:1px 5px;margin-left:5px\">CR\u00cdTICO</span>'")
        w("        :rDn>=5?'<span style=\"font-size:9px;font-weight:700;background:rgba(251,140,0,.12);color:var(--orange);border:1px solid rgba(251,140,0,.3);border-radius:3px;padding:1px 5px;margin-left:5px\">ATENCI\u00d3N</span>':'')")
        w("        +(hasDet?'<span class=\"pvch\" style=\"font-size:10px;color:var(--text3);margin-left:5px\">\u25b6</span>':'')+'</td>'")
        w("        +'<td style=\"text-align:center;color:var(--text2)\">'+fmtN(x.paid)+'</td>'")
        w("        +'<td style=\"color:var(--orange);text-align:center\">'+x.dev+'</td>'")
        w("        +'<td style=\"color:'+cD+';font-weight:700;text-align:center\">'+rD+'%</td>'")
        w("        +'<td style=\"color:'+cA+';font-weight:700;text-align:center\">'+(activos>0?activos+' \U0001F534':'\u2014')+'</td>'")
        w("        +'</tr>'+detRow;")
        w("    }).join('');")
        w("  }")
        w("}")
        w("function pvToggle(r){var n=r.nextElementSibling;if(n&&n.classList.contains('pvdet')){n.style.display=n.style.display==='none'?'table-row':'none';var c=r.querySelector('span.pvch');if(c)c.textContent=n.style.display!=='none'?'\u25bc':'\u25b6';}}")
        w("if(typeof showTab==='function')showTab(0);")
        w("")
        w("</script>")
        