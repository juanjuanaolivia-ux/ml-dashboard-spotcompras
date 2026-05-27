"""
ml_data.py — Fetchers de datos desde la API de Mercado Libre
Cubre: ventas, stock (Flex/Full), comisiones, envíos, reclamos, devoluciones
"""

import json
import os
from datetime import datetime, timedelta, timezone
from typing import Optional

from ml_auth import MLSession

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
DATA_DIR    = os.path.join(SCRIPT_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _save_json(name: str, data):
    """Guardado atómico: escribe a .tmp y renombra — evita truncación."""
    path = os.path.join(DATA_DIR, f"{name}.json")
    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, path)
    print(f"  💾 {name}.json guardado ({len(data) if isinstance(data, list) else 1} registros)")
    return path


def _date_range(days_back: int = 30) -> tuple[str, str]:
    """Devuelve (from_date, to_date) en formato ISO 8601 para el filtro de la API."""
    tz    = timezone.utc
    end   = datetime.now(tz)
    start = end - timedelta(days=days_back)
    fmt   = "%Y-%m-%dT%H:%M:%S.000-00:00"
    return start.strftime(fmt), end.strftime(fmt)


# ─────────────────────────────────────────────────────────────────────────────
# 1. VENTAS Y FACTURACIÓN
# ─────────────────────────────────────────────────────────────────────────────

def fetch_orders(session: MLSession, days_back: int = 30) -> list:
    """
    Trae todas las órdenes (ventas) del período.
    Incluye: fecha, GMV, unidades, estado, item, comprador, envío.
    """
    print(f"\n  📦 Fetching órdenes (últimos {days_back} días)...")
    date_from, date_to = _date_range(days_back)

    orders = session.get_paginated(
        f"/orders/search",
        params={
            "seller": session.user_id,
            "sort":   "date_desc",
            "order.date_created.from": date_from,
            "order.date_created.to":   date_to,
        },
        limit=50
    )

    # Enriquecer con datos básicos del ítem si no vienen completos
    enriched = []
    for o in orders:
        order_items = o.get("order_items", [])
        row = {
            "order_id":      o.get("id"),
            "date_created":  o.get("date_created"),
            "date_closed":   o.get("date_closed"),
            "status":        o.get("status"),
            "total_amount":  o.get("total_amount"),
            "paid_amount":   o.get("paid_amount"),
            "currency_id":   o.get("currency_id"),
            "buyer_id":      o.get("buyer", {}).get("id"),
            "shipping_id":   o.get("shipping", {}).get("id"),
            "items": [
                {
                    "item_id":    i.get("item", {}).get("id"),
                    "title":      i.get("item", {}).get("title"),
                    "category":   i.get("item", {}).get("category_id"),
                    "quantity":   i.get("quantity"),
                    "unit_price": i.get("unit_price"),
                    "sale_fee":   i.get("sale_fee"),
                    "listing_type": i.get("item", {}).get("listing_type_id"),
                }
                for i in order_items
            ],
        }
        enriched.append(row)

    _save_json("orders", enriched)
    return enriched


# ─────────────────────────────────────────────────────────────────────────────
# 2. STOCK — FLEX y FULL
# ─────────────────────────────────────────────────────────────────────────────

def fetch_items(session: MLSession) -> list:
    """Trae todos los ítems activos y pausados del vendedor con datos de stock y logística."""
    print("\n  🗂️  Fetching ítems del catálogo...")

    # 1. Listar IDs de ítems activos + pausados (pausados también venden)
    all_ids = []
    for status in ["active", "paused"]:
        raw = session.get_paginated(
            f"/users/{session.user_id}/items/search",
            params={"status": status},
            limit=100
        )
        if raw and isinstance(raw[0], str):
            all_ids.extend(raw)
        else:
            all_ids.extend([i.get("id") for i in raw if i.get("id")])
    item_ids = list(dict.fromkeys(all_ids))  # dedup preservando orden

    print(f"  → {len(item_ids)} ítems encontrados (activos + pausados)")

    # 2. Traer detalles en batches de 20
    details = []
    for i in range(0, len(item_ids), 20):
        batch = item_ids[i:i+20]
        ids_str = ",".join(batch)
        try:
            resp = session.get(f"/items", params={"ids": ids_str})
            # Puede ser lista de {code, body} o lista directa
            for entry in (resp if isinstance(resp, list) else [resp]):
                body = entry.get("body", entry)
                if body:
                    details.append(body)
        except Exception as e:
            print(f"    ⚠️  Error en batch {i}: {e}")

    # 3. Extraer datos relevantes
    items = []
    for d in details:
        items.append({
            "item_id":          d.get("id"),
            "title":            d.get("title"),
            "category_id":      d.get("category_id"),
            "price":            d.get("price"),
            "base_price":       d.get("base_price"),
            "status":           d.get("status"),
            "listing_type_id":  d.get("listing_type_id"),
            "available_quantity": d.get("available_quantity"),
            "sold_quantity":    d.get("sold_quantity"),
            "shipping_mode":    d.get("shipping", {}).get("mode"),         # "me2" = Flex, "fulfillment" = Full
            "free_shipping":    d.get("shipping", {}).get("free_shipping"),
            "logistic_type":    d.get("shipping", {}).get("logistic_type"),
            "health":           d.get("health"),
            "permalink":        d.get("permalink"),
            "thumbnail":        d.get("thumbnail"),
            "date_created":     d.get("date_created"),
        })

    _save_json("items", items)
    return items


def fetch_fulfillment_stock(session: MLSession) -> list:
    """
    Trae el stock en centros de distribución de Mercado Envíos Full.
    Endpoint: /user-inventories (o inventory por seller)
    """
    print("\n  🏭  Fetching stock Fulfillment (Full)...")
    try:
        # Obtener inventories del seller
        inventories = session.get(f"/users/{session.user_id}/inventories")
        inventory_id = None

        if isinstance(inventories, dict):
            inv_list = inventories.get("inventories", [inventories])
        else:
            inv_list = inventories

        if inv_list:
            inventory_id = inv_list[0].get("id") if isinstance(inv_list[0], dict) else None

        if not inventory_id:
            print("  ℹ️  No se encontró inventory ID (puede que no uses Full)")
            _save_json("fulfillment_stock", [])
            return []

        # Stock por fulfillment
        stock_data = session.get(
            f"/inventories/{inventory_id}/stock/fulfillment",
            params={"limit": 100}
        )
        stock_list = stock_data.get("results", []) if isinstance(stock_data, dict) else []
        _save_json("fulfillment_stock", stock_list)
        return stock_list

    except Exception as e:
        print(f"  ⚠️  No se pudo obtener stock Full: {e}")
        _save_json("fulfillment_stock", [])
        return []


# ─────────────────────────────────────────────────────────────────────────────
# 3. COMISIONES Y COSTOS
# ─────────────────────────────────────────────────────────────────────────────

def fetch_billing(session: MLSession, days_back: int = 30) -> list:
    """
    Trae los movimientos de cuenta del período (comisiones, cargos, acreditaciones).
    """
    print(f"\n  💰 Fetching billing/movimientos (últimos {days_back} días)...")
    date_from, date_to = _date_range(days_back)

    try:
        movements = session.get_paginated(
            f"/users/{session.user_id}/account/movements",
            params={
                "date_from": date_from[:10],  # YYYY-MM-DD
                "date_to":   date_to[:10],
                "type":      "all",
            },
            limit=50
        )
        _save_json("billing", movements)
        return movements
    except Exception as e:
        print(f"  ⚠️  Error en billing: {e}")
        # Fallback: sacar comisiones de las órdenes ya descargadas
        _save_json("billing", [])
        return []


def compute_commissions_from_orders(orders: list) -> list:
    """
    Calcula comisiones y sale_fees directamente desde las órdenes.
    Útil como fallback si el endpoint de billing no está disponible.
    """
    commissions = []
    for o in orders:
        for item in o.get("items", []):
            commissions.append({
                "order_id":    o.get("order_id"),
                "date":        o.get("date_created"),
                "item_id":     item.get("item_id"),
                "title":       item.get("title"),
                "quantity":    item.get("quantity"),
                "unit_price":  item.get("unit_price"),
                "sale_fee":    item.get("sale_fee"),
                "total_sale":  (item.get("unit_price") or 0) * (item.get("quantity") or 0),
                "listing_type": item.get("listing_type"),
            })
    _save_json("commissions", commissions)
    return commissions


# ─────────────────────────────────────────────────────────────────────────────
# 4. RECLAMOS Y DEVOLUCIONES
# ─────────────────────────────────────────────────────────────────────────────

def fetch_claims(session: MLSession, days_back: int = 60) -> list:
    """Trae reclamos del período (abiertos y cerrados)."""
    print(f"\n  ⚠️  Fetching reclamos (últimos {days_back} días)...")
    date_from, _ = _date_range(days_back)

    claims = []
    for status in ["opened", "closed"]:
        try:
            batch = session.get_paginated(
                "/claims/search",
                params={
                    "role":         "respondent",
                    "status":       status,
                    "date_created_from": date_from[:10],
                },
                limit=50
            )
            claims.extend(batch)
        except Exception as e:
            print(f"  ⚠️  Error fetching claims ({status}): {e}")

    _save_json("claims", claims)
    return claims


def fetch_returns(session: MLSession, days_back: int = 60) -> list:
    """Trae devoluciones/mediaciones del período."""
    print(f"\n  🔄 Fetching devoluciones (últimos {days_back} días)...")
    date_from, _ = _date_range(days_back)

    try:
        returns = session.get_paginated(
            "/mediations/search",
            params={
                "role":        "respondent",
                "date_from":   date_from[:10],
            },
            limit=50
        )
        _save_json("returns", returns)
        return returns
    except Exception as e:
        print(f"  ⚠️  Error fetching devoluciones: {e}")
        _save_json("returns", [])
        return []


def fetch_shipments_detail(session: MLSession, orders: list) -> list:
    """
    Trae detalle logístico de los envíos de las órdenes.
    Solo procesa órdenes con shipping_id.
    """
    print("\n  🚚 Fetching detalle de envíos...")
    shipping_ids = list(set(
        o["shipping_id"] for o in orders
        if o.get("shipping_id")
    ))

    shipments = []
    for sid in shipping_ids[:200]:  # limitar a 200 para no tardar mucho
        try:
            s = session.get(f"/shipments/{sid}")
            shipments.append({
                "shipment_id":    s.get("id"),
                "order_id":       s.get("order_id"),
                "status":         s.get("status"),
                "substatus":      s.get("substatus"),
                "date_created":   s.get("date_created"),
                "date_delivered": s.get("date_delivered") or s.get("date_first_printed"),
                "logistic_type":  s.get("logistic_type"),
                "mode":           s.get("mode"),
                "cost_amount":    s.get("shipping_items", [{}])[0].get("cost") if s.get("shipping_items") else None,
                "receiver_city":  s.get("receiver_address", {}).get("city", {}).get("name"),
                "receiver_state": s.get("receiver_address", {}).get("state", {}).get("name"),
            })
        except Exception:
            pass

    _save_json("shipments", shipments)
    return shipments


# ─────────────────────────────────────────────────────────────────────────────
# Orquestador de datos
# ─────────────────────────────────────────────────────────────────────────────

def fetch_all(days_back: int = 30) -> dict:
    """
    Corre todos los fetchers y devuelve un dict con todos los datos.
    """
    print("\n" + "="*60)
    print(f"  ACTUALIZANDO DATOS (últimos {days_back} días)")
    print("="*60)

    session = MLSession()

    orders           = fetch_orders(session, days_back)
    items            = fetch_items(session)
    fulfillment      = fetch_fulfillment_stock(session)
    billing          = fetch_billing(session, days_back)
    commissions      = compute_commissions_from_orders(orders)
    claims           = fetch_claims(session, days_back * 2)
    returns          = fetch_returns(session, days_back * 2)
    shipments        = fetch_shipments_detail(session, orders)

    summary = {
        "updated_at":   datetime.now().isoformat(),
        "days_back":    days_back,
        "orders":       len(orders),
        "items":        len(items),
        "fulfillment":  len(fulfillment),
        "billing":      len(billing),
        "commissions":  len(commissions),
        "claims":       len(claims),
        "returns":      len(returns),
        "shipments":    len(shipments),
    }
    _save_json("summary", summary)

    print(f"\n  ✅ Datos actualizados. Resumen:")
    for k, v in summary.items():
        if k not in ("updated_at", "days_back"):
            print(f"     {k}: {v}")
    print()

    return {
        "orders":      orders,
        "items":       items,
        "fulfillment": fulfillment,
        "billing":     billing,
        "commissions": commissions,
        "claims":      claims,
        "returns":     returns,
        "shipments":   shipments,
        "summary":     summary,
    }


# ─────────────────────────────────────────────────────────────────────────────
# NUEVAS FUNCIONES PARA PIPELINE YoY
# ─────────────────────────────────────────────────────────────────────────────

def fetch_orders_range(session, date_from: str, date_to: str, label: str = "") -> list:
    """Descarga órdenes para un rango de fechas específico."""
    print(f"\n  📦 Fetching órdenes {label}...")
    results, offset, limit = [], 0, 50
    while True:
        try:
            resp = session.get("/orders/search", params={
                "seller": session.user_id,
                "sort":   "date_desc",
                "order.date_created.from": date_from,
                "order.date_created.to":   date_to,
                "limit":  limit,
                "offset": offset,
            })
        except Exception as e:
            print(f"    ⚠️  Error en página {offset}: {e}")
            break

        raw_items = resp.get("results", [])
        for o in raw_items:
            order_items = o.get("order_items", o.get("items", []))
            payments = o.get("payments", []) or []
            p0 = payments[0] if payments else {}
            results.append({
                "order_id":     o.get("id"),
                "date_created": o.get("date_created"),
                "date_closed":  o.get("date_closed"),
                "status":       o.get("status"),
                "total_amount": o.get("total_amount"),
                "paid_amount":  o.get("paid_amount"),
                "installments": int(p0.get("installments") or 1),
                "currency_id":  o.get("currency_id"),
                "buyer_id":     (o.get("buyer") or {}).get("id") if isinstance(o.get("buyer"), dict) else o.get("buyer_id"),
                "shipping_id":  (o.get("shipping") or {}).get("id") if isinstance(o.get("shipping"), dict) else o.get("shipping_id"),
                # Campos extra para órdenes canceladas (análisis postventa)
                "cancel_detail": o.get("cancel_detail"),
                "tags":          o.get("tags") or [],
                "mediations":    o.get("mediations") or [],
                "items": [
                    {
                        "item_id":    (i.get("item") or {}).get("id") if isinstance(i.get("item"), dict) else i.get("item_id"),
                        "title":      (i.get("item") or {}).get("title") if isinstance(i.get("item"), dict) else i.get("title"),
                        "category":   (i.get("item") or {}).get("category_id") if isinstance(i.get("item"), dict) else i.get("category"),
                        "quantity":   i.get("quantity"),
                        "unit_price": i.get("unit_price"),
                        "sale_fee":   i.get("sale_fee"),
                        "listing_type": (i.get("item") or {}).get("listing_type_id") if isinstance(i.get("item"), dict) else i.get("listing_type"),
                    }
                    for i in order_items
                ],
            })

        total  = resp.get("paging", {}).get("total", 0)
        offset += limit
        if offset >= total or not raw_items or offset >= 10000:
            break

    print(f"    → {len(results)} órdenes")
    return results


def compute_metrics(orders: list) -> dict:
    """Computa todas las métricas de negocio desde una lista de órdenes."""
    from collections import defaultdict
    from datetime import datetime

    paid = [o for o in orders if o.get("status") == "paid"]
    canc = [o for o in orders if o.get("status") == "cancelled"]

    gmv   = sum(o.get("total_amount", 0) or 0 for o in paid)
    units = sum(i.get("quantity", 0) or 0 for o in paid for i in o.get("items", []))
    fees  = sum(i.get("sale_fee", 0)  or 0 for o in paid for i in o.get("items", []))

    # Daily GMV + stats (por número de día del mes)
    daily       = defaultdict(float)
    daily_units = defaultdict(int)
    daily_paid  = defaultdict(int)
    daily_canc  = defaultdict(int)
    dow         = defaultdict(float)
    dow_dates   = defaultdict(set)   # FIXED: fechas únicas por DOW (no órdenes)
    inst_dist   = defaultdict(int)
    unique_day_set = set()

    for o in paid:
        ds   = o.get("date_created", "") or ""
        inst = int(o.get("installments") or 1)
        inst_dist[inst] += 1
        try:
            day = int(ds[:10].split("-")[2])
            amt = o.get("total_amount", 0) or 0
            daily[day]       += amt
            daily_paid[day]  += 1
            for i in o.get("items", []):
                daily_units[day] += i.get("quantity") or 0
            dt = datetime.fromisoformat(ds[:19])
            dow[dt.weekday()]        += amt
            dow_dates[dt.weekday()].add(ds[:10])  # unique dates per DOW
            unique_day_set.add(ds[:10])
        except Exception:
            pass

    for o in canc:
        ds = o.get("date_created", "") or ""
        try:
            day = int(ds[:10].split("-")[2])
            daily_canc[day] += 1
        except Exception:
            pass

    # Products
    prods = defaultdict(lambda: {"gmv": 0, "units": 0, "fees": 0, "title": ""})
    for o in paid:
        for i in o.get("items", []):
            iid = i.get("item_id") or "?"
            prods[iid]["gmv"]   += (i.get("unit_price") or 0) * (i.get("quantity") or 0)
            prods[iid]["units"] += i.get("quantity") or 0
            prods[iid]["fees"]  += i.get("sale_fee") or 0
            prods[iid]["title"] = prods[iid]["title"] or (i.get("title") or iid)[:55]

    # Cancellations by product
    cprods = defaultdict(lambda: {"c": 0, "p": 0, "title": ""})
    for o in canc:
        for i in o.get("items", []):
            iid = i.get("item_id") or "?"
            cprods[iid]["c"] += 1
            cprods[iid]["title"] = cprods[iid]["title"] or (i.get("title") or iid)[:55]
    for o in paid:
        for i in o.get("items", []):
            iid = i.get("item_id") or "?"
            if iid in cprods:
                cprods[iid]["p"] += 1

    total_orders = len(paid) + len(canc)
    unique_days  = len(unique_day_set) or 1

    # Build daily_stats for JS period filter
    all_days = sorted(set(list(daily.keys()) + list(daily_canc.keys())))
    daily_stats = {}
    for d in all_days:
        daily_stats[str(d)] = {
            "gmv":  round(daily.get(d, 0), 2),
            "units": daily_units.get(d, 0),
            "paid":  daily_paid.get(d, 0),
            "canc":  daily_canc.get(d, 0),
        }

    return {
        "gmv":         round(gmv, 2),
        "units":       units,
        "fees":        round(fees, 2),
        "fee_rate":    round(fees / gmv * 100, 2) if gmv else 0,
        "net":         round(gmv - fees, 2),
        "paid":        len(paid),
        "cancelled":   len(canc),
        "cancel_rate": round(len(canc) / total_orders * 100, 2) if total_orders else 0,
        "avg_ticket":  round(gmv / len(paid), 2) if paid else 0,
        "unique_days": unique_days,
        "daily":       {str(k): round(v, 2) for k, v in sorted(daily.items())},
        "daily_stats": daily_stats,
        "dow":         [round(dow.get(i, 0), 2) for i in range(7)],
        "dow_count":   [len(dow_dates.get(i, set())) for i in range(7)],  # FIXED: días únicos
        "installment_dist": {str(k): v for k, v in sorted(inst_dist.items())},
        "net_real":    round(gmv - fees, 2),
        "top": sorted(
            [{"id": k, "gmv": round(v["gmv"]), "units": v["units"],
              "fees": round(v["fees"]), "title": v["title"]}
             for k, v in prods.items()],
            key=lambda x: -x["gmv"]
        )[:25],
        "cancels": sorted(
            [{"id": k, "c": v["c"], "p": v["p"], "title": v["title"],
              "rate": round(v["c"] / (v["c"] + v["p"]) * 100, 1) if (v["c"] + v["p"]) else 0}
             for k, v in cprods.items()],
            key=lambda x: -x["c"]
        )[:20],
    }


def enrich_metrics(orders: list, items: list, ship_map: dict = None) -> dict:
    """Enriquece métricas con dimensiones adicionales: categoría, tipo publicación, logística, hora."""
    from collections import defaultdict
    from datetime import datetime

    ship_map  = ship_map or {}
    item_map = {i["item_id"]: i for i in items}
    paid = [o for o in orders if o.get("status") == "paid"]

    by_cat  = defaultdict(lambda: {"gmv": 0, "units": 0, "name": ""})
    by_lt   = defaultdict(lambda: {"gmv": 0, "units": 0})
    by_log     = defaultdict(lambda: {"gmv": 0, "units": 0})
    by_hour    = defaultdict(float)
    hm         = [[0.0] * 24 for _ in range(7)]
    du         = defaultdict(int)
    # Per-day breakdown for listing_type and logistic (enables range filtering in dashboard)
    daily_lt   = defaultdict(lambda: defaultdict(float))   # {day: {label: gmv}}
    daily_log  = defaultdict(lambda: defaultdict(float))   # {day: {label: gmv}}
    daily_lt_u = defaultdict(lambda: defaultdict(int))     # {day: {label: units}}
    daily_log_u= defaultdict(lambda: defaultdict(int))     # {day: {label: units}}

    for o in paid:
        ds = o.get("date_created", "") or ""
        try:
            dt  = datetime.fromisoformat(ds[:19])
            h   = dt.hour
            dw  = dt.weekday()
            amt = o.get("total_amount", 0) or 0
            by_hour[h]   += amt
            hm[dw][h]    += amt
        except Exception:
            h, dw = 0, 0

        day = 0
        try:
            day = int(ds[:10].split("-")[2])
        except Exception:
            pass

        for i in o.get("items", []):
            iid   = i.get("item_id") or "?"
            gmv   = (i.get("unit_price") or 0) * (i.get("quantity") or 0)
            units = i.get("quantity") or 0
            cat     = i.get("category") or item_map.get(iid, {}).get("category_id", "?")
            lt_raw  = item_map.get(iid, {}).get("listing_type_id") or i.get("listing_type") or "unknown"
            lt      = {"gold_special": "Clásica", "gold_pro": "Premium",
                       "gold_premium": "Premium"}.get(lt_raw, "Sin clasificar" if lt_raw == "unknown" else lt_raw)
            # Prioridad: ship_map (API /shipments) > shipping inline > item_map
            _sid    = str(o.get("shipping_id", ""))
            log_raw = (ship_map.get(_sid)
                       or o.get("shipping", {}).get("logistic_type")
                       or item_map.get(iid, {}).get("logistic_type")
                       or "unknown")
            log     = {"fulfillment": "Full",
                       "self_service": "Flex",        # vendedor gestiona su propia logística
                       "cross_docking": "Colecta",    # vendedor lleva a ML, ML entrega (ME)
                       "default_buying_flow": "Colecta",
                       "drop_off": "Colecta",
                       "xd_drop_off": "Colecta"}.get(log_raw, None if log_raw == "unknown" else log_raw)
            if log is None: continue  # sin logistic_type confirmado → excluir del breakdown

            by_cat[cat]["gmv"]   += gmv
            by_cat[cat]["units"] += units
            by_lt[lt]["gmv"]     += gmv
            by_lt[lt]["units"]   += units
            by_log[log]["gmv"]   += gmv
            by_log[log]["units"] += units
            du[day]              += units
            if day:
                daily_lt[day][lt]   += gmv
                daily_log[day][log] += gmv
                daily_lt_u[day][lt]   += units
                daily_log_u[day][log] += units

    # Serialize daily_lt and daily_log as {day_str: {label: gmv_int}}
    daily_lt_out  = {str(d): {lbl: round(v) for lbl, v in lmap.items()}
                     for d, lmap in sorted(daily_lt.items())}
    daily_log_out = {str(d): {lbl: round(v) for lbl, v in lmap.items()}
                     for d, lmap in sorted(daily_log.items())}
    daily_lt_u_out  = {str(d): dict(lmap) for d, lmap in sorted(daily_lt_u.items())}
    daily_log_u_out = {str(d): dict(lmap) for d, lmap in sorted(daily_log_u.items())}

    return {
        "by_category":     {k: {"gmv": round(v["gmv"]), "units": v["units"], "name": v["name"]}
                            for k, v in by_cat.items()},
        "by_listing_type": {k: {"gmv": round(v["gmv"]), "units": v["units"]}
                            for k, v in by_lt.items()},
        "by_logistic":     {k: {"gmv": round(v["gmv"]), "units": v["units"]}
                            for k, v in by_log.items()},
        "by_hour":         {str(k): round(v) for k, v in sorted(by_hour.items())},
        "heatmap":         [[round(hm[d][h]) for h in range(24)] for d in range(7)],
        "daily_units":     {str(k): v for k, v in sorted(du.items())},
        "daily_lt":        daily_lt_out,
        "daily_log":       daily_log_out,
        "daily_lt_u":      daily_lt_u_out,
        "daily_log_u":     daily_log_u_out,
    }


def fetch_category_names(session, category_ids: list) -> dict:
    """Obtiene nombres legibles de categorías ML."""
    names = {}
    for cid in category_ids:
        if not cid or cid == "?":
            continue
        try:
            r = session.get(f"/categories/{cid}")
            names[cid] = r.get("name", cid)
        except Exception:
            names[cid] = cid
    print(f"  🏷️  Nombres de categorías: {len(names)} obtenidos")
    return names


def fetch_monthly_evolution_2026(session, existing: list = None) -> list:
    """
    Descarga métricas mensuales para todos los meses del año corriente.
    Cachea meses ya cerrados para no re-fetchear.
    """
    import calendar
    from datetime import date
    today = date.today()
    year  = today.year

    existing_months = {}
    if existing:
        for m in existing:
            if m.get("complete"):
                existing_months[m["month"]] = m

    MONTH_ES = ["","Ene","Feb","Mar","Abr","May","Jun",
                "Jul","Ago","Sep","Oct","Nov","Dic"]

    result = []
    for m in range(1, today.month + 1):
        if m < today.month and m in existing_months:
            result.append(existing_months[m])
            print(f"  📅 {year}-{m:02d}: cacheado ({existing_months[m]['paid']} órdenes)")
            continue

        last_day  = calendar.monthrange(year, m)[1]
        end_day   = today.day if m == today.month else last_day
        date_from = f"{year}-{m:02d}-01T00:00:00.000-00:00"
        date_to   = f"{year}-{m:02d}-{end_day:02d}T23:59:59.000-00:00"

        orders  = fetch_orders_range(session, date_from, date_to, f"{year}-{m:02d}")
        metrics = compute_metrics(orders)

        result.append({
            "month":       m,
            "year":        year,
            "label":       MONTH_ES[m],
            "complete":    m < today.month,
            "gmv":         metrics["gmv"],
            "units":       metrics["units"],
            "paid":        metrics["paid"],
            "cancelled":   metrics["cancelled"],
            "cancel_rate": metrics["cancel_rate"],
            "fees":        metrics["fees"],
            "avg_ticket":  metrics["avg_ticket"],
        })

    _save_json("monthly_2026", result)
    return result


def fetch_reputation(session) -> dict:
    """Trae datos de reputación del vendedor."""
    print("\n  ⭐ Fetching reputación...")
    try:
        user = session.get(f"/users/{session.user_id}")
        sr   = user.get("seller_reputation", {})
        metrics = sr.get("metrics", {})
        txn     = sr.get("transactions", {})
        result = {
            "level_id":           sr.get("level_id", ""),
            "power_seller_status": sr.get("power_seller_status", ""),
            "sales_60d":          metrics.get("sales", {}).get("completed", 0),
            "claims_rate":        metrics.get("claims", {}).get("rate", 0),
            "claims_value":       metrics.get("claims", {}).get("value", 0),
            "delayed_rate":       metrics.get("delayed_handling_time", {}).get("rate", 0),
            "delayed_value":      metrics.get("delayed_handling_time", {}).get("value", 0),
            "cancellations_rate": metrics.get("cancellations", {}).get("rate", 0),
            "cancellations_value": metrics.get("cancellations", {}).get("value", 0),
            "completed":          txn.get("completed", 0),
            "canceled":           txn.get("canceled", 0),
            "total":              txn.get("total", 0),
            "ratings_positive":   txn.get("ratings", {}).get("positive", 0),
            "ratings_negative":   txn.get("ratings", {}).get("negative", 0),
            "nickname":           user.get("nickname", ""),
            "fetched_at":         datetime.now().isoformat(),
        }
        _save_json("reputation", result)
        print(f"  ✅ {result['power_seller_status'].upper()} | Reclamos {result['claims_rate']*100:.1f}% | Cancelaciones {result['cancellations_rate']*100:.2f}%")
        return result
    except Exception as e:
        print(f"  ⚠️ Error reputación: {e}")
        return {}


def fetch_ads_data(session, advertiser_id: str = None) -> dict:
    """
    Intenta obtener datos de ads via API.
    Requiere scope advertising:read.
    """
    print("\n  📣 Fetching datos de Ads...")
    aid = advertiser_id or str(session.user_id)
    endpoints_to_try = [
        f"/advertising/product_ads/campaigns?advertiser_id={aid}",
        f"/advertising/product_ads/v2/campaigns?advertiser_id={aid}",
        f"/advertising/advertisers/{aid}/campaigns",
    ]
    for ep in endpoints_to_try:
        try:
            r = session.get(ep)
            if isinstance(r, (dict, list)):
                _save_json("ads_data", r if isinstance(r, dict) else {"campaigns": r})
                print(f"  ✅ Ads data: {ep}")
                return r if isinstance(r, dict) else {"campaigns": r}
        except Exception as e:
            if "403" in str(e):
                print(f"  ⚠️ Ads scope no habilitado (advertising:read). Activalo en ML Developers.")
                break
            continue
    _save_json("ads_data", {"status": "scope_required",
                            "message": "Activar advertising:read en ML Developers -> Re-autorizar"})
    return {}


def fetch_costs_summary(session) -> dict:
    print("\n  Fetching costos/billing...")
    try:
        for ep in [
            f"/billing/integration/v1/accounts/{session.user_id}/transactions",
            f"/users/{session.user_id}/account/movements",
        ]:
            try:
                r = session.get(ep, params={"limit": 5})
                _save_json("costs_data", r if isinstance(r, dict) else {"results": r})
                return r
            except Exception as e:
                if "403" in str(e):
                    break
                continue
    except Exception:
        pass
    _save_json("costs_data", {"status": "scope_required",
                              "message": "Activar billing:read en ML Developers -> Re-autorizar"})
    return {}


def process_status_xlsx(xlsx_path: str) -> list:
    """
    Lee STATUS.xlsx del vendedor (header fila 7):
      D=CODIGO, E=DESCRIPCION, F=CATEGORIA, I=SUB_CATEGORIA, J=MARCA
      N=STOCK GRAL (total), O=STOCK DIGITAL (deposito), P=STOCK FULLFILL, Q=ADUANA
      T=VTA ULTIMA SEMANA, U=VTA ULTIMO MES
    """
    try:
        import openpyxl, warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb['STATUS'] if 'STATUS' in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        result = []
        for row in rows[7:]:  # data starts at row 8 (index 7)
            if not row or all(v is None for v in row):
                continue
            def _v(col): return row[col] if col < len(row) else None
            def _num(col):
                v = _v(col)
                try: return int(float(v)) if v is not None else 0
                except: return 0
            codigo = str(_v(3) or "").strip()
            desc   = str(_v(4) or "").strip()
            if not codigo and not desc:
                continue
            result.append({
                "codigo":       codigo,
                "descripcion":  desc,
                "categoria":    str(_v(5) or "").strip(),
                "sub_cat":      str(_v(8) or "").strip(),
                "marca":        str(_v(9) or "").strip(),
                "dias_stock":   str(_v(6) or "").strip(),
                "activado":     str(_v(7) or "SI").strip(),
                "stock_total":  _num(13),
                "stock_dep":    _num(14),
                "stock_full":   _num(15),
                "stock_aduana": _num(16),
                "vta_semana":   _num(19),
                "vta_mes":      _num(20),
            })
        print(f"  STATUS.xlsx: {len(result)} SKUs procesados")
        _save_json("stock_status", result)
        return result
    except Exception as e:
        print(f"  No se pudo procesar STATUS.xlsx: {e}")
        return []


def load_sku_categories(xlsx_path: str) -> dict:
    """Carga archivo de categorias por SKU. Retorna {codigo: {categoria, sub_cat, descripcion}}"""
    try:
        import openpyxl, warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            codigo = str(row[0] or "").strip()
            if codigo:
                result[codigo] = {
                    "descripcion": str(row[1] or "").strip(),
                    "categoria":   str(row[2] or "").strip(),
                    "sub_cat":     str(row[3] or "").strip(),
                }
        print(f"  SKU categories: {len(result)} codigos cargados")
        return result
    except Exception as e:
        print(f"  No se pudo cargar categorias SKU: {e}")
        return {}


def compute_postventa(cur_raw: list, pri_raw: list = None) -> dict:
    """
    Genera postventa_current.json y cancelled_current.json desde órdenes crudas.
    Requiere que fetch_orders_range capture cancel_detail, tags, mediations.
    """
    from collections import defaultdict
    import json as _json
    import os as _os

    DATA_DIR = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'data')

    paid_orders = [o for o in cur_raw if o.get("status") == "paid"]
    canc_orders = [o for o in cur_raw if o.get("status") == "cancelled"]

    # --- Guardar cancelled_current.json ---
    _save_json("cancelled_current", canc_orders)

    # Cancel reason codes → categorías
    CODE_MAP = {
        'buyer_cancel_express':    'pre_envio',
        'pack_splitted':           'pack',
        'mediations':              'med',
        'shipment_not_delivered':  'log',
        'fraud':                   'fraud',
        'shipment_unfulfilled':    'log',
        'not_paid':                'pre_envio',
        'buyer_cancel_after_ship': 'pre_envio',
    }
    CODE_LABEL = {
        'pre_envio': 'Comprador canceló antes de despacho',
        'pack':      'División de pack (ML)',
        'med':       'Mediación / Reclamo',
        'log':       'No entregado por logística',
        'fraud':     'Fraude detectado',
        'other':     'Sin especificar',
    }

    total_paid  = len(paid_orders)
    total_canc  = len(canc_orders)

    # Count devolutions (returned) vs mediations per cancel_detail
    dev_count  = 0
    med_count  = 0
    gmv_dev    = 0.0
    gmv_med    = 0.0

    daily = defaultdict(lambda: {'dev':0,'med':0,'pre':0,'log':0,'int':0,'gmv_dev':0,'gmv_med':0})
    by_code_counts = defaultdict(int)

    # Per-SKU tracking
    sku_dev = defaultdict(lambda: {'id':'','title':'','dev':0,'med':0,'paid':0,'gmv_dev':0})
    sku_med = defaultdict(lambda: {'id':'','title':'','dev':0,'med':0,'paid':0,'gmv_med':0})

    for o in paid_orders:
        ds = o.get("date_created","")
        try:
            day = int(ds[8:10])
        except:
            day = None
        for i in o.get("items",[]):
            iid = i.get("item_id","")
            sku_dev[iid]['id']    = iid
            sku_dev[iid]['title'] = sku_dev[iid].get('title','') or (i.get("title","") or iid)[:55]
            sku_dev[iid]['paid'] += 1
            sku_med[iid]['id']    = iid
            sku_med[iid]['title'] = sku_med[iid].get('title','') or (i.get("title","") or iid)[:55]
            sku_med[iid]['paid'] += 1

    for o in canc_orders:
        cd   = o.get("cancel_detail") or {}
        code = cd.get("code","") or ""
        meds = o.get("mediations") or []
        tags = o.get("tags") or []
        gmv  = o.get("total_amount",0) or 0

        # Determine type
        has_med = bool(meds) or "mediation" in " ".join(tags).lower() or code in ("mediations","fraud")
        cat_key = CODE_MAP.get(code, 'other')

        ds = o.get("date_created","")
        try:
            day = int(ds[8:10])
        except:
            day = 0

        by_code_counts[CODE_LABEL.get(cat_key, 'Sin especificar')] += 1

        if has_med:
            med_count += 1
            gmv_med   += gmv
            daily[day]['med']     += 1
            daily[day]['gmv_med'] += gmv
            for i in o.get("items",[]):
                iid = i.get("item_id","")
                if iid:
                    sku_med[iid]['id']    = iid
                    sku_med[iid]['title'] = sku_med[iid].get('title','') or (i.get("title","") or iid)[:55]
                    sku_med[iid]['med']   += 1
                    sku_med[iid]['gmv_med'] += gmv
        else:
            dev_count += 1
            gmv_dev   += gmv
            daily[day]['dev']     += 1
            daily[day]['gmv_dev'] += gmv
            for i in o.get("items",[]):
                iid = i.get("item_id","")
                if iid:
                    sku_dev[iid]['id']    = iid
                    sku_dev[iid]['title'] = sku_dev[iid].get('title','') or (i.get("title","") or iid)[:55]
                    sku_dev[iid]['dev']   += 1
                    sku_dev[iid]['gmv_dev'] += gmv

        # sub-categories
        if cat_key == 'pre_envio': daily[day]['pre'] += 1
        elif cat_key == 'log':     daily[day]['log'] += 1
        elif cat_key == 'med':     daily[day]['int'] += 1

    base = total_paid + total_canc
    rate_dev = round(dev_count / base * 100, 2) if base else 0
    rate_med = round(med_count / base * 100, 2) if base else 0

    # Top 15 by dev / med
    top_dev = sorted(
        [dict(v) for v in sku_dev.values() if v['dev'] > 0],
        key=lambda x: -x['dev']
    )[:15]
    top_med = sorted(
        [dict(v) for v in sku_med.values() if v['med'] > 0],
        key=lambda x: -x['med']
    )[:15]

    result = {
        'paid':    total_paid,
        'dev':     dev_count,
        'med':     med_count,
        'gmv_dev': round(gmv_dev),
        'gmv_med': round(gmv_med),
        'rate_dev': rate_dev,
        'rate_med': rate_med,
        'daily':   {str(k): {kk: round(vv) for kk,vv in v.items()} for k,v in sorted(daily.items()) if k > 0},
        'by_code': dict(by_code_counts),
        'top_dev': top_dev,
        'top_med': top_med,
    }

    _save_json("postventa_current", result)
    print(f"  ✅ postventa_current: paid={total_paid}, dev={dev_count}, med={med_count}, cancelled_saved={len(canc_orders)}")
    return result
